import re
from django.contrib import admin
from django.contrib.admin import ModelAdmin
from django.contrib.admin.filters import SimpleListFilter
from django.contrib.admin.views.main import ChangeList
from django.contrib.admin.templatetags.admin_urls import admin_urlname
from django.contrib.admin.utils import unquote
from django.core.exceptions import PermissionDenied
from django.db import transaction
from .models import (
    Cliente, Agente, Ventas, Anticipo, TerminoCredito, MercadoDestino, PagoVenta,
    SaldoCliente, AntigüedadSaldo, EstadoCuentaCliente, ConfiguracionCuentasPorCobrar,
    ObligacionFiscal, DocumentoCFDI
)
from .forms import (
    VentasAdminForm,
    CFDIUploadForm,
    CFDIConfirmForm,
    AnticipoCFDIUploadForm,
    AnticipoCFDIConfirmForm,
)
from catalogo.models import Sucursal, Pais, Producto
from gastos.models import Cuenta
from import_export import resources
from import_export.admin import ImportExportModelAdmin
from import_export import resources, fields
from import_export.widgets import ForeignKeyWidget
from import_export.forms import ExportForm, ImportForm
from app.widgets import MoneyWidget
from django.utils.html import format_html
from app.media_utils import safe_file_url
from django.http import (
    FileResponse, Http404, HttpResponse, JsonResponse,
)
from django.shortcuts import render, redirect
from django.urls import path, reverse
from django.utils.safestring import mark_safe
from django.db import models
from django.db.models import Sum, Count, Avg, Q, F, Case, When, Value
from django.db.models.functions import Extract, TruncMonth, TruncDay, Coalesce
from django.utils import timezone
from django.utils.translation import gettext_lazy as _
from datetime import datetime, timedelta, date
from decimal import Decimal
from djmoney.money import Money
import json
from django.contrib import messages
from django.template.response import TemplateResponse
from .services.metrics_service import CuentasPorCobrarMetrics
from .services.cache_service import CuentasPorCobrarCache
from django.utils.html import format_html


CFDI_UPLOAD_LIMITS = {
    'xml': (1 * 1024 * 1024, '1 MB'),
    'pdf': (1 * 1024 * 1024, '1 MB'),
    'zip': (5 * 1024 * 1024, '5 MB'),
}


def _validar_archivo_cfdi(archivo):
    """Devuelve un mensaje de error si el archivo no cumple tipo o tamaño."""
    extension = (
        archivo.name.lower().rsplit('.', 1)[-1]
        if '.' in archivo.name else ''
    )
    limite = CFDI_UPLOAD_LIMITS.get(extension)
    if limite is None:
        return f'"{archivo.name}" no es .xml, .zip ni .pdf.'

    limite_bytes, limite_etiqueta = limite
    if archivo.size > limite_bytes:
        return f'"{archivo.name}" supera {limite_etiqueta}.'
    return None


def _extraer_archivos(archivos):
    """
    Extrae una lista de (nombre, bytes, ext) a partir de archivos subidos,
    soportando .xml y .pdf individuales y .zip que contenga ambos tipos.
    """
    import io
    import zipfile

    extraidos = []
    for f in archivos:
        nombre = f.name.lower()
        if nombre.endswith('.zip'):
            with zipfile.ZipFile(io.BytesIO(f.read())) as zf:
                for info in zf.infolist():
                    if info.is_dir():
                        continue
                    ext = info.filename.lower().rsplit('.', 1)[-1] if '.' in info.filename else ''
                    if ext in ('xml', 'pdf'):
                        extraidos.append((info.filename, zf.read(info), ext))
        else:
            ext = nombre.rsplit('.', 1)[-1] if '.' in nombre else ''
            if ext in ('xml', 'pdf'):
                extraidos.append((f.name, f.read(), ext))
    return extraidos


def _pdf_response(file_field):
    """Serve an authorized PDF through Django from local storage or R2."""
    try:
        pdf_file = file_field.open('rb')
    except (FileNotFoundError, OSError):
        raise Http404(_('No fue posible encontrar el archivo PDF.'))

    filename = file_field.name.rsplit('/', 1)[-1]
    response = FileResponse(
        pdf_file,
        content_type='application/pdf',
        as_attachment=False,
        filename=filename,
    )

    # XFrameOptionsMiddleware no reemplaza cabeceras ya definidas. La vista se
    # puede incrustar únicamente dentro del mismo sistema administrativo.
    response['X-Frame-Options'] = 'SAMEORIGIN'
    response['Content-Security-Policy'] = "frame-ancestors 'self'"
    response['Cache-Control'] = 'private, no-store'
    return response


def _pdf_preview_button(pdf_url, title, identifier):
    aria_label = _('Visualizar PDF %(identifier)s') % {
        'identifier': identifier,
    }
    return format_html(
        '<button type="button" class="cfdi-pdf-trigger js-cfdi-pdf-preview" '
        'data-pdf-url="{}" data-pdf-title="{}" aria-label="{}">'
        '<i class="fas fa-file-pdf" aria-hidden="true"></i>'
        '<span>{}</span></button>',
        pdf_url,
        title,
        aria_label,
        _('Ver PDF'),
    )


_UUID_RE = re.compile(
    r'[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}'
)


# =============================================================================
# FILTROS PERSONALIZADOS AVANZADOS
# =============================================================================

class RangoCreditoFilter(SimpleListFilter):
    """Filtro por rango de crédito disponible"""
    title = 'Rango de Crédito Disponible'
    parameter_name = 'rango_credito'
    
    def lookups(self, request, model_admin):
        return [
            ('0-1000', '$0 - $1,000'),
            ('1000-5000', '$1,000 - $5,000'),
            ('5000-10000', '$5,000 - $10,000'),
            ('10000-50000', '$10,000 - $50,000'),
            ('50000+', '$50,000+'),
            ('sin_credito', 'Sin Crédito Disponible'),
        ]
    
    def queryset(self, request, queryset):
        if self.value() == 'sin_credito':
            return queryset.filter(
                Q(tipo_cliente='Contado') | Q(limite_credito__amount=0)
            )
        elif self.value() == '0-1000':
            return queryset.filter(
                tipo_cliente__in=['Credito', 'Mixto'],
                limite_credito__amount__range=[0, 1000]
            )
        elif self.value() == '1000-5000':
            return queryset.filter(
                tipo_cliente__in=['Credito', 'Mixto'],
                limite_credito__amount__range=[1000, 5000]
            )
        elif self.value() == '5000-10000':
            return queryset.filter(
                tipo_cliente__in=['Credito', 'Mixto'],
                limite_credito__amount__range=[5000, 10000]
            )
        elif self.value() == '10000-50000':
            return queryset.filter(
                tipo_cliente__in=['Credito', 'Mixto'],
                limite_credito__amount__range=[10000, 50000]
            )
        elif self.value() == '50000+':
            return queryset.filter(
                tipo_cliente__in=['Credito', 'Mixto'],
                limite_credito__amount__gte=50000
            )
        return queryset

class VencimientoFilter(SimpleListFilter):
    """Filtro por estado de vencimiento de ventas"""
    title = 'Estado de Vencimiento'
    parameter_name = 'vencimiento'
    
    def lookups(self, request, model_admin):
        return [
            ('vence_hoy', 'Vence Hoy'),
            ('vence_semana', 'Vence Esta Semana'),
            ('vence_mes', 'Vence Este Mes'),
            ('vencido', 'Vencido'),
            ('vencido_30', 'Vencido +30 días'),
            ('vencido_60', 'Vencido +60 días'),
            ('vencido_90', 'Vencido +90 días'),
        ]
    
    def queryset(self, request, queryset):
        hoy = timezone.now().date()
        
        if self.value() == 'vence_hoy':
            return queryset.filter(
                modalidad_pago='Credito',
                fecha_vencimiento=hoy
            )
        elif self.value() == 'vence_semana':
            fin_semana = hoy + timedelta(days=7)
            return queryset.filter(
                modalidad_pago='Credito',
                fecha_vencimiento__range=[hoy, fin_semana]
            )
        elif self.value() == 'vence_mes':
            fin_mes = hoy + timedelta(days=30)
            return queryset.filter(
                modalidad_pago='Credito',
                fecha_vencimiento__range=[hoy, fin_mes]
            )
        elif self.value() == 'vencido':
            return queryset.filter(
                modalidad_pago='Credito',
                fecha_vencimiento__lt=hoy,
                estado_cobranza__in=['Pendiente', 'Parcial']
            )
        elif self.value() == 'vencido_30':
            hace_30 = hoy - timedelta(days=30)
            return queryset.filter(
                modalidad_pago='Credito',
                fecha_vencimiento__lt=hace_30,
                estado_cobranza__in=['Pendiente', 'Parcial']
            )
        elif self.value() == 'vencido_60':
            hace_60 = hoy - timedelta(days=60)
            return queryset.filter(
                modalidad_pago='Credito',
                fecha_vencimiento__lt=hace_60,
                estado_cobranza__in=['Pendiente', 'Parcial']
            )
        elif self.value() == 'vencido_90':
            hace_90 = hoy - timedelta(days=90)
            return queryset.filter(
                modalidad_pago='Credito',
                fecha_vencimiento__lt=hace_90,
                estado_cobranza__in=['Pendiente', 'Parcial']
            )
        return queryset

class MontoVentaFilter(SimpleListFilter):
    """Filtro por rango de monto de venta"""
    title = 'Rango de Monto'
    parameter_name = 'rango_monto'
    
    def lookups(self, request, model_admin):
        return [
            ('0-1000', '$0 - $1,000'),
            ('1000-5000', '$1,000 - $5,000'),
            ('5000-10000', '$5,000 - $10,000'),
            ('10000-25000', '$10,000 - $25,000'),
            ('25000-50000', '$25,000 - $50,000'),
            ('50000+', '$50,000+'),
        ]
    
    def queryset(self, request, queryset):
        if self.value() == '0-1000':
            return queryset.filter(monto__amount__range=[0, 1000])
        elif self.value() == '1000-5000':
            return queryset.filter(monto__amount__range=[1000, 5000])
        elif self.value() == '5000-10000':
            return queryset.filter(monto__amount__range=[5000, 10000])
        elif self.value() == '10000-25000':
            return queryset.filter(monto__amount__range=[10000, 25000])
        elif self.value() == '25000-50000':
            return queryset.filter(monto__amount__range=[25000, 50000])
        elif self.value() == '50000+':
            return queryset.filter(monto__amount__gte=50000)
        return queryset


class ClienteResource(resources.ModelResource):
    pais = fields.Field(    
        column_name='pais',
        attribute='pais',
        widget=ForeignKeyWidget(Pais, field='nombre'))
    
    class Meta:
        model = Cliente
        fields = (
            'id', 'nombre', 'rfc', 'residencia_fiscal',
            'numero_registro_fiscal', 'codigo_postal_fiscal',
            'regimen_fiscal', 'telefono', 'correo', 'direccion', 'pais',
            'fecha_registro',
        )
    
    def dehydrate_pais(self, cliente):
        return cliente.pais.nombre
    
    def before_import_row(self, row, **kwargs):
        # Asigna un ID específico basado en un rango disponible
        if not row['id']:
            last_cliente = Cliente.objects.order_by('-id').first()
            next_id = last_cliente.id + 1 if last_cliente else 1
            row['id'] = next_id
        

@admin.register(Cliente)
class ClienteAdmin(ImportExportModelAdmin, ModelAdmin):
    resource_class = ClienteResource
    import_form_class = ImportForm
    export_form_class = ExportForm
    
    list_display = (
        'nombre', 'rfc', 'get_pais', 'tipo_cliente', 'limite_credito',
        'calificacion_credito', 'get_credito_disponible', 'get_saldo_conciliado', 'activo'
    )
    
    list_filter = ('tipo_cliente', 'calificacion_credito', 'mercado_destino', 'activo', 'pais', RangoCreditoFilter)
    search_fields = ('nombre', 'rfc', 'numero_registro_fiscal', 'correo')
    list_per_page = 20
    
    fieldsets = (
        ('Información Básica', {
            'fields': ('nombre', 'telefono', 'correo', 'direccion', 'imagen')
        }),
        ('Información Fiscal', {
            'fields': (
                'rfc', 'residencia_fiscal', 'numero_registro_fiscal',
                'codigo_postal_fiscal', 'regimen_fiscal',
            )
        }),
        ('Ubicación y Mercado', {
            'fields': ('pais', 'mercado_destino')
        }),
        ('Configuración de Crédito', {
            'fields': ('tipo_cliente', 'limite_credito', 'termino_credito_predeterminado', 
                      'calificacion_credito')
        }),
        ('Estado', {
            'fields': ('activo', 'fecha_registro'),
        }),
    )
    
    readonly_fields = ('fecha_registro',)

    def get_pais(self, obj):
        return obj.pais.nombre
    get_pais.short_description = 'País'

    def get_bandera(self, obj):
        return obj.pais.mostrar_bandera()
    get_bandera.short_description = 'Bandera'
    
    def get_credito_disponible(self, obj):
        if obj.tipo_cliente == 'Contado':
            return 'N/A'
        disponible = obj.credito_disponible()
        color = 'green' if disponible > 0 else 'red'
        return format_html(
            '<span style="color: {}">${}</span>',
            color, f"{float(disponible):,.2f}"
        )
    get_credito_disponible.short_description = 'Crédito Disponible'

    def get_saldo_conciliado(self, obj):
        saldo = obj.saldo_conciliado()
        color = 'red' if saldo > 0 else 'green'
        return format_html(
            '<span style="color: {}">${}</span>',
            color, f"{saldo:,.2f}"
        )
    get_saldo_conciliado.short_description = 'Saldo CFDI (conciliado)'

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                '<client_id>/reporte-completo/',
                self.admin_site.admin_view(self.reporte_cliente_completo),
                name='%s_%s_reporte_completo' % (self.model._meta.app_label, self.model._meta.model_name),
            ),
            path(
                'conciliacion/',
                self.admin_site.admin_view(self.conciliacion_view),
                name='%s_%s_conciliacion' % (self.model._meta.app_label, self.model._meta.model_name),
            ),
        ]
        return custom_urls + urls

    def conciliacion_view(self, request):
        """Vista de conciliación de CFDI por cliente."""
        from .services.conciliacion_service import conciliacion_global
        filas = conciliacion_global()
        context = dict(
            self.admin_site.each_context(request),
            filas=filas,
            title='Conciliación de CFDI por cliente',
            opts=self.model._meta,
        )
        return TemplateResponse(request, 'admin/ventas/conciliacion.html', context)
    
    def reporte_cliente_completo(self, request, client_id):
        """Genera un reporte completo del cliente con todas sus transacciones."""
        cliente = self.get_object(request, unquote(client_id))
        if cliente is None:
            messages.error(request, 'Cliente no encontrado')
            return redirect('admin:ventas_cliente_changelist')
            
        # Obtener métricas del cliente
        ventas_totales = cliente.ventas_set.aggregate(
            total=Sum('monto'),
            count=Count('id'),
            promedio=Avg('monto')
        )
        
        ventas_por_estado = cliente.ventas_set.values('estado_cobranza').annotate(
            total=Sum('monto'),
            count=Count('id')
        ).order_by('estado_cobranza')
        
        context = dict(
            self.admin_site.each_context(request),
            cliente=cliente,
            ventas_totales=ventas_totales,
            ventas_por_estado=ventas_por_estado,
            title=f'Reporte Completo - {cliente.nombre}'
        )
        
        return TemplateResponse(request, 'admin/ventas/cliente/reporte_completo.html', context)

# Inline para PagoVenta - debe definirse antes de VentasAdmin
class PagoVentaInline(admin.TabularInline):
    model = PagoVenta
    extra = 0
    readonly_fields = ('fecha_registro',)
    fields = ('fecha_pago', 'monto_pago', 'cuenta_destino', 'metodo_pago', 'referencia', 'notas')

# Inline para DocumentoCFDI - debe definirse antes de VentasAdmin
class DocumentoCFDIInline(admin.TabularInline):
    model = DocumentoCFDI
    fk_name = 'venta'
    extra = 0
    fields = ('subtipo', 'folio', 'uuid', 'fecha_emision', 'monto', 'estado')
    readonly_fields = ('subtipo', 'folio', 'uuid', 'fecha_emision', 'monto', 'estado')
    can_delete = False
    show_change_link = True

@admin.register(Agente)
class AgenteAdmin(ModelAdmin):
    list_display = ('nombre', 'fecha_registro')
    list_per_page = 12
   
class VentasResource(resources.ModelResource):
    agente = fields.Field(
        column_name='agente',
        attribute='agente_id',
        widget=ForeignKeyWidget(Agente, field='nombre'))
    
    producto = fields.Field(    
        column_name='producto',
        attribute='producto',
        widget=ForeignKeyWidget(Producto, field='variedad'))
    
    cliente = fields.Field(
        column_name='cliente',
        attribute='cliente',
        widget=ForeignKeyWidget(Cliente, field='nombre'))
    
    sucursal = fields.Field(
        column_name='sucursal',
        attribute='sucursal_id',
        widget=ForeignKeyWidget(Sucursal, field='nombre'))
    
    cuenta = fields.Field(
        column_name='cuenta',
        attribute='cuenta',
        widget=ForeignKeyWidget(Cuenta, field='numero_cuenta'))
    
    monto = fields.Field(
        column_name='monto',
        attribute='monto',
        widget=MoneyWidget())

    class Meta:
        model = Ventas
        fields = ('id', 'fecha_salida_manifiesto', 'agente', 'fecha_deposito', 'carga', 'PO', 'producto', 'cantidad', 'monto', 'descripcion', 'tipo_registro', 'cliente', 'fecha_registro', 'sucursal','cuenta')
        import_id_fields = ('id',)
        
    def dehydrate_agente(self, ventas):
        return ventas.agente_id.nombre if ventas.agente_id else ''
    
    def dehydrate_producto(self, ventas):
        return ventas.producto.variedad if ventas.producto_id else ''
    
    def dehydrate_cliente(self, ventas):
        return ventas.cliente.nombre
    
    def dehydrate_sucursal(self, ventas):
        return ventas.sucursal_id.nombre
    
    def dehydrate_cuenta(self, ventas):
        return ventas.cuenta.numero_cuenta if ventas.cuenta else ""

    def before_import_row(self, row, **kwargs):
        # Asigna un ID específico basado en un rango disponible
        if not row['id']:
            last_ventas = Ventas.objects.order_by('-id').first()
            next_id = last_ventas.id + 1 if last_ventas else 1
            row['id'] = next_id
       
@admin.register(Ventas)
class VentasAdmin(ModelAdmin):
    form = VentasAdminForm
    change_list_template = 'admin/ventas/ventas/change_list.html'
    list_select_related = ('cliente', 'cliente__pais', 'mercado_destino')

    class Media:
        js = ('js/ventas_form_logic.js',)
    
    list_display = (
        'fecha_salida_manifiesto', 'carga', 'get_cliente_info', 'get_monto_formateado', 
        'modalidad_pago', 'get_estado_cobranza', 'get_dias_vencimiento', 
        'tipo_venta', 'get_mercado_destino', 'get_saldo_pendiente'
    )
    
    list_filter = (
        VencimientoFilter, MontoVentaFilter, 'tipo_registro', 'modalidad_pago', 'estado_cobranza',
        'tipo_venta', 'fecha_salida_manifiesto', 'mercado_destino', 'termino_credito',
        'cliente__calificacion_credito', 'cliente__tipo_cliente'
    )
    
    search_fields = (
        'carga', 'cliente__nombre', 'producto__variedad', 'descripcion',
        'PO', 'pedimento',
    )
    search_help_text = 'Buscar por carga, cliente, producto/servicio, PO o pedimento'
    
    list_per_page = 30
    date_hierarchy = 'fecha_salida_manifiesto'
    
    inlines = [PagoVentaInline, DocumentoCFDIInline]
    
    actions = [
        'generar_reporte_cliente', 
        'marcar_como_pagado', 
        'generar_estado_cuenta',
        'enviar_notificacion_vencimiento',
        'exportar_cuentas_vencidas',
        'export_to_excel',
    ]
    
    fieldsets = (
        ('Información Básica', {
            'fields': ('fecha_salida_manifiesto', 'agente_id', 'fecha_deposito',
                       'carga', 'PO', 'pedimento')
        }),
        ('Concepto y Cliente', {
            'fields': ('producto', 'cantidad', 'monto', 'cliente',
                       'sucursal_id', 'descripcion')
        }),
        ('Modalidad de Pago', {
            'fields': ('modalidad_pago', 'termino_credito', 'fecha_vencimiento',
                       'estado_cobranza', 'monto_pagado'),
            'classes': ('wide',)
        }),
        ('Mercado y Exportación', {
            'fields': ('tipo_venta', 'mercado_destino', 'incoterm',
                       'moneda_venta', 'tipo_cambio', 'numero_carga_comprador'),
        }),
        ('Contabilidad', {
            'fields': ('cuenta', 'anticipo'),
        }),
        ('Tipo de Registro', {
            'fields': ('tipo_registro',),
            'description': 'Indica si el ingreso corresponde a una venta, maquila o servicio.'
        }),
    )
    
    readonly_fields = ('fecha_registro', 'monto_pagado')

    def save_model(self, request, obj, form, change):
        if obj.modalidad_pago == 'Credito' and not change and obj.cliente_id:
            try:
                monto = float(obj.monto.amount)
                if not obj.cliente.puede_otorgar_credito(monto):
                    disponible = obj.cliente.credito_disponible()
                    messages.warning(
                        request,
                        f"Límite de crédito insuficiente para {obj.cliente}. "
                        f"Disponible: ${disponible:,.2f}. La venta se guardó igualmente."
                    )
            except Exception:
                pass  # No bloquear el guardado por errores de validación
        super().save_model(request, obj, form, change)
        from .services.cache_service import VentasBalancesCache
        VentasBalancesCache.invalidar()

    def delete_model(self, request, obj):
        super().delete_model(request, obj)
        from .services.cache_service import VentasBalancesCache
        VentasBalancesCache.invalidar()

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                'balances/',
                self.admin_site.admin_view(self.ventas_balances_admin_view),
                name='ventas_ventas_balances',
            ),
            path(
                'balances/export/',
                self.admin_site.admin_view(self.exportar_balances_xlsx_admin_view),
                name='ventas_exportar_balances_xlsx',
            ),
            path(
                'reporte-cobranza/',
                self.admin_site.admin_view(self.reporte_cobranza_admin_view),
                name='ventas_reporte_cobranza',
            ),
            path(
                'reporte-cobranza/export-excel/',
                self.admin_site.admin_view(self.exportar_reporte_cobranza_excel),
                name='ventas_reporte_cobranza_excel',
            ),
            path(
                'reporte-cobranza/export-pdf/',
                self.admin_site.admin_view(self.exportar_reporte_cobranza_pdf),
                name='ventas_reporte_cobranza_pdf',
            ),
            path(
                'dashboard-ventas/',
                self.admin_site.admin_view(self.dashboard_ventas),
                name='ventas_dashboard',
            ),
            path(
                'reporte-cliente/<int:cliente_id>/',
                self.admin_site.admin_view(self.reporte_detallado_cliente),
                name='ventas_reporte_cliente',
            ),
            path(
                'exportar-excel/',
                self.admin_site.admin_view(self.exportar_excel_personalizado),
                name='ventas_exportar_excel',
            ),
            path(
                'api/cliente-info/<int:pk>/',
                self.admin_site.admin_view(self.api_cliente_info),
                name='ventas_api_cliente_info',
            ),
            path(
                'api/termino-credito-info/<int:pk>/',
                self.admin_site.admin_view(self.api_termino_credito_info),
                name='ventas_api_termino_credito_info',
            ),
            path(
                'importar-cfdi/',
                self.admin_site.admin_view(self.importar_desde_cfdi),
                name='ventas_importar_cfdi',
            ),
            path(
                'importar-cfdi-masivo/',
                self.admin_site.admin_view(self.importar_cfdi_masivo),
                name='ventas_importar_cfdi_masivo',
            ),
        ]
        return custom_urls + urls

    def api_cliente_info(self, request, pk):
        """Returns client data for form auto-fill logic (JSON)."""
        try:
            cliente = Cliente.objects.select_related(
                'pais', 'mercado_destino', 'termino_credito_predeterminado'
            ).get(pk=pk)
        except Cliente.DoesNotExist:
            return JsonResponse({'error': 'Not found'}, status=404)

        pais_nombre = cliente.pais.nombre if cliente.pais else ''
        es_extranjero = pais_nombre.lower() not in ('méxico', 'mexico')

        mercado_id = None
        if cliente.mercado_destino_id:
            mercado_id = cliente.mercado_destino_id
        elif es_extranjero:
            md = MercadoDestino.objects.filter(
                paises=cliente.pais, activo=True
            ).first()
            mercado_id = md.pk if md else None

        return JsonResponse({
            'es_extranjero': es_extranjero,
            'pais_nombre': pais_nombre,
            'mercado_destino_id': mercado_id,
            'termino_credito_id': cliente.termino_credito_predeterminado_id,
        })

    def api_termino_credito_info(self, request, pk):
        """Returns dias_credito for a TerminoCredito so JS can calculate due date."""
        try:
            tc = TerminoCredito.objects.get(pk=pk)
        except TerminoCredito.DoesNotExist:
            return JsonResponse({'error': 'Not found'}, status=404)
        return JsonResponse({'dias_credito': tc.dias_credito})

    # =========================================================================
    # IMPORTAR VENTA DESDE CFDI XML
    # =========================================================================

    def importar_desde_cfdi(self, request):
        """
        Two-step view for importing a Ventas record from a CFDI XML file.

        GET  → show upload form (step 1)
        POST step=upload  → parse XML, show confirmation form (step 2)
        POST step=confirm → validate and create Ventas, redirect to change view
        """
        from .cfdi_parser import classify_subtipo, parse_cfdi
        from djmoney.money import Money

        opts = self.model._meta
        admin_url = reverse('admin:%s_%s_changelist' % (opts.app_label, opts.model_name))

        # ── Step 2 → confirm & save ──────────────────────────────────────
        if request.method == 'POST' and request.POST.get('_step') == 'confirm':
            form = CFDIConfirmForm(request.POST)
            if form.is_valid():
                cd = form.cleaned_data
                try:
                    venta = Ventas(
                        monto=Money(cd['monto'], cd['moneda_venta']),
                        moneda_venta=cd['moneda_venta'],
                        tipo_cambio=cd['tipo_cambio'],
                        incoterm=cd['incoterm'],
                        tipo_venta=cd['tipo_venta'],
                        modalidad_pago=cd['modalidad_pago'],
                        cantidad=cd['cantidad'],
                        descripcion=cd['descripcion'],
                        PO=cd['PO'],
                        cliente=cd['cliente'],
                        producto=cd.get('producto'),
                        fecha_salida_manifiesto=cd['fecha_salida_manifiesto'],
                        fecha_deposito=cd['fecha_deposito'],
                        agente_id=cd.get('agente_id'),
                        pedimento=cd['pedimento'],
                        carga=cd['carga'],
                        sucursal_id=cd['sucursal_id'],
                        cuenta=cd['cuenta'],
                        tipo_registro=cd['tipo_registro'],
                        termino_credito=cd.get('termino_credito'),
                    )
                    venta.full_clean()
                    venta.save()

                    if cd['tipo_registro'] == Ventas.TipoRegistro.SERVICIO:
                        subtipo = DocumentoCFDI.SubtipoDocumento.INGRESO_SERVICIO
                    else:
                        subtipo = (
                            DocumentoCFDI.SubtipoDocumento.VENTA_EXPORTACION
                            if cd['tipo_venta'] == Ventas.TipoVenta.EXPORTACION
                            else DocumentoCFDI.SubtipoDocumento.VENTA_NACIONAL
                        )
                    DocumentoCFDI.objects.create(
                        cliente=cd['cliente'],
                        tipo=DocumentoCFDI.TipoDocumento.INGRESO,
                        subtipo=subtipo,
                        folio=cd['folio_factura'] or None,
                        fecha_emision=cd['fecha_emision_cfdi'],
                        monto=Money(cd['monto'], cd['moneda_venta']),
                        moneda=cd['moneda_venta'],
                        tipo_cambio=cd['tipo_cambio'],
                        venta=venta,
                    )

                    messages.success(
                        request,
                        f'Venta importada correctamente desde CFDI — venta #{venta.pk}.'
                    )
                    change_url = reverse(
                        'admin:%s_%s_change' % (opts.app_label, opts.model_name),
                        args=[venta.pk]
                    )
                    return redirect(change_url)
                except Exception as exc:
                    messages.error(request, f'Error al guardar la venta: {exc}')

            context = dict(
                self.admin_site.each_context(request),
                form=form,
                step='confirm',
                title='Importar venta — confirmar datos',
                opts=opts,
            )
            return TemplateResponse(request, 'admin/ventas/importar_cfdi.html', context)

        # ── Step 1 → parse XML ───────────────────────────────────────────
        if request.method == 'POST' and request.POST.get('_step') == 'upload':
            upload_form = CFDIUploadForm(request.POST, request.FILES)
            if upload_form.is_valid():
                xml_bytes = upload_form.cleaned_data['xml_file'].read()
                try:
                    parsed = parse_cfdi(xml_bytes)
                except ValueError as exc:
                    messages.error(request, str(exc))
                    upload_form = CFDIUploadForm()
                    context = dict(
                        self.admin_site.each_context(request),
                        form=upload_form,
                        step='upload',
                        title='Importar venta desde CFDI (XML)',
                        opts=opts,
                    )
                    return TemplateResponse(request, 'admin/ventas/importar_cfdi.html', context)

                subtipo = classify_subtipo(parsed)
                if subtipo == 'ingreso_mixto':
                    messages.error(
                        request,
                        'El CFDI mezcla productos y servicios. No se puede '
                        'convertir automáticamente en una sola operación.',
                    )
                    return TemplateResponse(
                        request,
                        'admin/ventas/importar_cfdi.html',
                        dict(
                            self.admin_site.each_context(request),
                            form=CFDIUploadForm(), step='upload',
                            title='Importar venta desde CFDI (XML)', opts=opts,
                        ),
                    )
                if subtipo not in (
                    'venta_nacional', 'venta_exportacion', 'ingreso_servicio',
                ):
                    messages.error(
                        request,
                        'Este documento no es una factura de venta o servicio. '
                        'Utiliza la importación masiva para procesarlo.',
                    )
                    return TemplateResponse(
                        request,
                        'admin/ventas/importar_cfdi.html',
                        dict(
                            self.admin_site.each_context(request),
                            form=CFDIUploadForm(), step='upload',
                            title='Importar venta desde CFDI (XML)', opts=opts,
                        ),
                    )

                es_servicio = subtipo == 'ingreso_servicio'

                # Try to find best client match against receptor nombre
                receptor_nombre = parsed.get('_receptor_nombre', '')
                cliente_inicial = None
                cliente_sugerido_nombre = ''
                if receptor_nombre:
                    # 1) Exact full-name match (case-insensitive)
                    qs = Cliente.objects.filter(nombre__iexact=receptor_nombre, activo=True)
                    if qs.exists():
                        cliente_inicial = qs.first()
                    else:
                        # 2) Try each significant word (len > 3) from receptor against DB names
                        words = [w for w in receptor_nombre.split() if len(w) > 3]
                        for word in words:
                            qs = Cliente.objects.filter(nombre__icontains=word, activo=True)
                            if qs.count() == 1:
                                cliente_inicial = qs.first()
                                break
                            elif qs.exists():
                                cliente_inicial = qs.first()
                                # keep searching for a more specific match
                    cliente_sugerido_nombre = receptor_nombre

                # Detect product by scanning DB varieties against the CFDI description
                descripcion_cfdi = parsed.get('descripcion', '')
                no_id = parsed.get('_no_identificacion', '')
                fraccion = parsed.get('_fraccion_arancelaria', '')
                producto_inicial = None
                # 1) Match any Producto.variedad found inside the CFDI description text
                if descripcion_cfdi and not es_servicio:
                    for p in Producto.objects.filter(disponible=True).order_by('variedad'):
                        if p.variedad and p.variedad.strip().lower() in descripcion_cfdi.lower():
                            producto_inicial = p
                            break
                # 2) Fallback: match by NoIdentificacion field
                if not es_servicio and not producto_inicial and no_id:
                    producto_inicial = (
                        Producto.objects.filter(variedad__icontains=no_id).first()
                        or Producto.objects.filter(descripcion__icontains=no_id).first()
                    )
                # 3) Fallback: match by nombre (e.g. "Mango") when fraccion arancelaria present
                if not es_servicio and not producto_inicial and fraccion:
                    producto_inicial = Producto.objects.filter(nombre__icontains='Mango').first()

                fecha_cfdi = parsed.get('fecha_emision_cfdi') or timezone.now().date()
                initial = {
                    'folio_factura': parsed.get('folio_factura', ''),
                    'fecha_emision_cfdi': parsed.get('fecha_emision_cfdi'),
                    'monto': parsed.get('monto', '0'),
                    'moneda_venta': parsed.get('moneda_venta', 'MXN'),
                    'tipo_cambio': parsed.get('tipo_cambio', '1.0000'),
                    'incoterm': parsed.get('incoterm', ''),
                    'tipo_venta': parsed.get('tipo_venta', 'Nacional'),
                    'modalidad_pago': parsed.get('modalidad_pago', 'Contado'),
                    'cantidad': parsed.get('cantidad'),
                    'descripcion': parsed.get('descripcion', ''),
                    'PO': parsed.get('PO', ''),
                    'cliente': cliente_inicial,
                    'producto': producto_inicial,
                    'tipo_registro': (
                        Ventas.TipoRegistro.SERVICIO
                        if es_servicio else Ventas.TipoRegistro.VENTA
                    ),
                    'fecha_salida_manifiesto': fecha_cfdi,
                    'fecha_deposito': fecha_cfdi,
                }

                form = CFDIConfirmForm(initial=initial)
                context = dict(
                    self.admin_site.each_context(request),
                    form=form,
                    step='confirm',
                    title='Importar venta — confirmar datos',
                    opts=opts,
                    parsed=parsed,
                    cliente_sugerido_nombre=cliente_sugerido_nombre,
                    es_servicio=es_servicio,
                )
                return TemplateResponse(request, 'admin/ventas/importar_cfdi.html', context)

        # ── Step 0 → show upload form ────────────────────────────────────
        upload_form = CFDIUploadForm()
        context = dict(
            self.admin_site.each_context(request),
            form=upload_form,
            step='upload',
            title='Importar venta desde CFDI (XML)',
            opts=opts,
        )
        return TemplateResponse(request, 'admin/ventas/importar_cfdi.html', context)

    def importar_cfdi_masivo(self, request):
        """
        Importación masiva de CFDI (ventas, notas, anticipos y recibos de pago)
        a partir de múltiples XML o un ZIP descargado de Blikon.
        """
        from .cfdi_parser import parse_cfdi, classify_subtipo
        from .services.cfdi_import_service import (
            crear_cliente_desde_cfdi, crear_producto_desde_cfdi, importar_cfdi,
            match_cliente, match_producto, parsed_to_json, parsed_from_json,
            sugerir_pais, sugerir_producto_desde_cfdi,
        )

        opts = self.model._meta

        # ── Step 2 → confirm & import ─────────────────────────────────────
        if request.method == 'POST' and request.POST.get('_step') == 'confirm':
            total = 0
            try:
                total = int(request.POST.get('total', '0'))
            except (TypeError, ValueError):
                total = 0

            upload_token = request.POST.get('upload_token') or ''
            archivos_map = (
                request.session.pop(f'cfdi_archivos_{upload_token}', None)
                if upload_token else None
            ) or {}
            archivos_xml = archivos_map.get('xml', {})
            archivos_pdf = archivos_map.get('pdf', {})
            tempdir = archivos_map.get('dir')

            filas = []
            errores_preparacion = {}
            for i in range(total):
                if not request.POST.get(f'include_{i}'):
                    continue
                raw = request.POST.get(f'row_{i}', '')
                if not raw:
                    continue
                try:
                    parsed = parsed_from_json(raw)
                    fila = {
                        'i': i,
                        'parsed': parsed,
                        'cliente_id': request.POST.get(f'cliente_{i}') or None,
                        'producto_id': request.POST.get(f'producto_{i}') or None,
                        'crear_cliente': bool(request.POST.get(f'crear_cliente_{i}')),
                        'crear_producto': bool(request.POST.get(f'crear_producto_{i}')),
                        'pais_id': request.POST.get(f'pais_cliente_{i}') or None,
                    }
                    filas.append(fila)
                except Exception as exc:
                    errores_preparacion[i] = str(exc)

            prioridad_cfdi = {
                'venta_nacional': 0,
                'venta_exportacion': 0,
                'ingreso_servicio': 0,
                'ingreso_mixto': 0,
                'remanente_anticipo': 1,
                'nota_cargo': 2,
                'nota_credito': 2,
                'recibo_pago': 3,
            }
            filas.sort(key=lambda fila: prioridad_cfdi.get(
                classify_subtipo(fila['parsed']), 99
            ))

            # Crea primero los receptores elegidos para que una sola alta se
            # aplique a todos los CFDI con la misma identidad fiscal del lote.
            clientes_creados = set()
            for fila in filas:
                if fila['i'] in errores_preparacion:
                    continue
                if fila['cliente_id'] or not fila['crear_cliente']:
                    continue
                try:
                    if not request.user.has_perm('ventas.add_cliente'):
                        raise PermissionError('No tienes permiso para crear clientes.')
                    pais = Pais.objects.get(pk=fila['pais_id']) if fila['pais_id'] else None
                    cliente, creado = crear_cliente_desde_cfdi(
                        fila['parsed'], pais=pais,
                    )
                    fila['cliente_id'] = cliente.pk
                    if creado:
                        clientes_creados.add(cliente.pk)
                except Exception as exc:
                    errores_preparacion[fila['i']] = str(exc)

            # Crea primero los productos solicitados. Después vuelve a ejecutar
            # el matching para reutilizarlos en los demás CFDI del mismo lote.
            productos_creados = set()
            for fila in filas:
                if fila['i'] in errores_preparacion:
                    continue
                subtipo = classify_subtipo(fila['parsed'])
                if (
                    subtipo not in ('venta_nacional', 'venta_exportacion')
                    or fila['producto_id']
                    or not fila['crear_producto']
                ):
                    continue
                try:
                    if not request.user.has_perm('catalogo.add_producto'):
                        raise PermissionError('No tienes permiso para crear productos.')
                    producto, creado = crear_producto_desde_cfdi(fila['parsed'])
                    fila['producto_id'] = producto.pk
                    if creado:
                        productos_creados.add(producto.pk)
                except Exception as exc:
                    errores_preparacion[fila['i']] = str(exc)

            for fila in filas:
                if fila['i'] in errores_preparacion:
                    continue
                subtipo = classify_subtipo(fila['parsed'])
                if (
                    subtipo in ('venta_nacional', 'venta_exportacion')
                    and not fila['producto_id']
                ):
                    producto = match_producto(fila['parsed'])
                    if producto:
                        fila['producto_id'] = producto.pk
                    else:
                        errores_preparacion[fila['i']] = (
                            'Selecciona un producto o créalo con los datos del CFDI.'
                        )

            resultados = []
            errores = 0
            for fila in filas:
                i = fila['i']
                parsed = fila['parsed']
                subtipo = classify_subtipo(parsed)
                if i in errores_preparacion:
                    errores += 1
                    resultados.append({
                        'ok': False,
                        'folio': parsed.get('folio_factura') or parsed.get('uuid'),
                        'subtipo': subtipo,
                        'error': errores_preparacion[i],
                    })
                    continue
                try:
                    cliente = (
                        Cliente.objects.get(pk=fila['cliente_id'], activo=True)
                        if fila['cliente_id'] else None
                    )
                    producto = (
                        Producto.objects.get(pk=fila['producto_id'], disponible=True)
                        if fila['producto_id'] else None
                    )

                    from django.core.files.base import File
                    uuid_lower = (parsed.get('uuid') or '').strip().lower()
                    archivo_xml = None
                    archivo_pdf = None
                    if uuid_lower and archivos_xml.get(uuid_lower):
                        archivo_xml = File(
                            open(archivos_xml[uuid_lower], 'rb'),
                            name=f'{uuid_lower}.xml',
                        )
                    if uuid_lower and archivos_pdf.get(uuid_lower):
                        archivo_pdf = File(
                            open(archivos_pdf[uuid_lower], 'rb'),
                            name=f'{uuid_lower}.pdf',
                        )

                    _obj, doc, subtipo = importar_cfdi(
                        parsed, cliente=cliente, producto=producto,
                        archivo_pdf=archivo_pdf, archivo_xml=archivo_xml,
                    )
                    resultados.append({
                        'ok': True,
                        'folio': parsed.get('folio_factura') or parsed.get('uuid') or doc.pk,
                        'subtipo': subtipo,
                        'doc_id': doc.pk,
                        'cliente_creado': doc.cliente_id in clientes_creados,
                        'producto_creado': (
                            getattr(_obj, 'producto_id', None) in productos_creados
                        ),
                    })
                except Exception as exc:
                    errores += 1
                    resultados.append({
                        'ok': False,
                        'folio': parsed.get('folio_factura') or parsed.get('uuid'),
                        'subtipo': subtipo,
                        'error': str(exc),
                    })

            if tempdir:
                import shutil
                shutil.rmtree(tempdir, ignore_errors=True)


            context = dict(
                self.admin_site.each_context(request),
                step='done',
                resultados=resultados,
                errores=errores,
                importados=sum(1 for resultado in resultados if resultado.get('ok')),
                clientes_creados=len(clientes_creados),
                productos_creados=len(productos_creados),
                title='Resultado de importación masiva',
                opts=opts,
            )
            return TemplateResponse(request, 'admin/ventas/importar_cfdi_masivo.html', context)

        # ── Step 1 → parse XMLs ───────────────────────────────────────────
        if request.method == 'POST' and request.POST.get('_step') == 'upload':
            import os
            import tempfile
            import uuid as uuid_mod

            archivos = request.FILES.getlist('archivos')
            errores_upload = []
            archivos_validos = []
            for f in archivos:
                error = _validar_archivo_cfdi(f)
                if error:
                    errores_upload.append(error)
                else:
                    archivos_validos.append(f)

            if not archivos_validos and not errores_upload:
                errores_upload.append('Debes seleccionar al menos un archivo.')

            if not errores_upload and archivos_validos:
                extraidos = _extraer_archivos(archivos_validos)
                xmls = [(n, b) for n, b, e in extraidos if e == 'xml']
                pdfs = [(n, b) for n, b, e in extraidos if e == 'pdf']

                # UUID de cada PDF, extraído del nombre del archivo (Blikon).
                pdfs_por_uuid = {}
                for nombre, contenido in pdfs:
                    m = _UUID_RE.search(nombre)
                    if m:
                        pdfs_por_uuid[m.group(0).lower()] = contenido

                tempdir = tempfile.mkdtemp(prefix='cfdi_import_')
                archivos_map = {'xml': {}, 'pdf': {}, 'dir': tempdir}

                previews = []
                uuids_lote = set()
                for nombre, xml_bytes in xmls:
                    item = {'nombre': nombre}
                    try:
                        parsed = parse_cfdi(xml_bytes)
                        subtipo = classify_subtipo(parsed)
                        uuid = (parsed.get('uuid') or '').strip()
                        uuid_lower = uuid.lower()
                        duplicado = None
                        if uuid and uuid in uuids_lote:
                            duplicado = 'El UUID está repetido dentro de este lote.'
                        elif uuid and DocumentoCFDI.objects.filter(uuid__iexact=uuid).exists():
                            duplicado = 'Este UUID ya fue importado anteriormente.'
                        if uuid:
                            uuids_lote.add(uuid)

                        tiene_pdf = False
                        if uuid_lower:
                            xml_path = os.path.join(tempdir, f'{uuid_lower}.xml')
                            with open(xml_path, 'wb') as fh:
                                fh.write(xml_bytes)
                            archivos_map['xml'][uuid_lower] = xml_path

                            if uuid_lower in pdfs_por_uuid:
                                pdf_path = os.path.join(tempdir, f'{uuid_lower}.pdf')
                                with open(pdf_path, 'wb') as fh:
                                    fh.write(pdfs_por_uuid[uuid_lower])
                                archivos_map['pdf'][uuid_lower] = pdf_path
                                tiene_pdf = True

                        item.update({
                            'parsed': parsed,
                            'subtipo': subtipo,
                            'subtipo_label': DocumentoCFDI.SubtipoDocumento(subtipo).label,
                            'cliente': match_cliente(parsed),
                            'pais_sugerido': sugerir_pais(parsed),
                            'producto': match_producto(parsed),
                            'producto_sugerido': sugerir_producto_desde_cfdi(parsed),
                            'receptor_nombre': parsed.get('_receptor_nombre') or '',
                            'receptor_rfc': parsed.get('_receptor_rfc') or '',
                            'receptor_residencia': (
                                parsed.get('_receptor_residencia_fiscal') or ''
                            ),
                            'receptor_registro_fiscal': (
                                parsed.get('_receptor_num_reg_id_trib') or ''
                            ),
                            'json': parsed_to_json(parsed),
                            'duplicado': duplicado,
                            'tiene_pdf': tiene_pdf,
                        })
                        if subtipo == 'ingreso_mixto':
                            item['error'] = (
                                'El CFDI mezcla productos y servicios. Requiere '
                                'revisión manual y no se importará como una sola venta.'
                            )
                    except Exception as exc:
                        item['error'] = str(exc)
                    previews.append(item)

                token = uuid_mod.uuid4().hex
                request.session[f'cfdi_archivos_{token}'] = archivos_map

                context = dict(
                    self.admin_site.each_context(request),
                    step='confirm',
                    upload_token=token,
                    previews=previews,
                    total_archivos=len(previews),
                    total_listos=sum(
                        1 for item in previews
                        if not item.get('error') and not item.get('duplicado')
                    ),
                    total_sin_cliente=sum(
                        1 for item in previews
                        if not item.get('error') and not item.get('duplicado')
                        and not item.get('cliente')
                    ),
                    total_duplicados=sum(
                        1 for item in previews if item.get('duplicado')
                    ),
                    clientes=Cliente.objects.filter(activo=True).order_by('nombre'),
                    paises=Pais.objects.order_by('nombre'),
                    productos=Producto.objects.filter(disponible=True).order_by('variedad'),
                    puede_crear_cliente=request.user.has_perm('ventas.add_cliente'),
                    puede_crear_producto=request.user.has_perm('catalogo.add_producto'),
                    title='Confirmar importación masiva de CFDI',
                    opts=opts,
                )
                return TemplateResponse(request, 'admin/ventas/importar_cfdi_masivo.html', context)

            context = dict(
                self.admin_site.each_context(request),
                errores_upload=errores_upload,
                step='upload',
                title='Importar CFDI masivo',
                opts=opts,
            )
            return TemplateResponse(request, 'admin/ventas/importar_cfdi_masivo.html', context)

        # ── Step 0 → upload form ──────────────────────────────────────────
        context = dict(
            self.admin_site.each_context(request),
            step='upload',
            title='Importar CFDI masivo',
            opts=opts,
        )
        return TemplateResponse(request, 'admin/ventas/importar_cfdi_masivo.html', context)

    def get_cliente_info(self, obj):
        """Información del cliente con indicador de riesgo"""
        cliente = obj.cliente
        pais = getattr(cliente, 'pais', None)
        pais_nombre = getattr(pais, 'nombre', '')
        bandera = getattr(pais, 'bandera', None)
        bandera_url = ''
        if bandera and getattr(bandera, 'name', ''):
            try:
                bandera_url = bandera.url
            except (AttributeError, ValueError):
                bandera_url = ''

        if bandera_url:
            pais_html = format_html(
                '<span class="sales-country sales-country--flag" '
                'aria-label="País: {}" title="{}">'
                '<img class="sales-country__flag" src="{}" alt="" '
                'width="18" height="12" loading="lazy">'
                '</span>',
                pais_nombre,
                pais_nombre,
                bandera_url,
            )
        elif pais_nombre:
            pais_html = format_html(
                '<span class="sales-country sales-country--fallback">'
                '<i class="fas fa-map-marker-alt" aria-hidden="true"></i>'
                '<span>{}</span>'
                '</span>',
                pais_nombre,
            )
        else:
            pais_html = mark_safe('<span class="sales-muted">—</span>')

        risk_class = {
            'A+': 'sales-risk--a-plus',
            'A': 'sales-risk--a',
            'B': 'sales-risk--b',
            'C': 'sales-risk--c',
        }.get(cliente.calificacion_credito, '')
        
        return format_html(
            '<strong class="sales-client-name">{}</strong><br>'
            '<span class="sales-client-meta sales-risk {}">'
            '<i class="fas fa-chart-line" aria-hidden="true"></i>'
            '<span>{}</span>'
            '</span><br>'
            '{}',
            cliente.nombre,
            risk_class,
            cliente.get_calificacion_credito_display(),
            pais_html,
        )
    get_cliente_info.short_description = 'Cliente & Riesgo'
    
    def get_monto_formateado(self, obj):
        """Monto con formato mejorado"""
        money_class = 'sales-money--cash' if obj.modalidad_pago == 'Contado' else 'sales-money--credit'
        monto_str = f"{float(obj.monto.amount):,.2f}"
        return format_html(
            '<span class="sales-money {}">${}</span><br>'
            '<small class="sales-muted">{}</small>',
            money_class,
            monto_str,
            obj.moneda_venta
        )
    get_monto_formateado.short_description = 'Monto'
    
    def get_saldo_pendiente(self, obj):
        """Saldo pendiente con formato visual"""
        if obj.modalidad_pago == 'Contado':
            return mark_safe(
                '<span class="sales-badge sales-badge--paid">'
                '<i class="fas fa-check" aria-hidden="true"></i>Pagado'
                '</span>'
            )
        
        saldo = obj.saldo_pendiente()
        if saldo <= 0:
            return mark_safe(
                '<span class="sales-badge sales-badge--paid">'
                '<i class="fas fa-check" aria-hidden="true"></i>$0.00'
                '</span>'
            )
        
        balance_class = 'sales-balance--overdue' if obj.esta_vencida() else 'sales-balance--open'
        return format_html(
            '<span class="sales-balance {}">${}</span>',
            balance_class, f"{float(saldo):,.2f}"
        )
    get_saldo_pendiente.short_description = 'Saldo Pendiente'
    
    # Métodos originales del admin
    def get_estado_cobranza(self, obj):
        colors = {
            'Pagado': 'sales-badge--paid',
            'Pendiente': 'sales-badge--pending',
            'Parcial': 'sales-badge--partial',
            'Vencido': 'sales-badge--overdue',
            'Incobrable': 'sales-badge--bad',
        }
        badge_class = colors.get(obj.estado_cobranza, 'sales-badge--neutral')
        return format_html(
            '<span class="sales-badge {}">{}</span>',
            badge_class, obj.get_estado_cobranza_display()
        )
    get_estado_cobranza.short_description = 'Estado Cobranza'
    
    def get_dias_vencimiento(self, obj):
        if obj.modalidad_pago == 'Contado':
            return '-'
        dias = obj.dias_vencido()
        if dias > 0:
            return format_html('<span class="sales-due sales-due--overdue">+{} días</span>', dias)
        elif dias < 0:
            return format_html('<span class="sales-due sales-due--ok">{} días</span>', abs(dias))
        else:
            return mark_safe('<span class="sales-due sales-due--today">Vence hoy</span>')
    get_dias_vencimiento.short_description = 'Vencimiento'
    
    def get_mercado_destino(self, obj):
        return obj.mercado_destino.nombre if obj.mercado_destino else obj.tipo_venta
    get_mercado_destino.short_description = 'Mercado'
    
    # =============================================================================
    # ACCIONES PERSONALIZADAS
    # =============================================================================
    
    def generar_reporte_cliente(self, request, queryset):
        """Generar reporte consolidado por cliente"""
        clientes_ids = list(queryset.values_list('cliente_id', flat=True).distinct())
        
        if len(clientes_ids) > 10:
            messages.error(request, 'Selecciona máximo 10 clientes para el reporte.')
            return
            
        # Redirigir a vista de reporte personalizada
        return redirect('admin:ventas_reporte_consolidado') + '?clientes=' + ','.join(map(str, clientes_ids))
    
    generar_reporte_cliente.short_description = "📊 Generar reporte por cliente"
    
    def marcar_como_pagado(self, request, queryset):
        """Marcar ventas a crédito como pagadas"""
        ventas_credito = queryset.filter(modalidad_pago='Credito')
        count = 0
        
        for venta in ventas_credito:
            if venta.estado_cobranza != 'Pagado':
                venta.estado_cobranza = 'Pagado'
                venta.monto_pagado = venta.monto
                venta.save()
                count += 1
        
        if count > 0:
            messages.success(request, f'{count} ventas marcadas como pagadas.')
        else:
            messages.warning(request, 'No hay ventas a crédito pendientes para marcar como pagadas.')
    
    marcar_como_pagado.short_description = "💰 Marcar como pagado"
    
    def exportar_cuentas_vencidas(self, request, queryset):
        """Exportar solo cuentas vencidas a Excel"""
        from django.http import HttpResponse
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        hoy = timezone.now().date()
        vencidas = queryset.filter(
            modalidad_pago='Credito',
            fecha_vencimiento__lt=hoy,
            estado_cobranza__in=['Pendiente', 'Parcial']
        )
        
        if not vencidas.exists():
            messages.warning(request, 'No hay cuentas vencidas en la selección.')
            return
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Cuentas Vencidas'
        
        # Encabezados
        headers = [
            'Fecha Venta', 'Cliente', 'Carga', 'Monto Original', 
            'Saldo Pendiente', 'Fecha Vencimiento', 'Días Vencido', 
            'Calificación Cliente', 'Teléfono', 'Email'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
        
        # Datos
        for row, venta in enumerate(vencidas, 2):
            ws.cell(row=row, column=1, value=venta.fecha_salida_manifiesto)
            ws.cell(row=row, column=2, value=venta.cliente.nombre)
            ws.cell(row=row, column=3, value=venta.carga)
            ws.cell(row=row, column=4, value=float(venta.monto.amount))
            ws.cell(row=row, column=5, value=venta.saldo_pendiente())
            ws.cell(row=row, column=6, value=venta.fecha_vencimiento)
            ws.cell(row=row, column=7, value=venta.dias_vencido())
            ws.cell(row=row, column=8, value=venta.cliente.calificacion_credito)
            ws.cell(row=row, column=9, value=venta.cliente.telefono)
            ws.cell(row=row, column=10, value=venta.cliente.correo)
        
        # Ajustar columnas
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=cuentas_vencidas_{hoy.strftime("%Y%m%d")}.xlsx'
        
        wb.save(response)
        return response
    
    exportar_cuentas_vencidas.short_description = "📄 Exportar cuentas vencidas"

    def export_to_excel(self, request, queryset):
        """Exporta las ventas seleccionadas a Excel con tabla dinámica +
        hoja de resumen ejecutivo de cuentas por cobrar."""
        import openpyxl
        import datetime as dt
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from collections import defaultdict

        NAVY  = "1E3A5F"
        TEAL  = "1AADBC"
        LIGHT = "D6EAF8"
        WHITE = "FFFFFF"
        MONEY = '"$"#,##0.00'
        hoy   = dt.date.today()

        def _nav_hdr(cell, bg=NAVY):
            cell.font      = Font(bold=True, color=WHITE)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        def _section_title(ws, row, col_start, col_end, text):
            cell = ws.cell(row=row, column=col_start, value=text)
            cell.font = Font(bold=True, color=WHITE, size=11)
            cell.fill = PatternFill("solid", fgColor=NAVY)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if col_end > col_start:
                ws.merge_cells(start_row=row, start_column=col_start,
                               end_row=row, end_column=col_end)
            ws.row_dimensions[row].height = 20

        def _add_table(ws, name, hdr_row, data_end_row, col_end):
            if data_end_row < hdr_row + 1:
                return
            tab = Table(
                displayName=name,
                ref=f"A{hdr_row}:{get_column_letter(col_end)}{data_end_row}",
            )
            tab.tableStyleInfo = TableStyleInfo(
                name="TableStyleLight2",
                showFirstColumn=False, showLastColumn=False,
                showRowStripes=True, showColumnStripes=False,
            )
            ws.add_table(tab)

        qs = queryset.select_related(
            'cliente', 'producto', 'sucursal_id', 'cuenta',
            'termino_credito', 'mercado_destino'
        ).order_by('fecha_salida_manifiesto')
        data = list(qs)

        wb = openpyxl.Workbook()

        # ── HOJA 1 — Detalle ─────────────────────────────────────────────────
        ws = wb.active
        ws.title = "Ventas"

        COLUMNS = [
            ("Fecha",          lambda v: v.fecha_salida_manifiesto,                     13, "DD/MM/YYYY"),
            ("Cliente",        lambda v: v.cliente.nombre,                              28, "@"),
            ("Carga",          lambda v: v.carga or "",                                 16, "@"),
            ("Producto / Servicio", lambda v: v.producto.variedad if v.producto_id else (v.descripcion or "Servicio"), 28, "@"),
            ("Sucursal",       lambda v: v.sucursal_id.nombre,                          18, "@"),
            ("Tipo",           lambda v: v.tipo_venta,                                  14, "@"),
            ("Modalidad",      lambda v: v.modalidad_pago,                              14, "@"),
            ("Monto",          lambda v: float(v.monto.amount),                         16, MONEY),
            ("Monto Pagado",   lambda v: float(v.monto_pagado.amount),                  16, MONEY),
            ("Saldo Pendiente",lambda v: round(v.saldo_por_cobrar(), 2), 16, MONEY),
            ("Estado",         lambda v: v.estado_cobranza,                             14, "@"),
            ("Vencimiento",    lambda v: v.fecha_vencimiento,                           13, "DD/MM/YYYY"),
        ]

        for ci, (hdr, _, width, _) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=ci, value=hdr)
            _nav_hdr(cell)
            ws.column_dimensions[get_column_letter(ci)].width = width

        for ri, venta in enumerate(data, 2):
            for ci, (_, getter, _, fmt) in enumerate(COLUMNS, 1):
                cell = ws.cell(row=ri, column=ci, value=getter(venta))
                cell.number_format = fmt
                cell.alignment = Alignment(vertical="center")

        last_data = len(data) + 1

        if data:
            tab_main = Table(
                displayName="TablaVentas",
                ref=f"A1:{get_column_letter(len(COLUMNS))}{last_data}",
            )
            tab_main.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False, showLastColumn=False,
                showRowStripes=True, showColumnStripes=False,
            )
            ws.add_table(tab_main)

        # Fila TOTAL fuera de la tabla
        total_r = last_data + 1
        for ci, col_label in ((8, "Monto"), (9, "Pagado"), (10, "Pendiente")):
            lbl = ws.cell(row=total_r, column=ci - 1)
            if ci == 8:
                lbl.value = "TOTAL"
                lbl.font  = Font(bold=True)
                lbl.alignment = Alignment(horizontal="right")
            col_ltr = get_column_letter(ci)
            tc = ws.cell(row=total_r, column=ci,
                         value=f"=SUM({col_ltr}2:{col_ltr}{last_data})")
            tc.number_format = MONEY
            tc.font = Font(bold=True)
            tc.fill = PatternFill("solid", fgColor=LIGHT)

        ws.freeze_panes = "A2"

        # ── HOJA 2 — Cuentas por Cobrar (Resumen Ejecutivo) ──────────────────
        ws2 = wb.create_sheet("Cuentas por Cobrar")

        if not data:
            ws2.cell(row=1, column=1, value="Sin datos seleccionados.")
        else:
            total_monto     = sum(float(v.monto.amount) for v in data)
            total_pagado    = sum(float(v.monto_pagado.amount) for v in data)
            total_pendiente = round(sum(v.saldo_por_cobrar() for v in data), 2)

            credito  = [v for v in data if v.modalidad_pago == 'Credito']
            vencidas = [v for v in credito if v.fecha_vencimiento and v.fecha_vencimiento < hoy
                        and v.estado_cobranza in ('Pendiente', 'Parcial')]
            pendientes = [v for v in credito if v.estado_cobranza in ('Pendiente', 'Parcial')]

            monto_vencido   = sum(v.saldo_por_cobrar() for v in vencidas)
            monto_pendiente_cxc = sum(v.saldo_por_cobrar() for v in pendientes)

            by_estado   = defaultdict(lambda: {'count': 0, 'monto': 0.0, 'pendiente': 0.0})
            by_cliente  = defaultdict(lambda: {'count': 0, 'monto': 0.0, 'pendiente': 0.0})
            by_venc     = defaultdict(lambda: {'count': 0, 'monto': 0.0})   # antigüedad

            for v in data:
                amt  = float(v.monto.amount)
                pend = round(v.saldo_por_cobrar(), 2)
                est  = v.estado_cobranza
                cli  = v.cliente.nombre

                by_estado[est]['count']    += 1
                by_estado[est]['monto']    += amt
                by_estado[est]['pendiente']+= pend

                by_cliente[cli]['count']    += 1
                by_cliente[cli]['monto']    += amt
                by_cliente[cli]['pendiente']+= pend

                # Antigüedad solo para créditos pendientes
                if v.modalidad_pago == 'Credito' and v.fecha_vencimiento and pend > 0:
                    dias_v = (hoy - v.fecha_vencimiento).days
                    if dias_v <= 0:
                        bucket = "Por vencer"
                    elif dias_v <= 30:
                        bucket = "1-30 días"
                    elif dias_v <= 60:
                        bucket = "31-60 días"
                    elif dias_v <= 90:
                        bucket = "61-90 días"
                    else:
                        bucket = "+90 días"
                    by_venc[bucket]['count'] += 1
                    by_venc[bucket]['monto'] += pend

            # Anchos
            for col, w in zip('ABCDE', [30, 14, 18, 18, 14]):
                ws2.column_dimensions[col].width = w

            r = 1
            # Cabecera
            hdr_cell = ws2.cell(row=r, column=1,
                                value="  RESUMEN EJECUTIVO — CUENTAS POR COBRAR")
            hdr_cell.font      = Font(bold=True, color=WHITE, size=14)
            hdr_cell.fill      = PatternFill("solid", fgColor=NAVY)
            hdr_cell.alignment = Alignment(horizontal="left", vertical="center")
            ws2.merge_cells(f"A{r}:E{r}")
            ws2.row_dimensions[r].height = 30
            r += 1

            sub = ws2.cell(
                row=r, column=1,
                value=(f"  Generado: {hoy.strftime('%d/%m/%Y')}"
                       f"   |   Registros analizados: {len(data)}"
                       f"   |   Créditos activos: {len(credito)}"),
            )
            sub.font = Font(italic=True, size=9, color="555555")
            sub.fill = PatternFill("solid", fgColor="EBF5FB")
            ws2.merge_cells(f"A{r}:E{r}")
            r += 2

            # KPIs — numeric cells with number_format
            kpis = [
                ("Total Facturado",      total_monto,           MONEY),
                ("Total Cobrado",        total_pagado,          MONEY),
                ("Saldo por Cobrar",     total_pendiente,       MONEY),
                ("Saldo Vencido",        monto_vencido,         MONEY),
                ("N° Ventas",            len(data),             "0"),
                ("N° Créditos Activos",  len(credito),          "0"),
                ("N° Vencidas",          len(vencidas),         "0"),
                ("N° Pendientes",        len(pendientes),       "0"),
            ]
            for i, (label, value, fmt) in enumerate(kpis):
                kpi_row = r + (i // 2)
                kpi_col = 1 + (i % 2) * 2
                lc = ws2.cell(row=kpi_row, column=kpi_col, value=label)
                lc.font      = Font(bold=True, size=9, color="666666")
                lc.fill      = PatternFill("solid", fgColor="F8FBFD")
                lc.alignment = Alignment(horizontal="left")
                vc = ws2.cell(row=kpi_row, column=kpi_col + 1, value=value)
                vc.number_format = fmt
                vc.font      = Font(bold=True, size=11, color=NAVY)
                vc.fill      = PatternFill("solid", fgColor="F8FBFD")
                vc.alignment = Alignment(horizontal="right")
            r += 5  # 4 KPI rows + blank separator

            # ── Por Estado de Cobranza ────────────────────────────────────
            _section_title(ws2, r, 1, 5, "  RESUMEN POR ESTADO DE COBRANZA")
            r += 1
            for ci, h in enumerate(["Estado", "N° Ventas", "Monto Total", "Saldo Pendiente", "% Pendiente"], 1):
                _nav_hdr(ws2.cell(row=r, column=ci, value=h), TEAL)
            est_hdr = r
            r += 1
            for estado, v in sorted(by_estado.items(), key=lambda x: -x[1]['pendiente']):
                pct = v['pendiente'] / total_monto if total_monto else 0
                ws2.cell(row=r, column=1, value=estado)
                ws2.cell(row=r, column=2, value=v['count']).alignment = Alignment(horizontal="center")
                c3 = ws2.cell(row=r, column=3, value=v['monto'])
                c3.number_format = MONEY
                c4 = ws2.cell(row=r, column=4, value=v['pendiente'])
                c4.number_format = MONEY
                c5 = ws2.cell(row=r, column=5, value=pct)
                c5.number_format = '0.00%'
                r += 1
            est_end = r - 1
            _add_table(ws2, "TablaEstados", est_hdr, est_end, 5)
            # Total
            ws2.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
            ws2.cell(row=r, column=2, value=len(data)).font = Font(bold=True)
            ct3 = ws2.cell(row=r, column=3, value=total_monto)
            ct3.number_format = MONEY; ct3.font = Font(bold=True); ct3.fill = PatternFill("solid", fgColor=LIGHT)
            ct4 = ws2.cell(row=r, column=4, value=total_pendiente)
            ct4.number_format = MONEY; ct4.font = Font(bold=True); ct4.fill = PatternFill("solid", fgColor=LIGHT)
            r += 2

            # ── Antigüedad de Saldos ──────────────────────────────────────
            _section_title(ws2, r, 1, 4, "  ANTIGÜEDAD DE SALDOS (CRÉDITOS PENDIENTES)")
            r += 1
            BUCKET_ORDER = ["Por vencer", "1-30 días", "31-60 días", "61-90 días", "+90 días"]
            for ci, h in enumerate(["Rango", "N° Facturas", "Saldo Pendiente", "% del Total Pendiente"], 1):
                _nav_hdr(ws2.cell(row=r, column=ci, value=h), TEAL)
            venc_hdr = r
            r += 1
            for bucket in BUCKET_ORDER:
                if bucket not in by_venc:
                    continue
                v = by_venc[bucket]
                pct = v['monto'] / monto_pendiente_cxc if monto_pendiente_cxc else 0
                ws2.cell(row=r, column=1, value=bucket)
                ws2.cell(row=r, column=2, value=v['count']).alignment = Alignment(horizontal="center")
                b3 = ws2.cell(row=r, column=3, value=v['monto'])
                b3.number_format = MONEY
                b4 = ws2.cell(row=r, column=4, value=pct)
                b4.number_format = '0.00%'
                r += 1
            venc_end = r - 1
            _add_table(ws2, "TablaAntiguedad", venc_hdr, venc_end, 4)
            ws2.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
            ws2.cell(row=r, column=2, value=len(pendientes)).font = Font(bold=True)
            bt3 = ws2.cell(row=r, column=3, value=monto_pendiente_cxc)
            bt3.number_format = MONEY; bt3.font = Font(bold=True); bt3.fill = PatternFill("solid", fgColor=LIGHT)
            r += 2

            # ── Top Clientes por Saldo Pendiente ──────────────────────────
            _section_title(ws2, r, 1, 5, "  TOP CLIENTES POR SALDO PENDIENTE")
            r += 1
            for ci, h in enumerate(["Cliente", "N° Ventas", "Monto Total", "Saldo Pendiente", "% del Total"], 1):
                _nav_hdr(ws2.cell(row=r, column=ci, value=h), TEAL)
            cli_hdr = r
            r += 1
            top_clientes = sorted(by_cliente.items(), key=lambda x: -x[1]['pendiente'])[:15]
            for cli, v in top_clientes:
                pct = v['pendiente'] / total_monto if total_monto else 0
                ws2.cell(row=r, column=1, value=cli)
                ws2.cell(row=r, column=2, value=v['count']).alignment = Alignment(horizontal="center")
                cc3 = ws2.cell(row=r, column=3, value=v['monto'])
                cc3.number_format = MONEY
                cc4 = ws2.cell(row=r, column=4, value=v['pendiente'])
                cc4.number_format = MONEY
                cc5 = ws2.cell(row=r, column=5, value=pct)
                cc5.number_format = '0.00%'
                r += 1
            cli_end = r - 1
            _add_table(ws2, "TablaClientes", cli_hdr, cli_end, 5)

            ws2.freeze_panes = "A4"

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f'attachment; filename="ventas_{hoy.strftime("%Y%m%d")}.xlsx"'
        )
        wb.save(response)
        return response

    export_to_excel.short_description = "Exportar a Excel (.xlsx)"

    # =============================================================================
    # VISTAS PERSONALIZADAS DE REPORTES
    # =============================================================================

    def ventas_balances_admin_view(self, request):
        """Vista de análisis de balances de ventas integrada en el admin de Django."""
        import json
        from django.http import JsonResponse
        from django.template.loader import render_to_string
        from ventas.views import build_ventas_balances_context
        from app.services.filter_utils import FilterOptionsProvider

        context = build_ventas_balances_context(request)
        context.update(self.admin_site.each_context(request))
        context.update({
            'title': 'Análisis de Ventas',
            'subtitle': None,
            'opts': self.model._meta,
            'has_view_permission': self.has_view_permission(request),
            'months': FilterOptionsProvider.get_months_list(),
        })

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            table_html = render_to_string(
                'admin/ventas/partials/ventas_balances_table.html',
                context,
                request=request,
            )
            return JsonResponse({
                'kpis': {
                    'total_ventas': float(context.get('total_ventas') or 0),
                    'total_pagado': float(context.get('total_pagado') or 0),
                    'saldo_pendiente_total': float(context.get('saldo_pendiente_total') or 0),
                    'numero_transacciones': context.get('numero_transacciones', 0),
                    'venta_maxima': float(context.get('venta_maxima') or 0),
                    'promedio_ventas': float(context.get('promedio_ventas') or 0),
                },
                'chart_data': {
                    'modalidad': json.loads(context['ventas_por_modalidad']),
                    'estados': json.loads(context['ventas_por_estado']),
                    'clientes': json.loads(context['ventas_por_cliente']),
                },
                'table_html': table_html,
            })

        return TemplateResponse(request, 'admin/ventas/ventas_balances.html', context)

    def exportar_balances_xlsx_admin_view(self, request):
        """Exporta los balances filtrados a XLSX desde el admin."""
        from ventas.views import exportar_balances_xlsx
        return exportar_balances_xlsx(request)

    def reporte_cobranza_admin_view(self, request):
        """Vista del Reporte Global de Cobranza integrada en el admin de Django."""
        from ventas.views import reporte_cobranza_global
        from datetime import date
        from decimal import Decimal
        from ventas.services.reporte_cobranza_service import generar_reporte_cobranza

        hoy = date.today()
        default_inicio = date(hoy.year, 1, 1)
        default_fin = date(hoy.year, 12, 31)

        fecha_inicio_str = request.GET.get('fecha_inicio', default_inicio.isoformat())
        fecha_fin_str = request.GET.get('fecha_fin', default_fin.isoformat())
        tipo_cambio_str = request.GET.get('tipo_cambio', '')

        try:
            fecha_inicio = date.fromisoformat(fecha_inicio_str)
        except (ValueError, TypeError):
            fecha_inicio = default_inicio

        try:
            fecha_fin = date.fromisoformat(fecha_fin_str)
        except (ValueError, TypeError):
            fecha_fin = default_fin

        tipo_cambio_override = None
        if tipo_cambio_str:
            try:
                tipo_cambio_override = Decimal(tipo_cambio_str)
            except Exception:
                pass

        datos = generar_reporte_cobranza(
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            tipo_cambio_override=tipo_cambio_override,
        )

        context = {**datos}
        context.update(self.admin_site.each_context(request))
        context.update({
            'title': 'Reporte Global de Cobranza',
            'subtitle': None,
            'opts': self.model._meta,
            'has_view_permission': self.has_view_permission(request),
            'fecha_inicio_str': fecha_inicio.isoformat(),
            'fecha_fin_str': fecha_fin.isoformat(),
            'tipo_cambio_input': tipo_cambio_str,
            'hoy': hoy,
        })
        return TemplateResponse(request, 'admin/ventas/reporte_cobranza.html', context)

    def exportar_reporte_cobranza_excel(self, request):
        """Exporta el reporte de cobranza a formato Excel."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import date
        from decimal import Decimal
        from ventas.services.reporte_cobranza_service import generar_reporte_cobranza

        # Obtener parámetros
        hoy = date.today()
        default_inicio = date(hoy.year, 1, 1)
        default_fin = date(hoy.year, 12, 31)

        fecha_inicio_str = request.GET.get('fecha_inicio', default_inicio.isoformat())
        fecha_fin_str = request.GET.get('fecha_fin', default_fin.isoformat())
        tipo_cambio_str = request.GET.get('tipo_cambio', '')

        try:
            fecha_inicio = date.fromisoformat(fecha_inicio_str)
        except (ValueError, TypeError):
            fecha_inicio = default_inicio

        try:
            fecha_fin = date.fromisoformat(fecha_fin_str)
        except (ValueError, TypeError):
            fecha_fin = default_fin

        tipo_cambio_override = None
        if tipo_cambio_str:
            try:
                tipo_cambio_override = Decimal(tipo_cambio_str)
            except Exception:
                pass

        # Generar datos
        datos = generar_reporte_cobranza(
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            tipo_cambio_override=tipo_cambio_override,
        )

        # Crear workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # Estilos
        header_fill = PatternFill(start_color='FEF08A', end_color='FEF08A', fill_type='solid')
        header_font = Font(bold=True, color='1E3A8A', size=11)
        section_fill = PatternFill(start_color='FDE68A', end_color='FDE68A', fill_type='solid')
        total_fill = PatternFill(start_color='67E8F9', end_color='67E8F9', fill_type='solid')
        subtotal_fill = PatternFill(start_color='A5F3FC', end_color='A5F3FC', fill_type='solid')
        anticipo_fill = PatternFill(start_color='F0FDF4', end_color='F0FDF4', fill_type='solid')
        grand_total_fill = PatternFill(start_color='0C4A6E', end_color='0C4A6E', fill_type='solid')
        grand_total_font = Font(bold=True, color='FFFFFF', size=12)
        center_align = Alignment(horizontal='center', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        border = Border(
            left=Side(style='thin', color='9CA3AF'),
            right=Side(style='thin', color='9CA3AF'),
            top=Side(style='thin', color='9CA3AF'),
            bottom=Side(style='thin', color='9CA3AF')
        )

        # ============== SHEET 1: VENTAS X COBRAR ==============
        ws_ventas = wb.create_sheet('Ventas x Cobrar', 0)
        sucursales = datos['sucursales']
        total_anticipos = datos.get('total_anticipos', 0)
        
        # Header
        row = 1
        ws_ventas.merge_cells(f'A{row}:' + get_column_letter(len(sucursales) + 3) + f'{row}')
        cell = ws_ventas[f'A{row}']
        cell.value = f'TEMPORADA {fecha_inicio} — {fecha_fin}\nMAQUILA, SERVICIOS Y VENTAS X COBRAR'
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        
        # Column headers
        row += 1
        ws_ventas[f'A{row}'] = 'CLIENTE / CONCEPTO'
        ws_ventas[f'B{row}'] = 'MON.'
        col_idx = 3
        for suc in sucursales:
            ws_ventas.cell(row, col_idx, suc.nombre.upper()).alignment = center_align
            ws_ventas.cell(row, col_idx).fill = section_fill
            ws_ventas.cell(row, col_idx).font = header_font
            col_idx += 1
        ws_ventas.cell(row, col_idx, 'TOTAL X COBRAR').alignment = center_align
        ws_ventas.cell(row, col_idx).fill = section_fill
        ws_ventas.cell(row, col_idx).font = header_font
        
        for c in range(1, col_idx + 1):
            ws_ventas.cell(row, c).fill = section_fill
            ws_ventas.cell(row, c).font = header_font
            ws_ventas.cell(row, c).alignment = center_align
            ws_ventas.cell(row, c).border = border

        # Anticipos / Saldo a favor
        if total_anticipos > 0:
            row += 1
            ws_ventas[f'A{row}'] = 'ANTICIPOS / SALDO A FAVOR'
            ws_ventas.cell(row, col_idx, total_anticipos).number_format = '$#,##0.00'
            for c in range(1, col_idx + 1):
                ws_ventas.cell(row, c).fill = anticipo_fill
                ws_ventas.cell(row, c).border = border

        # Data rows
        for fila in datos['ventas_por_cliente']:
            row += 1
            ws_ventas[f'A{row}'] = fila['cliente'].nombre.upper()
            ws_ventas[f'B{row}'] = fila['moneda']
            ws_ventas[f'B{row}'].alignment = center_align
            
            col_idx = 3
            for suc in sucursales:
                monto = fila['por_sucursal'].get(suc.id, 0)
                if monto > 0:
                    cell = ws_ventas.cell(row, col_idx, monto)
                    cell.number_format = '$#,##0.00' if fila['moneda'] == 'MXN' else 'US$#,##0.00'
                    cell.alignment = right_align
                else:
                    ws_ventas.cell(row, col_idx, '-').alignment = center_align
                ws_ventas.cell(row, col_idx).border = border
                col_idx += 1
            
            cell = ws_ventas.cell(row, col_idx, fila['total'])
            cell.number_format = '$#,##0.00' if fila['moneda'] == 'MXN' else 'US$#,##0.00'
            cell.alignment = right_align
            cell.border = border
            cell.font = Font(bold=True)

        # Totales
        if datos.get('total_ventas_usd', 0) > 0:
            row += 1
            ws_ventas[f'A{row}'] = 'SUBTOTAL EXPORTACIÓN (USD)'
            ws_ventas[f'B{row}'] = 'USD'
            col_idx = 3
            for suc in sucursales:
                monto = datos['totales_ventas_usd']['por_sucursal'].get(suc.id, 0)
                if monto > 0:
                    ws_ventas.cell(row, col_idx, monto).number_format = 'US$#,##0.00'
                col_idx += 1
            ws_ventas.cell(row, col_idx, datos['total_ventas_usd']).number_format = 'US$#,##0.00'
            for c in range(1, col_idx + 1):
                ws_ventas.cell(row, c).fill = total_fill
                ws_ventas.cell(row, c).font = Font(bold=True)
                ws_ventas.cell(row, c).border = border

            row += 1
            ws_ventas[f'A{row}'] = f'TIPO DE CAMBIO {hoy} USD → MXN'
            ws_ventas.cell(row, 2, float(datos['tipo_cambio_ventas'])).number_format = '0.0000'
            for c in range(1, col_idx + 1):
                ws_ventas.cell(row, c).fill = subtotal_fill
                ws_ventas.cell(row, c).font = Font(bold=True)
                ws_ventas.cell(row, c).border = border

            row += 1
            ws_ventas[f'A{row}'] = 'EQUIVALENTE EN PESOS MEXICANOS'
            ws_ventas[f'B{row}'] = 'MXN'
            ws_ventas.cell(row, col_idx, datos['total_ventas_equiv_mxn']).number_format = '$#,##0.00'
            for c in range(1, col_idx + 1):
                ws_ventas.cell(row, c).fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
                ws_ventas.cell(row, c).font = Font(bold=True)
                ws_ventas.cell(row, c).border = border

        if datos.get('total_ventas_mxn_nat', 0) > 0:
            row += 1
            ws_ventas[f'A{row}'] = 'SUBTOTAL VENTAS NACIONALES'
            ws_ventas[f'B{row}'] = 'MXN'
            col_idx = 3
            for suc in sucursales:
                monto = datos['totales_ventas_mxn_obj']['por_sucursal'].get(suc.id, 0)
                if monto > 0:
                    ws_ventas.cell(row, col_idx, monto).number_format = '$#,##0.00'
                col_idx += 1
            ws_ventas.cell(row, col_idx, datos['total_ventas_mxn_nat']).number_format = '$#,##0.00'
            for c in range(1, col_idx + 1):
                ws_ventas.cell(row, c).fill = subtotal_fill
                ws_ventas.cell(row, c).font = Font(bold=True)
                ws_ventas.cell(row, c).border = border

        row += 1
        ws_ventas.merge_cells(f'A{row}:B{row}')
        ws_ventas[f'A{row}'] = 'TOTAL CARTERA A COBRAR (EQUIVALENTE MXN)'
        ws_ventas.cell(row, col_idx, datos['total_cartera_ventas_mxn']).number_format = '$#,##0.00'
        for c in range(1, col_idx + 1):
            ws_ventas.cell(row, c).fill = grand_total_fill
            ws_ventas.cell(row, c).font = grand_total_font
            ws_ventas.cell(row, c).border = border

        # Ajustar anchos
        ws_ventas.column_dimensions['A'].width = 30
        ws_ventas.column_dimensions['B'].width = 8
        for i in range(3, len(sucursales) + 4):
            ws_ventas.column_dimensions[get_column_letter(i)].width = 18

        # ============== SHEET 2: MAQUILA Y SERVICIOS X COBRAR ==============
        ws_maquila = wb.create_sheet('Maquila y servicios', 1)
        
        # Header
        row = 1
        ws_maquila.merge_cells(f'A{row}:' + get_column_letter(len(sucursales) + 2) + f'{row}')
        cell = ws_maquila[f'A{row}']
        cell.value = f'TEMPORADA {fecha_inicio} — {fecha_fin}\nMAQUILA Y SERVICIOS X COBRAR'
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        
        # Column headers
        row += 1
        ws_maquila[f'A{row}'] = 'CLIENTE / CONCEPTO'
        ws_maquila[f'B{row}'] = 'MONEDA USD'
        col_idx = 3
        for suc in sucursales:
            ws_maquila.cell(row, col_idx, suc.nombre.upper())
            ws_maquila.cell(row, col_idx).fill = section_fill
            ws_maquila.cell(row, col_idx).font = header_font
            ws_maquila.cell(row, col_idx).alignment = center_align
            col_idx += 1
        ws_maquila.cell(row, col_idx, 'TOTAL X COBRAR')
        ws_maquila.cell(row, col_idx).fill = section_fill
        ws_maquila.cell(row, col_idx).font = header_font
        ws_maquila.cell(row, col_idx).alignment = center_align
        
        for c in range(1, col_idx + 1):
            ws_maquila.cell(row, c).fill = section_fill
            ws_maquila.cell(row, c).font = header_font
            ws_maquila.cell(row, c).alignment = center_align
            ws_maquila.cell(row, c).border = border

        # Data rows
        for fila in datos['maquila_por_cliente']:
            row += 1
            ws_maquila[f'A{row}'] = fila['cliente'].nombre.upper()
            
            col_idx = 3
            for suc in sucursales:
                monto = fila['por_sucursal'].get(suc.id, 0)
                if monto > 0:
                    cell = ws_maquila.cell(row, col_idx, monto)
                    cell.number_format = '$#,##0.00'
                    cell.alignment = right_align
                else:
                    ws_maquila.cell(row, col_idx, '-').alignment = center_align
                ws_maquila.cell(row, col_idx).border = border
                col_idx += 1
            
            cell = ws_maquila.cell(row, col_idx, fila['total'])
            cell.number_format = '$#,##0.00'
            cell.alignment = right_align
            cell.font = Font(bold=True)
            cell.border = border

        # Totales
        row += 1
        ws_maquila[f'A{row}'] = 'TOTAL MAQUILA Y SERVICIOS X COBRAR'
        ws_maquila[f'B{row}'] = '$'
        col_idx = 3
        for suc in sucursales:
            monto = datos['totales_maquila']['por_sucursal'].get(suc.id, 0)
            if monto > 0:
                ws_maquila.cell(row, col_idx, monto).number_format = '$#,##0.00'
            col_idx += 1
        ws_maquila.cell(row, col_idx, datos['totales_maquila']['total']).number_format = '$#,##0.00'
        for c in range(1, col_idx + 1):
            ws_maquila.cell(row, c).fill = total_fill
            ws_maquila.cell(row, c).font = Font(bold=True)
            ws_maquila.cell(row, c).border = border

        row += 1
        ws_maquila[f'A{row}'] = f'TIPO CAMBIO HOY {hoy} PARA PAGOS'
        ws_maquila.cell(row, 2, float(datos['tipo_cambio'])).number_format = '0.0000'
        for c in range(1, col_idx + 1):
            ws_maquila.cell(row, c).fill = subtotal_fill
            ws_maquila.cell(row, c).font = Font(bold=True)
            ws_maquila.cell(row, c).border = border

        row += 1
        ws_maquila[f'A{row}'] = 'PESOS MEXICANOS'
        ws_maquila[f'B{row}'] = '$'
        ws_maquila.cell(row, col_idx, datos['totales_maquila']['total_mxn']).number_format = '$#,##0.00'
        for c in range(1, col_idx + 1):
            ws_maquila.cell(row, c).fill = PatternFill(start_color='34D399', end_color='34D399', fill_type='solid')
            ws_maquila.cell(row, c).font = Font(bold=True, color='064E3B')
            ws_maquila.cell(row, c).border = border

        # Ajustar anchos
        ws_maquila.column_dimensions['A'].width = 30
        ws_maquila.column_dimensions['B'].width = 12
        for i in range(3, len(sucursales) + 3):
            ws_maquila.column_dimensions[get_column_letter(i)].width = 18

        # ============== SHEET 3: IMPUESTOS ==============
        if datos.get('obligacion_fiscal'):
            ws_impuestos = wb.create_sheet('Impuestos a Pagar', 2)
            obligacion = datos['obligacion_fiscal']
            
            row = 1
            ws_impuestos.merge_cells(f'A{row}:B{row}')
            cell = ws_impuestos[f'A{row}']
            cell.value = f"IMPUESTOS A PAGAR — {obligacion.periodo or ''}"
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            
            row += 1
            ws_impuestos[f'A{row}'] = 'CONCEPTO'
            ws_impuestos[f'B{row}'] = 'IMPORTE'
            for c in [1, 2]:
                ws_impuestos.cell(row, c).fill = section_fill
                ws_impuestos.cell(row, c).font = header_font
                ws_impuestos.cell(row, c).alignment = center_align
                ws_impuestos.cell(row, c).border = border

            impuestos = [
                ('ISR INGRESOS PROPIOS', obligacion.isr_ingresos_propios.amount),
                ('ISR RESICO SERVICIOS PROFESIONALES', obligacion.isr_resico.amount),
                ('ISR RETENCIONES POR SALARIOS', obligacion.isr_retenciones_salarios.amount),
                ('IVA RETENCIONES SERVICIOS PROFESIONALES', obligacion.iva_retenciones_profesionales.amount),
            ]
            
            for concepto, importe in impuestos:
                row += 1
                ws_impuestos[f'A{row}'] = concepto
                ws_impuestos.cell(row, 2, float(importe)).number_format = '$#,##0.00'
                ws_impuestos.cell(row, 2).alignment = right_align
                for c in [1, 2]:
                    ws_impuestos.cell(row, c).border = border

            row += 1
            ws_impuestos[f'A{row}'] = 'TOTAL IMPUESTOS A PAGAR'
            ws_impuestos.cell(row, 2, obligacion.total_impuestos()).number_format = '$#,##0.00'
            for c in [1, 2]:
                ws_impuestos.cell(row, c).fill = total_fill
                ws_impuestos.cell(row, c).font = Font(bold=True)
                ws_impuestos.cell(row, c).alignment = right_align if c == 2 else None
                ws_impuestos.cell(row, c).border = border

            ws_impuestos.column_dimensions['A'].width = 45
            ws_impuestos.column_dimensions['B'].width = 18

        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'reporte_cobranza_{fecha_inicio}_{fecha_fin}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    def exportar_reporte_cobranza_pdf(self, request):
        """Exporta el reporte de cobranza a formato PDF."""
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT
        from datetime import date
        from decimal import Decimal
        from ventas.services.reporte_cobranza_service import generar_reporte_cobranza
        from io import BytesIO

        # Obtener parámetros
        hoy = date.today()
        default_inicio = date(hoy.year, 1, 1)
        default_fin = date(hoy.year, 12, 31)

        fecha_inicio_str = request.GET.get('fecha_inicio', default_inicio.isoformat())
        fecha_fin_str = request.GET.get('fecha_fin', default_fin.isoformat())
        tipo_cambio_str = request.GET.get('tipo_cambio', '')

        try:
            fecha_inicio = date.fromisoformat(fecha_inicio_str)
        except (ValueError, TypeError):
            fecha_inicio = default_inicio

        try:
            fecha_fin = date.fromisoformat(fecha_fin_str)
        except (ValueError, TypeError):
            fecha_fin = default_fin

        tipo_cambio_override = None
        if tipo_cambio_str:
            try:
                tipo_cambio_override = Decimal(tipo_cambio_str)
            except Exception:
                pass

        # Generar datos
        datos = generar_reporte_cobranza(
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            tipo_cambio_override=tipo_cambio_override,
        )

        # Crear PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), 
                              leftMargin=0.5*inch, rightMargin=0.5*inch,
                              topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []
        styles = getSampleStyleSheet()
        
        # Estilos personalizados
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#2f4550'),
            alignment=TA_CENTER,
            spaceAfter=12
        )
        
        section_style = ParagraphStyle(
            'SectionTitle',
            parent=styles['Heading2'],
            fontSize=12,
            textColor=colors.HexColor('#2f4550'),
            alignment=TA_CENTER,
            spaceAfter=8
        )

        sucursales = datos['sucursales']
        
        # ============== SECCIÓN 1: VENTAS X COBRAR ==============
        elements.append(Paragraph(f'REPORTE GLOBAL DE COBRANZA', title_style))
        elements.append(Paragraph(f'Período: {fecha_inicio} — {fecha_fin}', section_style))
        elements.append(Spacer(1, 0.2*inch))
        elements.append(Paragraph('MAQUILA, SERVICIOS Y VENTAS X COBRAR', section_style))
        
        # Tabla ventas
        ventas_data = [['CLIENTE', 'MON.'] + [s.nombre.upper() for s in sucursales] + ['TOTAL']]
        
        if datos.get('total_anticipos', 0) > 0:
            anticipo_row = ['ANTICIPOS / SALDO A FAVOR', ''] + [''] * len(sucursales) + [f"${datos['total_anticipos']:,.2f}"]
            ventas_data.append(anticipo_row)
        
        for fila in datos['ventas_por_cliente']:
            row = [fila['cliente'].nombre.upper()[:25], fila['moneda']]
            for suc in sucursales:
                monto = fila['por_sucursal'].get(suc.id, 0)
                prefix = 'US' if fila['moneda'] == 'USD' else ''
                row.append(f"{prefix}${monto:,.2f}" if monto > 0 else '-')
            prefix = 'US' if fila['moneda'] == 'USD' else ''
            row.append(f"{prefix}${fila['total']:,.2f}")
            ventas_data.append(row)

        # Totales
        if datos.get('total_ventas_usd', 0) > 0:
            subtotal_usd = ['SUBTOTAL EXPORTACIÓN (USD)', 'USD'] + [''] * len(sucursales) + [f"US${datos['total_ventas_usd']:,.2f}"]
            ventas_data.append(subtotal_usd)
            tc_row = [f'TIPO DE CAMBIO {hoy}', f"{datos['tipo_cambio_ventas']:.4f}"] + [''] * len(sucursales) + ['']
            ventas_data.append(tc_row)
            equiv_row = ['EQUIVALENTE MXN', 'MXN'] + [''] * len(sucursales) + [f"${datos['total_ventas_equiv_mxn']:,.2f}"]
            ventas_data.append(equiv_row)

        if datos.get('total_ventas_mxn_nat', 0) > 0:
            subtotal_mxn = ['SUBTOTAL NACIONALES', 'MXN'] + [''] * len(sucursales) + [f"${datos['total_ventas_mxn_nat']:,.2f}"]
            ventas_data.append(subtotal_mxn)

        total_row = ['TOTAL CARTERA (MXN)', ''] + [''] * len(sucursales) + [f"${datos['total_cartera_ventas_mxn']:,.2f}"]
        ventas_data.append(total_row)

        # Crear tabla
        col_widths = [2.0*inch, 0.5*inch] + [1.0*inch] * len(sucursales) + [1.2*inch]
        ventas_table = Table(ventas_data, colWidths=col_widths)
        ventas_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('rgba(201,162,39,.15)')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#2f4550')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#b8dbd9')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (-1, 1), (-1, -1), 'RIGHT'),
        ]))
        
        elements.append(ventas_table)
        elements.append(PageBreak())

        # ============== SECCIÓN 2: MAQUILA Y SERVICIOS X COBRAR ==============
        elements.append(Paragraph('MAQUILA Y SERVICIOS X COBRAR', section_style))
        elements.append(Spacer(1, 0.1*inch))
        
        maquila_data = [['CLIENTE', ''] + [s.nombre.upper() for s in sucursales] + ['TOTAL']]
        
        for fila in datos['maquila_por_cliente']:
            row = [fila['cliente'].nombre.upper()[:25], '']
            for suc in sucursales:
                monto = fila['por_sucursal'].get(suc.id, 0)
                row.append(f"${monto:,.2f}" if monto > 0 else '-')
            row.append(f"${fila['total']:,.2f}")
            maquila_data.append(row)

        total_maq = ['TOTAL MAQUILA/SERVICIOS', '$'] + [''] * len(sucursales) + [f"${datos['totales_maquila']['total']:,.2f}"]
        maquila_data.append(total_maq)
        tc_maq = [f'TIPO CAMBIO {hoy}', f"{datos['tipo_cambio']:.4f}"] + [''] * len(sucursales) + ['']
        maquila_data.append(tc_maq)
        total_mxn = ['PESOS MEXICANOS', '$'] + [''] * len(sucursales) + [f"${datos['totales_maquila']['total_mxn']:,.2f}"]
        maquila_data.append(total_mxn)

        maquila_table = Table(maquila_data, colWidths=col_widths)
        maquila_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('rgba(201,162,39,.15)')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#2f4550')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (0, -3), (-1, -1), colors.HexColor('#A5F3FC')),
            ('FONTNAME', (0, -3), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (-1, 1), (-1, -1), 'RIGHT'),
        ]))
        
        elements.append(maquila_table)

        # ============== SECCIÓN 3: IMPUESTOS ==============
        if datos.get('obligacion_fiscal'):
            elements.append(PageBreak())
            elements.append(Paragraph('IMPUESTOS A PAGAR', section_style))
            elements.append(Spacer(1, 0.1*inch))
            
            obligacion = datos['obligacion_fiscal']
            impuestos_data = [
                ['CONCEPTO', 'IMPORTE'],
                ['ISR INGRESOS PROPIOS', f"${obligacion.isr_ingresos_propios.amount:,.2f}"],
                ['ISR RESICO SERVICIOS PROFESIONALES', f"${obligacion.isr_resico.amount:,.2f}"],
                ['ISR RETENCIONES POR SALARIOS', f"${obligacion.isr_retenciones_salarios.amount:,.2f}"],
                ['IVA RETENCIONES SERVICIOS PROFESIONALES', f"${obligacion.iva_retenciones_profesionales.amount:,.2f}"],
                ['TOTAL IMPUESTOS A PAGAR', f"${obligacion.total_impuestos():,.2f}"],
            ]
            
            impuestos_table = Table(impuestos_data, colWidths=[5*inch, 2*inch])
            impuestos_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('rgba(201,162,39,.15)')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#2f4550')),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#b8dbd9')),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            
            elements.append(impuestos_table)

        # Construir PDF
        doc.build(elements)
        
        # Preparar respuesta
        buffer.seek(0)
        response = HttpResponse(buffer.read(), content_type='application/pdf')
        filename = f'reporte_cobranza_{fecha_inicio}_{fecha_fin}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def dashboard_ventas(self, request):
        """
        Dashboard principal de ventas con métricas clave.
        
        **OPTIMIZADO CON REDIS CACHE**:
        - Cache TTL: 5 minutos (300s)
        - Mejora de performance: ~3s → <0.2s (95% reducción)
        - Cache hit esperado: >85%
        - Invalidación automática en cambios de ventas/pagos
        """
        try:
            # ═══════════════════════════════════════════════════════════════
            # OBTENER DATOS DESDE CACHE (5 min TTL)
            # ═══════════════════════════════════════════════════════════════
            datos_dashboard = CuentasPorCobrarCache.get_dashboard_ventas()
            
            if datos_dashboard is None:
                # Fallback: Si el cache falla, mostrar error amigable
                messages.warning(
                    request,
                    'El dashboard está temporalmente lento. '
                    'Los datos se están recalculando.'
                )
                # Intentar cálculo directo como fallback
                from dateutil.relativedelta import relativedelta
                dso_metrics = CuentasPorCobrarMetrics.calcular_dso()
                hoy = timezone.now().date()
                inicio_mes = hoy.replace(day=1)
                inicio_año = hoy.replace(month=1, day=1)
                
                ventas_mes = Ventas.objects.filter(
                    fecha_salida_manifiesto__gte=inicio_mes
                ).aggregate(total=Sum('monto'), count=Count('id'))
                
                vencidas = Ventas.objects.filter(
                    modalidad_pago='Credito',
                    fecha_vencimiento__lt=hoy,
                    estado_cobranza__in=['Pendiente', 'Parcial']
                ).aggregate(total=Sum('monto'), count=Count('id'))
                
                datos_dashboard = {
                    'dso_metrics': dso_metrics,
                    'ventas_mes': ventas_mes,
                    'vencidas': vencidas,
                    'tasa_morosidad': 0,
                    'cartera_aging': {'corriente': 0, 'vencida_30': 0, 'vencida_60': 0, 'vencida_90': 0},
                    'recuperacion_mes_anterior': 0,
                    'top_clientes': []
                }
            
            # ═══════════════════════════════════════════════════════════════
            # CONSTRUIR CONTEXTO CON DATOS CACHEADOS
            # ═══════════════════════════════════════════════════════════════
            context = dict(
                self.admin_site.each_context(request),
                dso_metrics=datos_dashboard['dso_metrics'],
                ventas_mes=datos_dashboard['ventas_mes'],
                vencidas=datos_dashboard['vencidas'],
                top_clientes=datos_dashboard['top_clientes'],
                tasa_morosidad=datos_dashboard['tasa_morosidad'],
                cartera_aging=datos_dashboard['cartera_aging'],
                recuperacion_mes_anterior=datos_dashboard['recuperacion_mes_anterior'],
                title='Dashboard de Ventas',
            )

            return TemplateResponse(request, 'admin/ventas/dashboard.html', context)

        except Exception as e:
            messages.error(request, f'Error al generar dashboard: {str(e)}')
            return redirect('admin:ventas_ventas_changelist')
    
    def reporte_detallado_cliente(self, request, cliente_id):
        """Vista de reporte detallado por cliente desde VentasAdmin"""
        from .models import Cliente as ClienteModel
        try:
            cliente = ClienteModel.objects.get(pk=cliente_id)
        except ClienteModel.DoesNotExist:
            messages.error(request, 'Cliente no encontrado')
            return redirect('admin:ventas_ventas_changelist')
        
        ventas_totales = cliente.ventas_set.aggregate(
            total=Sum('monto'),
            count=Count('id'),
            promedio=Avg('monto')
        )
        
        ventas_por_estado = cliente.ventas_set.values('estado_cobranza').annotate(
            total=Sum('monto'),
            count=Count('id')
        ).order_by('estado_cobranza')
        
        context = dict(
            self.admin_site.each_context(request),
            cliente=cliente,
            ventas_totales=ventas_totales,
            ventas_por_estado=ventas_por_estado,
            title=f'Reporte: {cliente.nombre}'
        )
        return TemplateResponse(request, 'admin/ventas/cliente/reporte_completo.html', context)
    
    def exportar_excel_personalizado(self, request):
        """Vista de exportación con opciones avanzadas"""
        return redirect('admin:ventas_ventas_changelist')
    
    def get_estado_cobranza(self, obj):
        colors = {
            'Pagado': 'green',
            'Pendiente': 'orange', 
            'Parcial': 'blue',
            'Vencido': 'red',
            'Incobrable': 'darkred'
        }
        color = colors.get(obj.estado_cobranza, 'black')
        return format_html(
            '<span style="color: {}; font-weight: bold;">{}</span>',
            color, obj.get_estado_cobranza_display()
        )
    get_estado_cobranza.short_description = 'Estado Cobranza'
    
    def get_dias_vencimiento(self, obj):
        if obj.modalidad_pago == 'Contado':
            return '-'
        dias = obj.dias_vencido()
        if dias > 0:
            return format_html('<span style="color: red;">+{} días</span>', dias)
        elif dias < 0:
            return format_html('<span style="color: green;">{} días</span>', abs(dias))
        else:
            return 'Vence hoy'
    get_dias_vencimiento.short_description = 'Vencimiento'
    
    def get_mercado_destino(self, obj):
        return obj.mercado_destino.nombre if obj.mercado_destino else obj.tipo_venta
    get_mercado_destino.short_description = 'Mercado'

# Administración para PagoVenta
class PagoVentaResource(resources.ModelResource):
    venta = fields.Field(
        column_name='venta',
        attribute='venta',
        widget=ForeignKeyWidget(Ventas, field='carga'))
    
    cuenta_destino = fields.Field(
        column_name='cuenta_destino',
        attribute='cuenta_destino',
        widget=ForeignKeyWidget(Cuenta, field='numero_cuenta'))
    
    monto_pago = fields.Field(
        column_name='monto_pago',
        attribute='monto_pago',
        widget=MoneyWidget())

    class Meta:
        model = PagoVenta
        fields = ('id', 'fecha_pago', 'venta', 'monto_pago', 'cuenta_destino', 'metodo_pago', 'referencia')
        import_id_fields = ('id',)

@admin.register(PagoVenta)
class PagoVentaAdmin(ImportExportModelAdmin, ModelAdmin):
    """
    Admin para PagoVenta con controles bancarios de integridad.
    Implementa RF01, RF02, RF03: validaciones de nivel financiero.
    """
    resource_class = PagoVentaResource
    change_list_template = 'admin/ventas/pagoventa/change_list.html'
    list_display = ('fecha_pago', 'get_venta_info', 'monto_pago', 'metodo_pago', 'get_saldo_pendiente', 'folio_rep', 'get_comprobante', 'referencia', 'fecha_registro')
    list_select_related = ('venta', 'venta__cliente', 'documento_cfdi')
    list_filter = ('fecha_pago', 'metodo_pago', 'venta__cliente', 'venta__estado_cobranza')
    search_fields = ('venta__carga', 'venta__cliente__nombre', 'referencia', 'notas', 'folio_rep', 'uuid_rep')
    date_hierarchy = 'fecha_pago'
    readonly_fields = ('fecha_registro', 'get_saldo_venta', 'preview_comprobante')
    
    # Usar formulario personalizado con validaciones bancarias
    from .forms_banking import PagoVentaForm
    form = PagoVentaForm

    def get_urls(self):
        custom_urls = [
            path(
                '<path:object_id>/pdf/',
                self.admin_site.admin_view(self.pdf_view),
                name='ventas_pagoventa_pdf',
            ),
        ]
        return custom_urls + super().get_urls()

    @staticmethod
    def _linked_document(obj):
        try:
            return obj.documento_cfdi
        except DocumentoCFDI.DoesNotExist:
            return None

    def _pdf_source(self, obj):
        documento = self._linked_document(obj)
        if documento and documento.archivo_pdf:
            return documento.archivo_pdf, documento

        comprobante = obj.comprobante_pago
        if (
            comprobante
            and comprobante.name.lower().rsplit('.', 1)[-1] == 'pdf'
        ):
            return comprobante, None
        return None, documento

    def pdf_view(self, request, object_id):
        pago = self.get_object(request, unquote(object_id))
        if pago is None:
            raise Http404(_('No se encontró el pago solicitado.'))
        if not self.has_view_permission(request, pago):
            raise PermissionDenied

        pdf_file, _documento = self._pdf_source(pago)
        if not pdf_file:
            raise Http404(_('Este pago no tiene un PDF asociado.'))
        return _pdf_response(pdf_file)
    
    fieldsets = (
        ('Información del Pago', {
            'fields': ('venta', 'get_saldo_venta', 'fecha_pago', 'monto_pago'),
            'description': mark_safe(
                '<strong style="color:#5a7d6b;">⚠️ Controles Bancarios Activos:</strong> '
                'No se permiten pagos a ventas completadas ni sobrepagos.'
            ),
        }),
        ('Detalles Transaccionales', {
            'fields': ('metodo_pago', 'referencia', 'cuenta_destino', 'comprobante_pago', 'preview_comprobante'),
        }),
        ('Recibo Electrónico de Pago (REP)', {
            'fields': ('folio_rep', 'uuid_rep', 'banco_origen'),
            'description': mark_safe(
                'Folio y UUID del Recibo Electrónico de Pago (complemento de pago) '
                'que respalda el depósito del cliente.'
            ),
        }),
        ('Notas y Auditoría', {
            'fields': ('notas', 'fecha_registro'),
        }),
    )
    
    def get_saldo_venta(self, obj):
        """Muestra el saldo pendiente de la venta"""
        if obj and obj.venta:
            saldo = obj.venta.saldo_pendiente()
            color = '#5a7d6b' if saldo > 0 else '#586f7c'
            return format_html(
                '<strong style="color:{}">Saldo pendiente: ${}</strong>',
                color, f'{float(saldo):,.2f}'
            )
        return '-'
    get_saldo_venta.short_description = 'Saldo de la Venta'
    
    def get_venta_info(self, obj):
        """Información de la venta con indicador de estado"""
        estado_color = {
            'Pagado': '#5a7d6b',
            'Pendiente': '#c9a227',
            'Parcial': '#586f7c',
            'Vencido': '#b85450',
        }.get(obj.venta.estado_cobranza, '#586f7c')
        
        return format_html(
            '{} - {} <span style="color:{}; font-weight:600;">●</span>',
            obj.venta.carga,
            obj.venta.cliente.nombre,
            estado_color
        )
    get_venta_info.short_description = 'Venta - Cliente'
    
    def get_saldo_pendiente(self, obj):
        """Muestra el saldo pendiente después de este pago"""
        saldo = obj.venta.saldo_pendiente()
        if saldo <= 0:
            return mark_safe('<span style="color:#5a7d6b; font-weight:600;">✓ Pagado</span>')
        return format_html('<span style="color:#c9a227;">${}</span>', f'{float(saldo):,.2f}')
    get_saldo_pendiente.short_description = 'Saldo Restante'
    
    def get_comprobante(self, obj):
        """Prioriza el PDF del REP y conserva el comprobante manual."""
        pdf_file, documento = self._pdf_source(obj)
        if pdf_file:
            identifier = (
                (documento.folio or documento.uuid or str(documento.pk))
                if documento
                else (obj.folio_rep or obj.uuid_rep or str(obj.pk))
            )
            pdf_url = reverse('admin:ventas_pagoventa_pdf', args=[obj.pk])
            title = (
                _('REP %(identifier)s') % {'identifier': identifier}
                if documento
                else _('Comprobante del pago %(identifier)s') % {
                    'identifier': identifier,
                }
            )
            return _pdf_preview_button(pdf_url, title, identifier)

        comprobante_url = safe_file_url(obj.comprobante_pago)
        if comprobante_url:
            file_ext = obj.comprobante_pago.name.split('.')[-1].lower()
            if file_ext in ['jpg', 'jpeg', 'png', 'gif', 'webp']:
                icon = '🖼️'
                file_type = 'Imagen'
            else:
                icon = '📎'
                file_type = 'Archivo'
            
            return format_html(
                '<a href="{}" target="_blank" class="comprobante-link" '
                'data-file-type="{}" data-file-url="{}" '
                'onclick="return previewComprobante(event, this);" '
                'style="text-decoration:none; cursor:pointer;" '
                'title="Click para ver comprobante">'
                '{} <span style="color:#586f7c; text-decoration:underline;">{}</span>'
                '</a>',
                comprobante_url,
                file_type.lower(),
                comprobante_url,
                icon,
                file_type
            )
        return format_html(
            '<span class="cfdi-pdf-missing" title="{}">'
            '<i class="far fa-file" aria-hidden="true"></i>'
            '<span>{}</span></span>',
            _('Este pago no tiene un comprobante asociado.'),
            _('Sin archivo'),
        )
    get_comprobante.short_description = 'Comprobante'
    
    def preview_comprobante(self, obj):
        """Muestra preview del comprobante en el formulario de detalle"""
        comprobante_url = safe_file_url(obj.comprobante_pago)
        if not comprobante_url:
            return mark_safe('<p style="color:#b8dbd9;">No hay comprobante adjunto</p>')
        
        file_ext = obj.comprobante_pago.name.split('.')[-1].lower()
        
        if file_ext in ['jpg', 'jpeg', 'png', 'gif', 'webp']:
            return format_html(
                '<div style="margin-top:10px;">'
                '<a href="{}" target="_blank">'
                '<img src="{}" style="max-width:400px; max-height:400px; '
                'border:2px solid #d8dce6; border-radius:8px; cursor:pointer;" '
                'alt="Comprobante de pago" />'
                '</a>'
                '<p style="margin-top:5px; color:#586f7c; font-size:12px;">'
                'Click en la imagen para verla en tamaño completo'
                '</p>'
                '</div>',
                comprobante_url,
                comprobante_url
            )
        elif file_ext == 'pdf':
            return format_html(
                '<div style="margin-top:10px;">'
                '<a href="{}" target="_blank" '
                'style="display:inline-block; padding:10px 20px; '
                'background:#586f7c; color:white; text-decoration:none; '
                'border-radius:6px; font-weight:500;">'
                '📄 Abrir PDF en nueva pestaña'
                '</a>'
                '<p style="margin-top:10px; color:#586f7c; font-size:12px;">'
                'Archivo: {}'
                '</p>'
                '</div>',
                comprobante_url,
                obj.comprobante_pago.name.split('/')[-1]
            )
        else:
            return format_html(
                '<div style="margin-top:10px;">'
                '<a href="{}" target="_blank" '
                'style="display:inline-block; padding:10px 20px; '
                'background:#586f7c; color:white; text-decoration:none; '
                'border-radius:6px; font-weight:500;">'
                '📎 Descargar archivo'
                '</a>'
                '<p style="margin-top:10px; color:#586f7c; font-size:12px;">'
                'Archivo: {}'
                '</p>'
                '</div>',
                comprobante_url,
                obj.comprobante_pago.name.split('/')[-1]
            )
    preview_comprobante.short_description = 'Vista Previa'
    
    class Media:
        js = ('admin/js/comprobante_preview.js',)
        css = {
            'all': ('admin/css/comprobante_preview.css',)
        }
    
    def save_model(self, request, obj, form, change):
        """Override para registrar usuario en auditoría"""
        super().save_model(request, obj, form, change)
        
        # Registrar en auditoría con usuario actual
        try:
            from auditoria.models import LogActividad
            LogActividad.objects.create(
                usuario=request.user,
                nombre_usuario=request.user.username,
                tipo_accion='update' if change else 'create',
                descripcion=f'Pago de ${obj.monto_pago.amount:,.2f} registrado para venta {obj.venta.carga}',
                modelo_afectado='PagoVenta',
                objeto_id=str(obj.pk),
            )
        except Exception:
            pass
    
class AnticiposResource(resources.ModelResource):
    cliente = fields.Field(
        column_name='cliente',
        attribute='cliente',
        widget=ForeignKeyWidget(Cliente, field='nombre'))

    cuenta = fields.Field(
        column_name='cuenta',
        attribute='cuenta',
        widget=ForeignKeyWidget(Cuenta, field='numero_cuenta'))

    monto = fields.Field(
        column_name='monto',
        attribute='monto',
        widget=MoneyWidget())

    class Meta:
        model = Anticipo
        fields = ('id', 'fecha', 'cliente', 'cuenta', 'monto', 'descripcion', 'estado_anticipo')
        import_id_fields = ('id',)

    def dehydrate_cliente(self, anticipo):
        return anticipo.cliente.nombre

    def dehydrate_cuenta(self, anticipo):
        return anticipo.cuenta.numero_cuenta if anticipo.cuenta else ""

    def before_import_row(self, row, **kwargs):
        # Asigna un ID específico basado en un rango disponible
        if not row['id']:
            last_anticipo = Anticipo.objects.order_by('-id').first()
            next_id = last_anticipo.id + 1 if last_anticipo else 1
            row['id'] = next_id
     
@admin.register(Anticipo)
class AnticipoAdmin(ImportExportModelAdmin, ModelAdmin):
    resource_class = AnticiposResource
    import_form_class = ImportForm
    export_form_class = ExportForm
    list_display = (
        'fecha',
        'cliente',
        'cuenta',
        'monto',
        'get_saldo_disponible',
        'uuid_cfdi',
        'es_remanente',
        'descripcion',
        'estado_anticipo',
    )
    list_per_page = 20
    list_filter = ('fecha', 'cliente', 'cuenta', 'monto', 'estado_anticipo')
    fields = (
        'fecha',
        'cliente',
        'cuenta',
        'monto',
        'monto_aplicado',
        'uuid_cfdi',
        'es_remanente',
        'descripcion',
        'estado_anticipo',
    )
    readonly_fields = ('monto_aplicado',)
    actions = ['aplicar_anticipo_a_ventas_pendientes']
    change_list_template = 'admin/ventas/anticipo/change_list.html'

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                'importar-cfdi/',
                self.admin_site.admin_view(self.importar_anticipo_desde_cfdi),
                name='ventas_importar_anticipo_cfdi',
            ),
        ]
        return custom_urls + urls

    def get_saldo_disponible(self, obj):
        return f"${obj.saldo_disponible():,.2f}"
    get_saldo_disponible.short_description = 'Saldo disponible'

    def aplicar_anticipo_a_ventas_pendientes(self, request, queryset):
        """Aplica los anticipos seleccionados a las ventas pendientes del cliente."""
        aplicados = 0
        omitidos = 0
        for anticipo in queryset.select_related('cliente', 'cuenta'):
            try:
                with transaction.atomic():
                    anticipo_bloqueado = Anticipo.objects.select_for_update().get(pk=anticipo.pk)
                    saldo_anticipo = Decimal(str(anticipo_bloqueado.saldo_disponible()))
                    if saldo_anticipo <= Decimal('0'):
                        omitidos += 1
                        continue

                    ventas_pendientes = list(
                        Ventas.objects.select_for_update().filter(
                            cliente=anticipo_bloqueado.cliente,
                            modalidad_pago=Ventas.ModalidadPago.CREDITO,
                            estado_cobranza__in=[
                                Ventas.EstadoCobranza.PENDIENTE,
                                Ventas.EstadoCobranza.PARCIAL,
                                Ventas.EstadoCobranza.VENCIDO,
                            ],
                        ).order_by('fecha_vencimiento', 'id')
                    )
                    if not ventas_pendientes:
                        omitidos += 1
                        continue

                    monto_aplicado = Decimal('0')
                    for venta in ventas_pendientes:
                        if saldo_anticipo <= Decimal('0'):
                            break

                        saldo_venta = Decimal(str(venta.monto.amount)) - Decimal(str(venta.monto_pagado.amount))
                        if saldo_venta <= Decimal('0'):
                            continue

                        abono = min(saldo_anticipo, saldo_venta).quantize(Decimal('0.01'))
                        if abono <= Decimal('0'):
                            continue

                        PagoVenta.objects.create(
                            venta=venta,
                            fecha_pago=timezone.now().date(),
                            monto_pago=Money(abono, venta.monto.currency),
                            cuenta_destino=anticipo_bloqueado.cuenta,
                            metodo_pago=PagoVenta.MetodoPago.TRANSFERENCIA,
                            referencia=f'ANTICIPO-{anticipo_bloqueado.pk}',
                            notas='Aplicacion automatica de anticipo desde admin.',
                        )

                        if not venta.anticipo_id:
                            venta.anticipo = anticipo_bloqueado
                            venta.save(update_fields=['anticipo'])

                        saldo_anticipo -= abono
                        monto_aplicado += abono

                    if monto_aplicado <= Decimal('0'):
                        omitidos += 1
                        continue

                    anticipo_bloqueado.registrar_aplicacion(monto_aplicado)
                    anticipo_bloqueado.save(update_fields=[
                        'monto_aplicado',
                        'monto_aplicado_currency',
                        'estado_anticipo',
                    ])
                    aplicados += 1
            except Exception as exc:
                self.message_user(
                    request,
                    f"Error al aplicar anticipo #{anticipo.pk}: {exc}",
                    messages.ERROR,
                )
                omitidos += 1

        if aplicados:
            self.message_user(
                request,
                f"{aplicados} anticipo(s) aplicado(s) exitosamente.",
                messages.SUCCESS,
            )
        if omitidos:
            self.message_user(
                request,
                f"{omitidos} anticipo(s) omitido(s) (sin ventas pendientes o error).",
                messages.WARNING,
            )

    aplicar_anticipo_a_ventas_pendientes.short_description = "Aplicar anticipo a ventas pendientes del cliente"

    def importar_anticipo_desde_cfdi(self, request):
        """Importa un anticipo desde XML CFDI con confirmacion previa."""
        from .cfdi_parser import parse_cfdi

        opts = self.model._meta

        if request.method == 'POST' and request.POST.get('_step') == 'confirm':
            form = AnticipoCFDIConfirmForm(request.POST)
            if form.is_valid():
                cd = form.cleaned_data
                try:
                    anticipo = Anticipo(
                        fecha=cd['fecha'],
                        cliente=cd['cliente'],
                        cuenta=cd['cuenta'],
                        monto=Money(cd['monto'], 'MXN'),
                        descripcion=cd['descripcion'],
                        estado_anticipo=Anticipo.Estado_anticipo.Pendiente,
                    )
                    anticipo.full_clean()
                    anticipo.save()

                    DocumentoCFDI.objects.create(
                        cliente=cd['cliente'],
                        tipo=DocumentoCFDI.TipoDocumento.INGRESO,
                        subtipo=DocumentoCFDI.SubtipoDocumento.REMANENTE_ANTICIPO,
                        folio=cd['folio_factura_anticipo'] or None,
                        fecha_emision=cd['fecha'],
                        monto=Money(cd['monto'], 'MXN'),
                        moneda='MXN',
                        anticipo=anticipo,
                    )

                    messages.success(
                        request,
                        f'Anticipo importado correctamente desde CFDI: #{anticipo.pk}.',
                    )
                    change_url = reverse(
                        'admin:%s_%s_change' % (opts.app_label, opts.model_name),
                        args=[anticipo.pk],
                    )
                    return redirect(change_url)
                except Exception as exc:
                    messages.error(request, f'Error al guardar anticipo: {exc}')

            context = dict(
                self.admin_site.each_context(request),
                form=form,
                step='confirm',
                title='Importar anticipo - confirmar datos',
                opts=opts,
            )
            return TemplateResponse(request, 'admin/ventas/importar_anticipo_cfdi.html', context)

        if request.method == 'POST' and request.POST.get('_step') == 'upload':
            upload_form = AnticipoCFDIUploadForm(request.POST, request.FILES)
            if upload_form.is_valid():
                xml_bytes = upload_form.cleaned_data['xml_file'].read()
                try:
                    parsed = parse_cfdi(xml_bytes)
                except ValueError as exc:
                    messages.error(request, str(exc))
                    upload_form = AnticipoCFDIUploadForm()
                    context = dict(
                        self.admin_site.each_context(request),
                        form=upload_form,
                        step='upload',
                        title='Importar anticipo desde CFDI (XML)',
                        opts=opts,
                    )
                    return TemplateResponse(request, 'admin/ventas/importar_anticipo_cfdi.html', context)

                receptor_nombre = parsed.get('_receptor_nombre', '')
                cliente_inicial = None
                if receptor_nombre:
                    cliente_inicial = Cliente.objects.filter(
                        nombre__iexact=receptor_nombre,
                        activo=True,
                    ).first()

                monto_mxn = Decimal(str(parsed.get('monto', 0)))
                moneda = (parsed.get('moneda_venta') or 'MXN').upper()
                tipo_cambio = Decimal(str(parsed.get('tipo_cambio') or 1))
                if moneda != 'MXN' and tipo_cambio > 0:
                    monto_mxn = (monto_mxn * tipo_cambio).quantize(Decimal('0.01'))

                initial = {
                    'fecha': parsed.get('fecha_emision_cfdi') or timezone.now().date(),
                    'monto': monto_mxn,
                    'folio_factura_anticipo': parsed.get('folio_factura', ''),
                    'descripcion': parsed.get('descripcion', ''),
                    'cliente': cliente_inicial,
                }
                form = AnticipoCFDIConfirmForm(initial=initial)
                context = dict(
                    self.admin_site.each_context(request),
                    form=form,
                    step='confirm',
                    title='Importar anticipo - confirmar datos',
                    opts=opts,
                )
                return TemplateResponse(request, 'admin/ventas/importar_anticipo_cfdi.html', context)

        upload_form = AnticipoCFDIUploadForm()
        context = dict(
            self.admin_site.each_context(request),
            form=upload_form,
            step='upload',
            title='Importar anticipo desde CFDI (XML)',
            opts=opts,
        )
        return TemplateResponse(request, 'admin/ventas/importar_anticipo_cfdi.html', context)

# Administración para TerminoCredito
@admin.register(TerminoCredito)
class TerminoCreditoAdmin(ModelAdmin):
    list_display = ('nombre', 'dias_credito', 'tasa_interes_mensual', 'activo', 'fecha_creacion')
    list_filter = ('activo', 'dias_credito')
    search_fields = ('nombre',)
    list_editable = ('activo',)
    ordering = ['dias_credito']
    
    fieldsets = (
        ('Información Básica', {
            'fields': ('nombre', 'dias_credito', 'activo')
        }),
        ('Configuración Financiera', {
            'fields': ('tasa_interes_mensual', 'descripcion')
        }),
    )
    
    readonly_fields = ('fecha_creacion',)


# Administración para MercadoDestino
@admin.register(MercadoDestino)
class MercadoDestinoAdmin(ModelAdmin):
    list_display = (
        'nombre', 'get_num_paises', 'requiere_documentacion_especial', 
        'moneda_preferida', 'factor_riesgo', 'activo'
    )
    list_filter = ('activo', 'requiere_documentacion_especial', 'moneda_preferida')
    search_fields = ('nombre',)
    filter_horizontal = ('paises',)
    list_editable = ('activo', 'requiere_documentacion_especial')
    
    fieldsets = (
        ('Información Básica', {
            'fields': ('nombre', 'activo')
        }),
        ('Países Asociados', {
            'fields': ('paises',)
        }),
        ('Configuración Comercial', {
            'fields': ('moneda_preferida', 'factor_riesgo', 'requiere_documentacion_especial')
        }),
    )
    
    readonly_fields = ('fecha_creacion',)
    
    def get_num_paises(self, obj):
        count = obj.paises.count()
        return format_html(
            '<span style="font-weight: bold;">{}</span> país{}',
            count, 'es' if count != 1 else ''
        )
    get_num_paises.short_description = 'Países'


# =========================================================================
# ADMINISTRACIÓN CUENTAS POR COBRAR
# =========================================================================

from .models import SaldoCliente, AntigüedadSaldo, EstadoCuentaCliente, ConfiguracionCuentasPorCobrar

# @admin.register(SaldoCliente)  # OCULTO DE LA SIDEBAR
class SaldoClienteAdmin(ModelAdmin):
    """
    Administración para Saldos de Clientes - RF1
    Permite visualizar y gestionar los saldos pendientes por cliente.
    """
    
    list_display = (
        'get_cliente_info', 'venta', 'fecha_creacion',
        'get_monto_original', 'get_saldo_pendiente', 'get_estado_visual',
        'get_dias_vencido', 'fecha_vencimiento'
    )
    
    list_filter = (
        'estado', 'fecha_vencimiento', 'fecha_creacion',
        'venta__modalidad_pago', 'venta__tipo_venta'
    )
    
    search_fields = (
        'cliente__nombre', 'venta__carga', 'venta__PO'
    )
    
    date_hierarchy = 'fecha_vencimiento'
    list_per_page = 25
    
    readonly_fields = (
        'venta', 'cliente', 'monto_original',
        'fecha_creacion', 'fecha_ultimo_pago',
        'get_monto_pagado', 'get_dias_vencido'
    )
    
    fieldsets = (
        ('Información de la Venta', {
            'fields': ('venta', 'cliente', 'monto_original')
        }),
        ('Estado del Saldo', {
            'fields': ('saldo_pendiente', 'estado', 'fecha_vencimiento', 'get_dias_vencido')
        }),
        ('Control de Pagos', {
            'fields': ('get_monto_pagado', 'fecha_ultimo_pago'),
        }),
        ('Auditoria', {
            'fields': ('fecha_creacion',),
        })
    )
    
    actions = ['marcar_como_incobrable', 'recalcular_saldos']
    
    def get_cliente_info(self, obj):
        return format_html(
            '<strong>{}</strong><br><small>{}</small>',
            obj.cliente.nombre,
            obj.cliente.pais.nombre if obj.cliente.pais else 'Sin país'
        )
    get_cliente_info.short_description = 'Cliente'
    
    def get_monto_original(self, obj):
        return format_html(
            '<span style="font-weight: bold;">${}</span>',
            f"{float(obj.monto_original.amount):,.2f}"
        )
    get_monto_original.short_description = 'Monto Original'
    
    def get_saldo_pendiente(self, obj):
        color = 'red' if obj.saldo_pendiente.amount > 0 else 'green'
        return format_html(
            '<span style="color: {}; font-weight: bold;">${}</span>',
            color, f"{float(obj.saldo_pendiente.amount):,.2f}"
        )
    get_saldo_pendiente.short_description = 'Saldo Pendiente'
    
    def get_estado_visual(self, obj):
        colors = {
            'P': 'orange',     # Pendiente
            'PP': 'blue',      # Pago Parcial  
            'PA': 'green',     # Pagado
            'V': 'red',        # Vencido
            'I': 'darkred',    # Incobrable
            'A': 'purple'      # Anulado
        }
        color = colors.get(obj.estado, 'black')
        return format_html(
            '<span style="color: {}; font-weight: bold;">{}</span>',
            color, obj.get_estado_display()
        )
    get_estado_visual.short_description = 'Estado'
    
    def get_dias_vencido(self, obj):
        dias = obj.dias_vencido()
        if dias > 0:
            return format_html('<span style="color: red; font-weight: bold;">+{}</span>', dias)
        elif dias < 0:
            return format_html('<span style="color: green;">{}</span>', dias)
        else:
            return 'Vence hoy'
    get_dias_vencido.short_description = 'Días Vencido'
    
    def get_monto_pagado(self, obj):
        pagado = obj.monto_original - obj.saldo_pendiente 
        return f"${float(pagado.amount):,.2f}"
    get_monto_pagado.short_description = 'Monto Pagado'
    
    def marcar_como_incobrable(self, request, queryset):
        """Acción para marcar saldos como incobrables"""
        count = 0
        for saldo in queryset.filter(estado__in=['P', 'PP', 'V']):
            saldo.estado = SaldoCliente.EstadosSaldo.INCOBRABLE
            saldo.save()
            count += 1
        
        self.message_user(
            request, 
            f'{count} saldo(s) marcado(s) como incobrable(s).'
        )
    marcar_como_incobrable.short_description = "Marcar como incobrable"
    
    def recalcular_saldos(self, request, queryset):
        """Acción para recalcular saldos basado en pagos"""
        from .services.cuentas_por_cobrar_service import CuentasPorCobrarService
        
        count = 0
        for saldo in queryset:
            try:
                CuentasPorCobrarService.sincronizar_deuda_venta(saldo.venta.id)
                count += 1
            except Exception:
                pass
        
        self.message_user(
            request,
            f'{count} saldo(s) recalculado(s).'
        )
    recalcular_saldos.short_description = "Recalcular saldos"


# @admin.register(AntigüedadSaldo)  # OCULTO DE LA SIDEBAR
class AntigüedadSaldoAdmin(ModelAdmin):
    """
    Administración para Análisis de Antigüedad - RF3
    Permite visualizar el aging de cartera por cliente.
    """
    
    list_display = (
        'get_cliente_info', 'fecha_calculo', 'get_total_saldo',
        'get_corriente', 'get_vencido_1', 'get_vencido_2', 'get_vencido_3',
        'get_porcentaje_critico'
    )
    
    list_filter = ('fecha_calculo',)
    search_fields = ('cliente__nombre',)
    date_hierarchy = 'fecha_calculo'
    list_per_page = 50
    
    readonly_fields = (
        'fecha_calculo', 'cliente', 'total_saldo', 'corriente', 
        'vencido_1', 'vencido_2', 'vencido_3', 'calculado_por'
    )
    
    fieldsets = (
        ('Cliente y Fecha', {
            'fields': ('cliente', 'fecha_calculo')
        }),
        ('Distribución por Antigüedad', {
            'fields': ('total_saldo', 'corriente', 'vencido_1', 'vencido_2', 'vencido_3')
        }),
        ('Auditoria', {
            'fields': ('calculado_por',),
        })
    )
    
    actions = ['exportar_aging_excel']
    
    def get_cliente_info(self, obj):
        return format_html(
            '<strong>{}</strong><br><small>ID: {}</small>',
            obj.cliente.nombre, obj.cliente.id
        )
    get_cliente_info.short_description = 'Cliente'
    
    def get_total_saldo(self, obj):
        return format_html(
            '<span style="font-weight: bold; color: navy;">${}</span>',
            f"{float(obj.total_saldo.amount):,.2f}"
        )
    get_total_saldo.short_description = 'Total Saldo'
    
    def get_corriente(self, obj):
        return format_html(
            '<span style="color: green;">${}</span>',
            f"{float(obj.corriente.amount):,.2f}"
        )
    get_corriente.short_description = 'Corriente'
    
    def get_vencido_1(self, obj):
        return format_html(
            '<span style="color: orange;">${}</span>',
            f"{float(obj.vencido_1.amount):,.2f}"
        )
    get_vencido_1.short_description = '1-30 días'
    
    def get_vencido_2(self, obj):
        return format_html(
            '<span style="color: red;">${}</span>',
            f"{float(obj.vencido_2.amount):,.2f}"
        )
    get_vencido_2.short_description = '31-60 días'
    
    def get_vencido_3(self, obj):
        return format_html(
            '<span style="color: darkred; font-weight: bold;">${}</span>',
            f"{float(obj.vencido_3.amount):,.2f}"
        )
    get_vencido_3.short_description = '+60 días'
    
    def get_porcentaje_critico(self, obj):
        """Muestra porcentaje de cartera crítica (>60 días)"""
        if obj.total_saldo.amount > 0:
            pct = (obj.vencido_3.amount / obj.total_saldo.amount) * 100
            color = 'red' if pct > 25 else 'orange' if pct > 10 else 'green'
            return format_html(
                '<span style="color: {}; font-weight: bold;">{:.1f}%</span>',
                color, pct
            )
        return '0%'
    get_porcentaje_critico.short_description = '% Crítico'
    
    def exportar_aging_excel(self, request, queryset):
        """Exporta análisis de aging a Excel"""
        # Esta funcionalidad se implementaría con openpyxl
        self.message_user(request, "Funcionalidad de exportación en desarrollo")
    exportar_aging_excel.short_description = "Exportar a Excel"


@admin.register(EstadoCuentaCliente)  
class EstadoCuentaClienteAdmin(ModelAdmin):
    """
    Administración para Estados de Cuenta - RF4
    Permite visualizar el historial completo por cliente.
    """
    
    list_display = (
        'get_cliente_info', 'fecha_generacion',
        'periodo_inicio', 'periodo_fin', 'get_saldo_inicial',
        'get_total_ventas', 'get_total_abonos', 'get_saldo_final', 'formato_generado'
    )
    
    list_filter = ('fecha_generacion', 'periodo_inicio', 'periodo_fin', 'formato_generado')
    search_fields = ('cliente__nombre', 'generado_por')
    date_hierarchy = 'periodo_inicio'
    list_per_page = 20
    
    readonly_fields = (
        'fecha_generacion', 'saldo_inicial', 'total_ventas', 'total_abonos',
        'saldo_final', 'numero_facturas'
    )
    
    fieldsets = (
        ('Información del Reporte', {
            'fields': ('cliente', 'fecha_generacion', 'generado_por')
        }),
        ('Período de Análisis', {
            'fields': ('periodo_inicio', 'periodo_fin')
        }),
        ('Resumen Financiero', {
            'fields': ('saldo_inicial', 'total_ventas', 'total_abonos', 'saldo_final', 'numero_facturas'),
            'description': 'Saldo Final = Saldo Inicial + Total Ventas − Total Abonos'
        }),
        ('Archivo', {
            'fields': ('formato_generado', 'archivo_generado', 'notas'),
        })
    )
    
    actions = ['regenerar_reporte', 'enviar_por_email']
    
    def get_cliente_info(self, obj):
        return format_html(
            '<strong>{}</strong><br><small>{}</small>',
            obj.cliente.nombre,
            obj.cliente.correo if obj.cliente.correo else 'Sin email'
        )
    get_cliente_info.short_description = 'Cliente'
    
    def get_saldo_inicial(self, obj):
        if obj.saldo_inicial.amount == 0:
            return format_html('<span style="color: gray;">$0.00</span>')
        color = 'darkorange' if obj.saldo_inicial.amount > 0 else 'green'
        return format_html(
            '<span style="color: {}; font-weight: bold;">${}</span>',
            color, f"{float(obj.saldo_inicial.amount):,.2f}"
        )
    get_saldo_inicial.short_description = 'Saldo Inicial'

    def get_total_ventas(self, obj):
        return format_html(
            '<span style="color: navy; font-weight: bold;">${}</span>',
            f"{float(obj.total_ventas.amount):,.2f}"
        )
    get_total_ventas.short_description = 'Total Ventas'

    def get_total_abonos(self, obj):
        return format_html(
            '<span style="color: green;">${}</span>',
            f"{float(obj.total_abonos.amount):,.2f}"
        )
    get_total_abonos.short_description = 'Total Abonos'
    
    def get_saldo_final(self, obj):
        color = 'red' if obj.saldo_final.amount > 0 else 'green'
        return format_html(
            '<span style="color: {}; font-weight: bold;">${}</span>',
            color, f"{float(obj.saldo_final.amount):,.2f}"
        )
    get_saldo_final.short_description = 'Saldo Final'
    
    def regenerar_reporte(self, request, queryset):
        """Regenera reportes de estado de cuenta"""
        from .services.cuentas_por_cobrar_service import CuentasPorCobrarService
        
        count = 0
        for estado in queryset:
            try:
                CuentasPorCobrarService.generar_estado_cuenta(
                    estado.cliente.id,
                    estado.periodo_inicio,
                    estado.periodo_fin
                )
                count += 1
            except Exception:
                pass
        
        self.message_user(request, f'{count} reporte(s) regenerado(s)')
    regenerar_reporte.short_description = "Regenerar reportes"
    
    def enviar_por_email(self, request, queryset):
        """Envía estados de cuenta por email"""
        # Funcionalidad para implementar con sistema de emails
        self.message_user(request, "Funcionalidad de envío por email en desarrollo")
    enviar_por_email.short_description = "Enviar por email"


# @admin.register(ConfiguracionCuentasPorCobrar)  # OCULTO DE LA SIDEBAR
class ConfiguracionCuentasPorCobrarAdmin(ModelAdmin):
    """
    Administración para Configuración del Sistema CxC
    Permite configurar parámetros globales del sistema.
    """
    
    list_display = (
        'get_config_str', 'dias_corriente', 'calculo_automatico_aging',
        'frecuencia_calculo', 'enviar_alertas_vencimiento', 'fecha_creacion'
    )
    
    list_filter = ('calculo_automatico_aging', 'frecuencia_calculo', 'enviar_alertas_vencimiento')
    search_fields = ('email_responsable_cobranza',)
    
    fieldsets = (
        ('Parámetros de Antigüedad (Aging)', {
            'fields': ('dias_corriente', 'dias_vencido_1', 'dias_vencido_2')
        }),
        ('Automatización', {
            'fields': ('calculo_automatico_aging', 'hora_calculo_aging', 'frecuencia_calculo')
        }),
        ('Tipo de Cambio', {
            'fields': ('tipo_cambio_usd',),
            'description': 'Tipo de cambio USD→MXN vigente. Actualizar diariamente.'
        }),
        ('Alertas y Notificaciones', {
            'fields': ('enviar_alertas_vencimiento', 'dias_previos_alerta', 'email_responsable_cobranza')
        }),
        ('Límites de Crédito', {
            'fields': ('permitir_sobregiro_credito', 'porcentaje_sobregiro_permitido'),
        }),
        ('Auditoria', {
            'fields': ('fecha_creacion', 'fecha_modificacion'),
        })
    )
    
    readonly_fields = ('fecha_creacion', 'fecha_modificacion')

    def get_config_str(self, obj):
        return str(obj)
    get_config_str.short_description = 'Configuración'

# Administración para MercadoDestino
class PaisInline(admin.TabularInline):
    model = MercadoDestino.paises.through
list_editable = ('activo',)
inlines = [PaisInline]
exclude = ('paises',)

# Inline para PagoVenta - debe definirse antes de VentasAdmin
class PagoVentaInline(admin.TabularInline):
    model = PagoVenta
    extra = 0
    readonly_fields = ('fecha_registro',)
    fields = ('fecha_pago', 'monto_pago', 'cuenta_destino', 'metodo_pago', 'referencia', 'notas')


# =============================================================================
# OBLIGACIONES FISCALES
# =============================================================================

@admin.register(ObligacionFiscal)
class ObligacionFiscalAdmin(ModelAdmin):
    list_display = ('periodo', 'isr_ingresos_propios', 'isr_resico',
                    'isr_retenciones_salarios', 'iva_retenciones_profesionales',
                    'get_total', 'fecha_registro')
    list_filter = ('fecha_registro',)
    search_fields = ('periodo',)
    readonly_fields = ('fecha_registro',)

    fieldsets = (
        ('Período', {
            'fields': ('periodo', 'fecha_registro')
        }),
        ('Impuestos', {
            'fields': (
                'isr_ingresos_propios',
                'isr_resico',
                'isr_retenciones_salarios',
                'iva_retenciones_profesionales',
            )
        }),
    )

    def get_total(self, obj):
        total = float(obj.total_impuestos())
        return format_html('<strong>${}</strong>', f'{total:,.2f}')
    get_total.short_description = 'Total Impuestos'


# =============================================================================
# DOCUMENTOS CFDI
# =============================================================================

@admin.register(DocumentoCFDI)
class DocumentoCFDIAdmin(ModelAdmin):
    change_list_template = 'admin/ventas/documentocfdi/change_list.html'
    list_display = (
        'subtipo', 'folio', 'uuid', 'get_cliente', 'fecha_emision', 'monto',
        'estado', 'pdf_preview',
    )
    list_filter = ('tipo', 'subtipo', 'estado', 'fecha_emision')
    search_fields = ('folio', 'uuid', 'cliente__nombre', 'venta__carga')
    date_hierarchy = 'fecha_emision'
    list_per_page = 30
    list_select_related = ('cliente',)
    readonly_fields = ('creado_en', 'actualizado_en')

    def get_urls(self):
        custom_urls = [
            path(
                '<path:object_id>/pdf/',
                self.admin_site.admin_view(self.pdf_view),
                name='ventas_documentocfdi_pdf',
            ),
        ]
        return custom_urls + super().get_urls()

    def pdf_view(self, request, object_id):
        """Redirect an authorized user to a fresh local or R2 PDF URL."""
        documento = self.get_object(request, unquote(object_id))
        if documento is None or not documento.archivo_pdf:
            raise Http404(_('Este CFDI no tiene un PDF asociado.'))
        if not self.has_view_permission(request, documento):
            raise PermissionDenied
        return _pdf_response(documento.archivo_pdf)

    def pdf_preview(self, obj):
        if not obj.archivo_pdf:
            return format_html(
                '<span class="cfdi-pdf-missing" title="{}">'
                '<i class="far fa-file-pdf" aria-hidden="true"></i>'
                '<span>{}</span></span>',
                _('Este CFDI no tiene un PDF asociado.'),
                _('Sin PDF'),
            )

        identifier = obj.folio or obj.uuid or str(obj.pk)
        pdf_url = reverse('admin:ventas_documentocfdi_pdf', args=[obj.pk])
        return _pdf_preview_button(
            pdf_url,
            _('CFDI %(identifier)s') % {'identifier': identifier},
            identifier,
        )
    pdf_preview.short_description = _('Documento')

    fieldsets = (
        ('Clasificación', {
            'fields': ('tipo', 'subtipo', 'estado')
        }),
        ('Identificación SAT', {
            'fields': ('serie', 'folio', 'uuid', 'fecha_emision', 'fecha_timbrado')
        }),
        ('Montos', {
            'fields': ('monto', 'moneda', 'tipo_cambio')
        }),
        ('Relaciones de negocio', {
            'fields': ('cliente', 'venta', 'anticipo', 'pago_venta')
        }),
        ('Relación fiscal', {
            'fields': ('cfdi_relacionado', 'tipo_relacion')
        }),
        ('Archivos', {
            'fields': ('archivo_xml', 'archivo_pdf')
        }),
        ('Auditoría', {
            'fields': ('creado_en', 'actualizado_en')
        }),
    )

    def get_cliente(self, obj):
        return obj.cliente.nombre
    get_cliente.short_description = 'Cliente'
