from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required, permission_required
from django.db.models import Sum, Count, Avg, Max, Min, Q
from django.db.models.functions import Extract, TruncMonth, TruncWeek, TruncDate
from django.utils import timezone
from datetime import datetime, timedelta, date
from decimal import Decimal
from collections import defaultdict, OrderedDict
import json
from .models import Anticipo, Ventas, Cliente, TerminoCredito, MercadoDestino, PagoVenta
from .forms import AnticipoForm
from catalogo.models import Sucursal, Pais
from gastos.models import Cuenta
from .services.reporte_cobranza_service import generar_reporte_cobranza

@login_required
@permission_required('ventas.view_anticipo', raise_exception=True)
def lista_anticipos(request):
    anticipos = Anticipo.objects.all()
    return render(request, 'lista_anticipos.html', {'anticipos': anticipos})

@login_required
@permission_required('ventas.add_anticipo', raise_exception=True)
def crear_anticipo(request):
    if request.method == 'POST':
        form = AnticipoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('lista_anticipos')
    else:
        form = AnticipoForm()
    return render(request, 'crear_anticipo.html', {'form': form})

@login_required
@permission_required('ventas.view_ventas', raise_exception=True)
def detalle_venta(request, venta_id):
    venta = get_object_or_404(Ventas, id=venta_id)
    monto_final = venta.calcular_monto_final()
    return render(request, 'detalle_venta.html', {'venta': venta, 'monto_final': monto_final})

def build_ventas_balances_context(request):
    """Construye y retorna el contexto para la vista de balances de ventas."""    # Obtener parámetros de filtro
    selected_cliente_id = request.GET.get('cliente_id', '')
    selected_cuenta_id = request.GET.get('cuenta_id', '')
    selected_sucursal_id = request.GET.get('sucursal_id', '')
    selected_mercado_id = request.GET.get('mercado_id', '')
    selected_modalidad = request.GET.get('modalidad_pago', '')
    selected_estado = request.GET.get('estado_cobranza', '')
    selected_year = request.GET.get('year', str(timezone.now().year))
    selected_months = request.GET.getlist('months')
    selected_periodo = request.GET.get('periodo', 'mensual')
    
    # Filtros de fecha específicos para período diario
    selected_dia = request.GET.get('dia', '')
    selected_fecha_inicio = request.GET.get('fecha_inicio', '')
    selected_fecha_fin = request.GET.get('fecha_fin', '')
    tipo_fecha = request.GET.get('tipo_fecha', 'dia')
    
    # Obtener listas para los filtros
    clientes = Cliente.objects.filter(activo=True).order_by('nombre')
    cuentas = Cuenta.objects.all().order_by('numero_cuenta')
    sucursales = Sucursal.objects.all().order_by('nombre')
    mercados = MercadoDestino.objects.filter(activo=True).order_by('nombre')
    
    # Años disponibles basados en las ventas registradas
    available_years = Ventas.objects.dates('fecha_registro', 'year').order_by('-fecha_registro')
    current_year = timezone.now().year
    
    # Construir filtros base
    filters = Q()
    
    # Filtros por selecciones
    if selected_cliente_id:
        filters &= Q(cliente_id=selected_cliente_id)
    if selected_cuenta_id:
        filters &= Q(cuenta_id=selected_cuenta_id)
    if selected_sucursal_id:
        filters &= Q(sucursal_id=selected_sucursal_id)
    if selected_mercado_id:
        filters &= Q(mercado_destino_id=selected_mercado_id)
    if selected_modalidad:
        filters &= Q(modalidad_pago=selected_modalidad)
    if selected_estado:
        filters &= Q(estado_cobranza=selected_estado)
    
    # Filtros de fecha según el período seleccionado
    if selected_periodo == 'diario':
        if tipo_fecha == 'dia' and selected_dia:
            try:
                fecha_dia = datetime.strptime(selected_dia, '%Y-%m-%d').date()
                filters &= Q(fecha_deposito=fecha_dia)
            except ValueError:
                pass
        elif tipo_fecha == 'rango' and selected_fecha_inicio and selected_fecha_fin:
            try:
                fecha_inicio = datetime.strptime(selected_fecha_inicio, '%Y-%m-%d').date()
                fecha_fin = datetime.strptime(selected_fecha_fin, '%Y-%m-%d').date()
                filters &= Q(fecha_deposito__range=[fecha_inicio, fecha_fin])
            except ValueError:
                pass
    else:
        # Filtro por año
        if selected_year:
            filters &= Q(fecha_deposito__year=int(selected_year))
        
        # Filtro por meses
        if selected_months and len(selected_months) > 0:
            try:
                months_int = [int(m) for m in selected_months if m.isdigit()]
                if months_int:
                    filters &= Q(fecha_deposito__month__in=months_int)
            except ValueError:
                pass
    
    # Obtener ventas filtradas
    ventas_queryset = Ventas.objects.filter(filters).select_related(
        'cliente', 'cuenta', 'sucursal_id', 'mercado_destino', 'producto', 'agente_id'
    ).prefetch_related('pagos')
    
    # Preparar datos para la tabla según el período
    balances = []
    
    if selected_periodo == 'diario':
        # Agrupar por día
        ventas_agrupadas = ventas_queryset.extra(
            select={'fecha_grupo': 'DATE(fecha_deposito)'}
        ).values(
            'fecha_grupo', 'cliente__nombre', 'cuenta__numero_cuenta', 
            'cuenta__id_banco__nombre', 'sucursal_id__nombre'
        ).annotate(
            total_ventas=Sum('monto'),
            total_pagado=Sum('monto_pagado'),
            estado_cobranza_principal=Max('estado_cobranza'),
            fecha_vencimiento_proxima=Min('fecha_vencimiento'),
            venta_maxima=Max('monto'),
            venta_minima=Min('monto'),
            venta_promedio=Avg('monto')
        ).order_by('fecha_grupo', 'cliente__nombre')
        
        acumulado = 0
        for idx, grupo in enumerate(ventas_agrupadas):
            total = float(grupo['total_ventas'] or 0)
            acumulado += total
            
            _pagado = float(grupo['total_pagado'] or 0)
            _venc   = grupo.get('fecha_vencimiento_proxima')
            balances.append({
                'numero_secuencial': idx + 1,
                'cliente_nombre': grupo['cliente__nombre'],
                'cuenta_numero': grupo['cuenta__numero_cuenta'],
                'banco_nombre': grupo['cuenta__id_banco__nombre'] or 'N/A',
                'sucursal_nombre': grupo['sucursal_id__nombre'],
                'fecha': grupo['fecha_grupo'],
                'total_ventas': total,
                'total_pagado': _pagado,
                'saldo_pendiente': total - _pagado,
                'estado_cobranza': Ventas.derive_estado_desde_totales(total, _pagado, _venc),
                'fecha_vencimiento': _venc,
                'venta_maxima': float(grupo['venta_maxima'] or 0),
                'venta_minima': float(grupo['venta_minima'] or 0),
                'venta_promedio': float(grupo['venta_promedio'] or 0),
                'acumulado': acumulado
            })
    
    elif selected_periodo == 'semanal':
        # Agrupar por semana
        ventas_agrupadas = ventas_queryset.annotate(
            fecha_grupo=TruncWeek('fecha_deposito')
        ).values(
            'fecha_grupo', 'cliente__nombre', 'cuenta__numero_cuenta',
            'cuenta__id_banco__nombre', 'sucursal_id__nombre'
        ).annotate(
            total_ventas=Sum('monto'),
            total_pagado=Sum('monto_pagado'),
            numero_transacciones=Count('id'),
            venta_maxima=Max('monto'),
            venta_minima=Min('monto'),
            venta_promedio=Avg('monto'),
            fecha_vencimiento_proxima=Min('fecha_vencimiento')
        ).order_by('fecha_grupo', 'cliente__nombre')
        
        acumulado = 0
        for idx, grupo in enumerate(ventas_agrupadas):
            total = float(grupo['total_ventas'] or 0)
            acumulado += total
            
            _pagado = float(grupo['total_pagado'] or 0)
            _venc   = grupo.get('fecha_vencimiento_proxima')
            balances.append({
                'numero_secuencial': idx + 1,
                'cliente_nombre': grupo['cliente__nombre'],
                'cuenta_numero': grupo['cuenta__numero_cuenta'],
                'banco_nombre': grupo['cuenta__id_banco__nombre'] or 'N/A',
                'sucursal_nombre': grupo['sucursal_id__nombre'],
                'fecha': grupo['fecha_grupo'],
                'total_ventas': total,
                'total_pagado': _pagado,
                'saldo_pendiente': total - _pagado,
                'estado_cobranza': Ventas.derive_estado_desde_totales(total, _pagado, _venc),
                'fecha_vencimiento': _venc,
                'venta_maxima': float(grupo['venta_maxima'] or 0),
                'venta_minima': float(grupo['venta_minima'] or 0),
                'venta_promedio': float(grupo['venta_promedio'] or 0),
                'acumulado': acumulado
            })
    
    else:  # mensual
        # Agrupar por mes
        ventas_agrupadas = ventas_queryset.annotate(
            fecha_grupo=TruncMonth('fecha_deposito')
        ).values(
            'fecha_grupo', 'cliente__nombre', 'cuenta__numero_cuenta',
            'cuenta__id_banco__nombre', 'sucursal_id__nombre'
        ).annotate(
            total_ventas=Sum('monto'),
            total_pagado=Sum('monto_pagado'),
            numero_transacciones=Count('id'),
            venta_maxima=Max('monto'),
            venta_minima=Min('monto'),
            venta_promedio=Avg('monto'),
            fecha_vencimiento_proxima=Min('fecha_vencimiento')
        ).order_by('fecha_grupo', 'cliente__nombre')
        
        acumulado = 0
        for idx, grupo in enumerate(ventas_agrupadas):
            total = float(grupo['total_ventas'] or 0)
            acumulado += total
            
            _pagado = float(grupo['total_pagado'] or 0)
            _venc   = grupo.get('fecha_vencimiento_proxima')
            balances.append({
                'numero_secuencial': idx + 1,
                'cliente_nombre': grupo['cliente__nombre'],
                'cuenta_numero': grupo['cuenta__numero_cuenta'],
                'banco_nombre': grupo['cuenta__id_banco__nombre'] or 'N/A',
                'sucursal_nombre': grupo['sucursal_id__nombre'],
                'fecha': grupo['fecha_grupo'],
                'total_ventas': total,
                'total_pagado': _pagado,
                'saldo_pendiente': total - _pagado,
                'estado_cobranza': Ventas.derive_estado_desde_totales(total, _pagado, _venc),
                'fecha_vencimiento': _venc,
                'venta_maxima': float(grupo['venta_maxima'] or 0),
                'venta_minima': float(grupo['venta_minima'] or 0),
                'venta_promedio': float(grupo['venta_promedio'] or 0),
                'acumulado': acumulado
            })
    
    # Calcular métricas generales
    totales = ventas_queryset.aggregate(
        total_ventas=Sum('monto'),
        total_pagado=Sum('monto_pagado'),
        numero_transacciones=Count('id'),
        venta_maxima=Max('monto'),
        venta_minima=Min('monto'),
        venta_promedio=Avg('monto')
    )
    
    total_ventas = float(totales['total_ventas'] or 0)
    total_pagado = float(totales['total_pagado'] or 0)
    saldo_pendiente_total = total_ventas - total_pagado
    numero_transacciones = totales['numero_transacciones'] or 0
    venta_maxima = float(totales['venta_maxima'] or 0)
    venta_minima = float(totales['venta_minima'] or 0)
    promedio_ventas = float(totales['venta_promedio'] or 0)
    
    # Calcular ventas por modalidad de pago para gráficos
    ventas_por_modalidad = ventas_queryset.values('modalidad_pago').annotate(
        total=Sum('monto')
    ).order_by('-total')
    
    # Convertir Decimal a float para JavaScript
    ventas_por_modalidad = [
        {'modalidad_pago': item['modalidad_pago'], 'total': float(item['total'] or 0)}
        for item in ventas_por_modalidad
    ]
    
    # Calcular ventas por estado de cobranza para gráficos
    ventas_por_estado = ventas_queryset.values('estado_cobranza').annotate(
        total=Sum('monto')
    ).order_by('-total')
    
    # Convertir Decimal a float para JavaScript
    ventas_por_estado = [
        {'estado_cobranza': item['estado_cobranza'], 'total': float(item['total'] or 0)}
        for item in ventas_por_estado
    ]
    
    # Calcular ventas por cliente para gráficos
    ventas_por_cliente = ventas_queryset.values('cliente__nombre').annotate(
        total=Sum('monto')
    ).order_by('-total')[:10]  # Top 10 clientes
    
    # Convertir Decimal a float para JavaScript
    ventas_por_cliente = [
        {'cliente__nombre': item['cliente__nombre'], 'total': float(item['total'] or 0)}
        for item in ventas_por_cliente
    ]
    
    # Calcular métricas adicionales
    ventas_contado = ventas_queryset.filter(modalidad_pago='Contado').aggregate(
        total=Sum('monto'), count=Count('id')
    )
    ventas_credito = ventas_queryset.filter(modalidad_pago='Credito').aggregate(
        total=Sum('monto'), count=Count('id')
    )
    
    ventas_vencidas = ventas_queryset.filter(
        estado_cobranza='Vencido'
    ).aggregate(total=Sum('monto'), count=Count('id'))
    
    context = {
        # Datos para filtros
        'clientes': clientes,
        'cuentas': cuentas,
        'sucursales': sucursales,
        'mercados': mercados,
        'available_years': available_years,
        'current_year': current_year,
        
        # Valores seleccionados en filtros
        'selected_cliente_id': selected_cliente_id,
        'selected_cuenta_id': selected_cuenta_id,
        'selected_sucursal_id': selected_sucursal_id,
        'selected_mercado_id': selected_mercado_id,
        'selected_modalidad': selected_modalidad,
        'selected_estado': selected_estado,
        'selected_year': selected_year,
        'selected_months': selected_months,
        'selected_periodo': selected_periodo,
        'selected_dia': selected_dia,
        'selected_fecha_inicio': selected_fecha_inicio,
        'selected_fecha_fin': selected_fecha_fin,
        'tipo_fecha': tipo_fecha,
        
        # Datos para la tabla
        'balances': balances,
        
        # Métricas generales
        'total_ventas': total_ventas,
        'total_pagado': total_pagado,
        'saldo_pendiente_total': saldo_pendiente_total,
        'numero_transacciones': numero_transacciones,
        'venta_maxima': venta_maxima,
        'venta_minima': venta_minima,
        'promedio_ventas': promedio_ventas,
        
        # Métricas por modalidad
        'ventas_contado_total': float(ventas_contado['total'] or 0),
        'ventas_contado_count': ventas_contado['count'] or 0,
        'ventas_credito_total': float(ventas_credito['total'] or 0),
        'ventas_credito_count': ventas_credito['count'] or 0,
        
        # Métricas de cobranza
        'ventas_vencidas_total': float(ventas_vencidas['total'] or 0),
        'ventas_vencidas_count': ventas_vencidas['count'] or 0,
        
        # Datos para gráficos (serializados como JSON)
        'ventas_por_modalidad': json.dumps(ventas_por_modalidad),
        'ventas_por_estado': json.dumps(ventas_por_estado),
        'ventas_por_cliente': json.dumps(ventas_por_cliente),
        
        # Opciones para filtros
        'modalidad_choices': Ventas.ModalidadPago.choices,
        'estado_choices': Ventas.EstadoCobranza.choices,
    }
    
    return context


@login_required
@permission_required('ventas.view_ventas', raise_exception=True)
def ventas_balances(request):
    """Vista para mostrar balances y análisis de ventas con filtros avanzados"""
    context = build_ventas_balances_context(request)
    return render(request, 'ventas/ventas_balances.html', context)


@login_required
@permission_required('ventas.view_ventas', raise_exception=True)
def exportar_balances_xlsx(request):
    """Exporta los balances filtrados actuales a un archivo XLSX."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    context = build_ventas_balances_context(request)
    balances = context['balances']

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balances de Ventas"

    # Encabezados
    headers = [
        '#', 'Cliente', 'Cuenta', 'Banco', 'Sucursal', 'Fecha',
        'Total Ventas', 'Total Pagado', 'Saldo Pendiente',
        'Estado Cobranza', 'Fecha Vencimiento', 'Acumulado',
    ]
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Filas de datos
    estado_colores = {
        'Pagado': 'C6EFCE',
        'Parcial': 'FFEB9C',
        'Pendiente': 'FFEB9C',
        'Vencido': 'FFC7CE',
    }

    for row_num, balance in enumerate(balances, start=2):
        estado = balance.get('estado_cobranza', '')
        row_fill = PatternFill(
            start_color=estado_colores.get(estado, 'FFFFFF'),
            end_color=estado_colores.get(estado, 'FFFFFF'),
            fill_type='solid',
        )
        fila = [
            balance.get('numero_secuencial', row_num - 1),
            balance.get('cliente_nombre', ''),
            balance.get('cuenta_numero', ''),
            balance.get('banco_nombre', ''),
            balance.get('sucursal_nombre', ''),
            str(balance.get('fecha', '')),
            balance.get('total_ventas', 0),
            balance.get('total_pagado', 0),
            balance.get('saldo_pendiente', 0),
            estado,
            str(balance.get('fecha_vencimiento', '') or ''),
            balance.get('acumulado', 0),
        ]
        for col_num, value in enumerate(fila, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.fill = row_fill

    # Ajustar anchos de columna
    for col_num in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_num)].auto_size = True

    # Totals row
    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
    ws.cell(row=total_row, column=7, value=context['total_ventas']).font = Font(bold=True)
    ws.cell(row=total_row, column=8, value=context['total_pagado']).font = Font(bold=True)
    ws.cell(row=total_row, column=9, value=context['saldo_pendiente_total']).font = Font(bold=True)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="balances_ventas.xlsx"'
    wb.save(response)
    return response


@login_required
@permission_required('ventas.view_ventas', raise_exception=True)
def reporte_cobranza_global(request):
    """
    Vista principal del Reporte Global de Cobranza.

    GET params:
        fecha_inicio  (YYYY-MM-DD)  — inicio del período
        fecha_fin     (YYYY-MM-DD)  — fin del período
        tipo_cambio   (decimal)     — override manual del tipo de cambio USD→MXN
    """
    hoy = date.today()
    # Default: temporada / año fiscal actual (1-ene al 31-dic)
    default_inicio = date(hoy.year, 1, 1)
    default_fin = date(hoy.year, 12, 31)

    fecha_inicio_str = request.GET.get('fecha_inicio', default_inicio.isoformat())
    fecha_fin_str = request.GET.get('fecha_fin', default_fin.isoformat())
    tipo_cambio_str = request.GET.get('tipo_cambio', '')

    # Parsear fechas
    try:
        fecha_inicio = date.fromisoformat(fecha_inicio_str)
    except (ValueError, TypeError):
        fecha_inicio = default_inicio

    try:
        fecha_fin = date.fromisoformat(fecha_fin_str)
    except (ValueError, TypeError):
        fecha_fin = default_fin

    # Parsear tipo de cambio
    tipo_cambio_override = None
    if tipo_cambio_str:
        try:
            tipo_cambio_override = Decimal(tipo_cambio_str)
        except Exception:
            pass

    datos = generar_reporte_cobranza(
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        tipo_cambio_override=tipo_cambio_override,
    )

    context = {
        **datos,
        'fecha_inicio_str': fecha_inicio.isoformat(),
        'fecha_fin_str': fecha_fin.isoformat(),
        'tipo_cambio_input': tipo_cambio_str,
        'hoy': hoy,
    }
    return render(request, 'ventas/reporte_cobranza.html', context)