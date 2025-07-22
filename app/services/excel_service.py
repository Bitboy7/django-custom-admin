"""
Servicio para exportación de reportes a Excel
"""
import openpyxl
from openpyxl.styles import PatternFill, NamedStyle, Font, Alignment
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.utils import get_column_letter
from datetime import datetime
import calendar

from gastos.models import Gastos, Cuenta, Compra, SaldoMensual
from ventas.models import Ventas, Anticipo


class ExcelReportService:
    """Servicio para crear reportes en Excel"""
    
    def __init__(self):
        self.header_style = None
        self.contable_style = None
        self.date_style = None
    
    def _setup_styles(self):
        """Configura los estilos para el Excel"""
        self.header_style = NamedStyle(name="header_style")
        self.header_style.font = Font(bold=True, color="FFFFFF")
        self.header_style.fill = PatternFill("solid", fgColor="366092")
        self.header_style.alignment = Alignment(horizontal="center", vertical="center")
        
        self.contable_style = NamedStyle(name="contable_style")
        self.contable_style.number_format = "$#,##0.00"
        self.contable_style.alignment = Alignment(horizontal="right")
        
        self.date_style = NamedStyle(name="date_style")
        self.date_style.number_format = "DD/MM/YYYY"
        self.date_style.alignment = Alignment(horizontal="center")
    
    def remove_timezone(self, dt):
        """Elimina la información de zona horaria de un objeto datetime de manera segura."""
        if dt is None:
            return None
        elif hasattr(dt, 'tzinfo') and dt.tzinfo:
            return dt.replace(tzinfo=None)
        return dt
    
    def convert_money_to_float(self, value):
        """Convierte un objeto Money a float de manera segura."""
        if value is None:
            return 0.0
        elif hasattr(value, 'amount'):
            return float(value.amount)
        else:
            return float(value)
    
    def create_full_report(self):
        """Crea un reporte completo con todas las hojas"""
        wb = openpyxl.Workbook()
        
        # Eliminar la hoja por defecto
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        # Configurar estilos
        self._setup_styles()
        
        # Crear hojas
        ws_gastos = wb.create_sheet("Gastos")
        self.create_gastos_sheet(ws_gastos)
        
        ws_compras = wb.create_sheet("Compras")
        self.create_compras_sheet(ws_compras)
        
        ws_ventas = wb.create_sheet("Ventas")
        self.create_ventas_sheet(ws_ventas)
        
        ws_anticipos = wb.create_sheet("Anticipos")
        self.create_anticipos_sheet(ws_anticipos)
        
        ws_balance = wb.create_sheet("Balance Mensual")
        self.create_balance_sheet(ws_balance, wb)
        
        return wb
    
    def create_gastos_sheet(self, ws):
        """Crea la hoja de gastos con datos y formateo"""
        headers = [
            "ID", "Fecha", "Sucursal", "Categoría", "Cuenta", "Monto", 
            "Descripción", "Fecha de Registro"
        ]
        ws.append(headers)
        
        # Aplicar estilo a los encabezados
        for col_num, _ in enumerate(headers, 1):
            ws.cell(row=1, column=col_num).style = self.header_style
        
        # Obtener datos de gastos
        gastos = Gastos.objects.all().order_by('-fecha')
        
        # Agregar datos a la hoja
        for gasto in gastos:
            row = [
                gasto.id,
                self.remove_timezone(gasto.fecha),
                gasto.id_sucursal.nombre,
                gasto.id_cat_gastos.nombre,
                gasto.id_cuenta_banco.numero_cuenta,
                self.convert_money_to_float(gasto.monto),
                gasto.descripcion,
                self.remove_timezone(gasto.fecha_registro)
            ]
            ws.append(row)
        
        # Aplicar estilos a las celdas
        for row_idx, _ in enumerate(gastos, 2):
            ws.cell(row=row_idx, column=2).style = self.date_style
            ws.cell(row=row_idx, column=8).style = self.date_style
            ws.cell(row=row_idx, column=6).style = self.contable_style
        
        # Agregar totales
        total_row = len(gastos) + 2
        ws.cell(row=total_row, column=5, value="Total:")
        ws.cell(row=total_row, column=5).font = Font(bold=True)
        
        total_formula = f"=SUM(F2:F{total_row-1})"
        ws.cell(row=total_row, column=6, value=total_formula)
        ws.cell(row=total_row, column=6).style = self.contable_style
        ws.cell(row=total_row, column=6).font = Font(bold=True)
        
        self._adjust_column_widths(ws)
    
    def create_compras_sheet(self, ws):
        """Crea la hoja de compras con datos y formateo"""
        headers = [
            "ID", "Fecha", "Productor", "Producto", "Cantidad", "Precio Unitario", 
            "Monto Total", "Cuenta", "Tipo de Pago", "Fecha de Registro"
        ]
        ws.append(headers)
        
        for col_num, _ in enumerate(headers, 1):
            ws.cell(row=1, column=col_num).style = self.header_style
        
        compras = Compra.objects.all().order_by('-fecha_compra')
        
        for compra in compras:
            row = [
                compra.id,
                self.remove_timezone(compra.fecha_compra),
                compra.productor.nombre_completo,
                f"{compra.producto.nombre} - {compra.producto.variedad}",
                compra.cantidad,
                self.convert_money_to_float(compra.precio_unitario),
                self.convert_money_to_float(compra.monto_total),
                compra.cuenta.numero_cuenta if compra.cuenta else "",
                compra.tipo_pago,
                self.remove_timezone(compra.fecha_registro)
            ]
            ws.append(row)
        
        for row_idx, _ in enumerate(compras, 2):
            ws.cell(row=row_idx, column=2).style = self.date_style
            ws.cell(row=row_idx, column=10).style = self.date_style
            ws.cell(row=row_idx, column=6).style = self.contable_style
            ws.cell(row=row_idx, column=7).style = self.contable_style
        
        # Agregar totales
        total_row = len(compras) + 2
        ws.cell(row=total_row, column=6, value="Total:")
        ws.cell(row=total_row, column=6).font = Font(bold=True)
        
        total_formula = f"=SUM(G2:G{total_row-1})"
        ws.cell(row=total_row, column=7, value=total_formula)
        ws.cell(row=total_row, column=7).style = self.contable_style
        ws.cell(row=total_row, column=7).font = Font(bold=True)
        
        self._adjust_column_widths(ws)
    
    def create_ventas_sheet(self, ws):
        """Crea la hoja de ventas con datos y formateo"""
        headers = [
            "ID", "Fecha Salida", "Agente", "Fecha Depósito", "Carga", "PO",
            "Producto", "Cantidad", "Monto", "Cliente", "Sucursal", "Cuenta"
        ]
        ws.append(headers)
        
        for col_num, _ in enumerate(headers, 1):
            ws.cell(row=1, column=col_num).style = self.header_style
        
        ventas = Ventas.objects.all().order_by('-fecha_salida_manifiesto')
        
        for venta in ventas:
            row = [
                venta.id,
                self.remove_timezone(venta.fecha_salida_manifiesto),
                venta.agente_id.nombre if hasattr(venta.agente_id, 'nombre') else str(venta.agente_id),
                self.remove_timezone(venta.fecha_deposito),
                venta.carga,
                venta.PO,
                venta.producto.nombre,
                venta.cantidad,
                self.convert_money_to_float(venta.monto),
                venta.cliente.nombre if hasattr(venta.cliente, 'nombre') else str(venta.cliente),
                venta.sucursal_id.nombre,
                venta.cuenta.numero_cuenta if hasattr(venta, 'cuenta') and venta.cuenta else ""
            ]
            ws.append(row)
        
        for row_idx, _ in enumerate(ventas, 2):
            ws.cell(row=row_idx, column=2).style = self.date_style
            ws.cell(row=row_idx, column=4).style = self.date_style
            ws.cell(row=row_idx, column=9).style = self.contable_style
        
        # Agregar totales
        total_row = len(ventas) + 2
        ws.cell(row=total_row, column=8, value="Total:")
        ws.cell(row=total_row, column=8).font = Font(bold=True)
        
        total_formula = f"=SUM(I2:I{total_row-1})"
        ws.cell(row=total_row, column=9, value=total_formula)
        ws.cell(row=total_row, column=9).style = self.contable_style
        ws.cell(row=total_row, column=9).font = Font(bold=True)
        
        self._adjust_column_widths(ws)
    
    def create_anticipos_sheet(self, ws):
        """Crea la hoja de anticipos con datos y formateo"""
        headers = [
            "ID", "Fecha", "Cliente", "Sucursal", "Cuenta", "Monto", 
            "Descripción", "Estado del Anticipo"
        ]
        ws.append(headers)
        
        for col_num, _ in enumerate(headers, 1):
            ws.cell(row=1, column=col_num).style = self.header_style
        
        anticipos = Anticipo.objects.all().order_by('-fecha')
        
        for anticipo in anticipos:
            row = [
                anticipo.id,
                self.remove_timezone(anticipo.fecha),
                anticipo.cliente.nombre if hasattr(anticipo.cliente, 'nombre') else str(anticipo.cliente),
                anticipo.sucursal.nombre if hasattr(anticipo.sucursal, 'nombre') else str(anticipo.sucursal),
                anticipo.cuenta.numero_cuenta if hasattr(anticipo.cuenta, 'numero_cuenta') else str(anticipo.cuenta),
                self.convert_money_to_float(anticipo.monto),
                anticipo.descripcion,
                anticipo.estado_anticipo
            ]
            ws.append(row)
        
        for row_idx, _ in enumerate(anticipos, 2):
            ws.cell(row=row_idx, column=2).style = self.date_style
            ws.cell(row=row_idx, column=6).style = self.contable_style
        
        # Agregar totales
        total_row = len(anticipos) + 2
        ws.cell(row=total_row, column=5, value="Total:")
        ws.cell(row=total_row, column=5).font = Font(bold=True)
        
        total_formula = f"=SUM(F2:F{total_row-1})"
        ws.cell(row=total_row, column=6, value=total_formula)
        ws.cell(row=total_row, column=6).style = self.contable_style
        ws.cell(row=total_row, column=6).font = Font(bold=True)
        
        self._adjust_column_widths(ws)
    
    def create_balance_sheet(self, ws, wb):
        """Crea la hoja de balance mensual con datos y formateo"""
        headers = [
            "Año", "Mes", "Cuenta", "Banco", "Saldo Inicial", "Gastos", 
            "Compras", "Ventas", "Anticipos", "Saldo Final"
        ]
        ws.append(headers)
        
        for col_num, _ in enumerate(headers, 1):
            ws.cell(row=1, column=col_num).style = self.header_style
        
        cuentas = Cuenta.objects.all()
        current_year = datetime.now().year
        saldos_acumulados = {}
        
        row_idx = 2
        for cuenta in cuentas:
            saldos_acumulados[cuenta.id] = 0
            
            # Verificar saldo inicial de enero
            try:
                saldo_enero = SaldoMensual.objects.filter(cuenta=cuenta, mes=1)
                for saldo in saldo_enero:
                    year_value = None
                    for field_name in ['año', 'anio', 'year']:
                        try:
                            year_value = getattr(saldo, field_name, None)
                            if year_value is not None and int(year_value) == current_year:
                                saldos_acumulados[cuenta.id] = self.convert_money_to_float(saldo.saldo_inicial)
                                break
                        except:
                            continue
                    if saldos_acumulados[cuenta.id] != 0:
                        break
            except:
                pass
            
            for month in range(1, 13):
                month_name = calendar.month_name[month]
                saldo_inicial = saldos_acumulados[cuenta.id]
                
                # Calcular movimientos del mes usando agregaciones
                from django.db.models import Sum
                
                gastos_mes = Gastos.objects.filter(
                    id_cuenta_banco=cuenta,
                    fecha__year=current_year,
                    fecha__month=month
                ).aggregate(total=Sum('monto'))['total'] or 0
                
                compras_mes = Compra.objects.filter(
                    cuenta=cuenta,
                    fecha_compra__year=current_year,
                    fecha_compra__month=month
                ).aggregate(total=Sum('monto_total'))['total'] or 0
                
                ventas_mes = Ventas.objects.filter(
                    cuenta=cuenta,
                    fecha_deposito__year=current_year,
                    fecha_deposito__month=month
                ).aggregate(total=Sum('monto'))['total'] or 0
                
                anticipos_mes = Anticipo.objects.filter(
                    cuenta=cuenta,
                    fecha__year=current_year,
                    fecha__month=month
                ).aggregate(total=Sum('monto'))['total'] or 0
                
                # Convertir a float - manejar objetos Money
                gastos_mes = self.convert_money_to_float(gastos_mes)
                compras_mes = self.convert_money_to_float(compras_mes)
                ventas_mes = self.convert_money_to_float(ventas_mes)
                anticipos_mes = self.convert_money_to_float(anticipos_mes)
                
                saldo_final = saldo_inicial - gastos_mes + compras_mes + ventas_mes + anticipos_mes
                saldos_acumulados[cuenta.id] = saldo_final
                
                ws.append([
                    current_year, month_name, cuenta.numero_cuenta, cuenta.id_banco.nombre,
                    saldo_inicial, gastos_mes, compras_mes, ventas_mes, anticipos_mes, saldo_final
                ])
                
                for col in range(5, 11):
                    ws.cell(row=row_idx, column=col).style = self.contable_style
                
                row_idx += 1
        
        # Agregar totales
        total_row = row_idx
        ws.cell(row=total_row, column=4, value="TOTALES:")
        ws.cell(row=total_row, column=4).font = Font(bold=True)
        
        for col in range(5, 11):
            col_letter = get_column_letter(col)
            formula = f"=SUM({col_letter}2:{col_letter}{total_row-1})"
            ws.cell(row=total_row, column=col, value=formula)
            ws.cell(row=total_row, column=col).style = self.contable_style
            ws.cell(row=total_row, column=col).font = Font(bold=True)
        
        self._adjust_column_widths(ws)
        self._add_balance_chart(ws, total_row)
        
        # Crear hoja de evolución de saldos
        ws_charts = wb.create_sheet("Evolución de Saldos")
        self.create_saldo_evolution_charts(wb, ws_charts, cuentas, current_year)
    
    def create_saldo_evolution_charts(self, wb, ws, cuentas, year):
        """Crea gráficos de línea que muestran la evolución del saldo"""
        ws.cell(row=1, column=1, value="Evolución de Saldos por Cuenta")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        
        balance_sheet = wb["Balance Mensual"]
        current_row = 3
        
        for cuenta in cuentas:
            chart = LineChart()
            chart.title = f"Evolución de Saldo - {cuenta.numero_cuenta} ({cuenta.id_banco.nombre})"
            chart.style = 13
            chart.x_axis.title = "Mes"
            chart.y_axis.title = "Saldo ($)"
            chart.height = 10
            chart.width = 20
            
            meses = []
            saldos_iniciales = []
            saldos_finales = []
            
            for row_idx in range(2, balance_sheet.max_row + 1):
                cuenta_num = balance_sheet.cell(row=row_idx, column=3).value
                if cuenta_num == cuenta.numero_cuenta:
                    mes = balance_sheet.cell(row=row_idx, column=2).value
                    saldo_inicial = balance_sheet.cell(row=row_idx, column=5).value
                    saldo_final = balance_sheet.cell(row=row_idx, column=10).value
                    
                    meses.append(mes)
                    saldos_iniciales.append(saldo_inicial)
                    saldos_finales.append(saldo_final)
            
            if meses:
                ws.cell(row=current_row, column=1, value=f"Cuenta: {cuenta.numero_cuenta} - {cuenta.id_banco.nombre}")
                ws.cell(row=current_row, column=1).font = Font(bold=True)
                current_row += 1
                
                ws.cell(row=current_row, column=1, value="Mes")
                ws.cell(row=current_row, column=2, value="Saldo Inicial")
                ws.cell(row=current_row, column=3, value="Saldo Final")
                for col in range(1, 4):
                    ws.cell(row=current_row, column=col).font = Font(bold=True)
                current_row += 1
                
                start_row = current_row
                for i, mes in enumerate(meses):
                    ws.cell(row=current_row, column=1, value=mes)
                    ws.cell(row=current_row, column=2, value=saldos_iniciales[i])
                    ws.cell(row=current_row, column=3, value=saldos_finales[i])
                    ws.cell(row=current_row, column=2).number_format = "$#,##0.00"
                    ws.cell(row=current_row, column=3).number_format = "$#,##0.00"
                    current_row += 1
                
                data = Reference(ws, min_col=2, min_row=current_row-len(meses)-1, max_row=current_row-1, max_col=3)
                cats = Reference(ws, min_col=1, min_row=current_row-len(meses), max_row=current_row-1)
                
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                
                ws.add_chart(chart, f"E{start_row}")
                current_row += 15
            else:
                ws.cell(row=current_row, column=1, value=f"No hay datos para la cuenta {cuenta.numero_cuenta}")
                current_row += 2
        
        self._adjust_column_widths(ws, max_col=4)
    
    def _adjust_column_widths(self, ws, max_col=None):
        """Ajusta el ancho de las columnas automáticamente"""
        max_col = max_col or ws.max_column
        for col_idx in range(1, max_col + 1):
            max_length = 0
            column = ws[get_column_letter(col_idx)]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
    
    def _add_balance_chart(self, ws, total_row):
        """Agrega un gráfico de barras al balance"""
        chart = BarChart()
        chart.title = "Saldo Final Mensual por Cuenta"
        chart.style = 10
        chart.x_axis.title = "Cuentas y Meses"
        chart.y_axis.title = "Monto ($)"
        
        data = Reference(ws, min_col=10, min_row=1, max_row=total_row-1, max_col=10)
        cats = Reference(ws, min_col=2, min_row=2, max_row=total_row-1)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        ws.add_chart(chart, "M2")
