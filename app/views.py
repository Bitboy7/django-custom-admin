from django.shortcuts import render
from datetime import datetime
from django.contrib.auth.decorators import login_required
import numpy as np
from django.db.models.functions import TruncMonth, TruncWeek, TruncDay
from django.contrib.auth.decorators import user_passes_test
from django.http import HttpResponse
from django.db.models import Sum, Avg, Max, Min, Count
import openpyxl
from openpyxl.styles import PatternFill, NamedStyle, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.utils import get_column_letter
from gastos.models import Gastos, Cuenta, Compra, SaldoMensual
from ventas.models import Ventas, Anticipo
import calendar

def is_admin(user):
    return user.is_superuser

@user_passes_test(is_admin)
def export_full_report_to_excel(request):
    """
    Exporta un informe completo con gastos, compras, ventas, anticipos y un balance mensual.
    """
    # Crear un libro de trabajo y definir estilos
    wb = openpyxl.Workbook()
    
    # Eliminar la hoja por defecto
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Definir estilos globales
    header_style = NamedStyle(name="header_style")
    header_style.font = Font(bold=True, color="FFFFFF")
    header_style.fill = PatternFill("solid", fgColor="366092")
    header_style.alignment = Alignment(horizontal="center", vertical="center")
    
    contable_style = NamedStyle(name="contable_style")
    contable_style.number_format = "$#,##0.00"
    contable_style.alignment = Alignment(horizontal="right")
    
    date_style = NamedStyle(name="date_style")
    date_style.number_format = "DD/MM/YYYY"
    date_style.alignment = Alignment(horizontal="center")
    
    # Crear hoja para gastos
    ws_gastos = wb.create_sheet("Gastos")
    _create_gastos_sheet(ws_gastos, header_style, contable_style, date_style)
    
    # Crear hoja para compras
    ws_compras = wb.create_sheet("Compras")
    _create_compras_sheet(ws_compras, header_style, contable_style, date_style)
    
    # Crear hoja para ventas
    ws_ventas = wb.create_sheet("Ventas")
    _create_ventas_sheet(ws_ventas, header_style, contable_style, date_style)
    
    # Crear hoja para anticipos
    ws_anticipos = wb.create_sheet("Anticipos")
    _create_anticipos_sheet(ws_anticipos, header_style, contable_style, date_style)
    
    # Crear hoja para balance mensual
    ws_balance = wb.create_sheet("Balance Mensual")
    _create_balance_sheet(ws_balance, header_style, contable_style, wb)
    
    # Generar la respuesta HTTP con el archivo Excel
    current_date = datetime.now().strftime("%Y%m%d")
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reporte_financiero_{current_date}.xlsx"'
    wb.save(response)
    
    return response

def remove_timezone(dt):
    """Elimina la información de zona horaria de un objeto datetime de manera segura."""
    if dt is None:
        return None
    elif hasattr(dt, 'tzinfo') and dt.tzinfo:
        return dt.replace(tzinfo=None)
    return dt

def _create_gastos_sheet(ws, header_style, contable_style, date_style):
    """Crea la hoja de gastos con datos y formateo"""
    # Definir encabezados
    headers = [
        "ID", "Fecha", "Sucursal", "Categoría", "Cuenta", "Monto", 
        "Descripción", "Fecha de Registro"
    ]
    ws.append(headers)
    
    # Aplicar estilo a los encabezados
    for col_num, _ in enumerate(headers, 1):
        ws.cell(row=1, column=col_num).style = header_style
    
    # Obtener datos de gastos
    gastos = Gastos.objects.all().order_by('-fecha')
    
    # Agregar datos a la hoja
    for gasto in gastos:
        # Convertir fechas a formato sin zona horaria
        row = [
            gasto.id,
            remove_timezone(gasto.fecha),
            gasto.id_sucursal.nombre,
            gasto.id_cat_gastos.nombre,
            gasto.id_cuenta_banco.numero_cuenta,
            gasto.monto,
            gasto.descripcion,
            remove_timezone(gasto.fecha_registro)
        ]
        ws.append(row)
    
    # Aplicar estilos a las celdas
    for row_idx, _ in enumerate(gastos, 2):
        # Estilo para fechas
        ws.cell(row=row_idx, column=2).style = date_style
        ws.cell(row=row_idx, column=8).style = date_style
        
        # Estilo para montos
        ws.cell(row=row_idx, column=6).style = contable_style
    
    # Agregar totales
    total_row = len(gastos) + 2
    ws.cell(row=total_row, column=5, value="Total:")
    ws.cell(row=total_row, column=5).font = Font(bold=True)
    
    total_formula = f"=SUM(F2:F{total_row-1})"
    ws.cell(row=total_row, column=6, value=total_formula)
    ws.cell(row=total_row, column=6).style = contable_style
    ws.cell(row=total_row, column=6).font = Font(bold=True)
    
    # Ajustar ancho de columnas
    for col_idx, column in enumerate(ws.columns, 1):
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

def _create_compras_sheet(ws, header_style, contable_style, date_style):
    """Crea la hoja de compras con datos y formateo"""
    # Definir encabezados
    headers = [
        "ID", "Fecha", "Productor", "Producto", "Cantidad", "Precio Unitario", 
        "Monto Total", "Cuenta", "Tipo de Pago", "Fecha de Registro"
    ]
    ws.append(headers)
    
    # Aplicar estilo a los encabezados
    for col_num, _ in enumerate(headers, 1):
        ws.cell(row=1, column=col_num).style = header_style
    
    # Obtener datos de compras
    compras = Compra.objects.all().order_by('-fecha_compra')
    
    # Agregar datos a la hoja
    for compra in compras:
        row = [
            compra.id,
            remove_timezone(compra.fecha_compra),
            compra.productor.nombre_completo,
            f"{compra.producto.nombre} - {compra.producto.variedad}",
            compra.cantidad,
            compra.precio_unitario,
            compra.monto_total,
            compra.cuenta.numero_cuenta if compra.cuenta else "",
            compra.tipo_pago,
            remove_timezone(compra.fecha_registro)
        ]
        ws.append(row)
    
    # Aplicar estilos a las celdas
    for row_idx, _ in enumerate(compras, 2):
        # Estilo para fechas
        ws.cell(row=row_idx, column=2).style = date_style
        ws.cell(row=row_idx, column=10).style = date_style
        
        # Estilo para montos
        ws.cell(row=row_idx, column=6).style = contable_style
        ws.cell(row=row_idx, column=7).style = contable_style
    
    # Agregar totales
    total_row = len(compras) + 2
    ws.cell(row=total_row, column=6, value="Total:")
    ws.cell(row=total_row, column=6).font = Font(bold=True)
    
    total_formula = f"=SUM(G2:G{total_row-1})"
    ws.cell(row=total_row, column=7, value=total_formula)
    ws.cell(row=total_row, column=7).style = contable_style
    ws.cell(row=total_row, column=7).font = Font(bold=True)
    
    # Ajustar ancho de columnas
    for col_idx, column in enumerate(ws.columns, 1):
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

def _create_ventas_sheet(ws, header_style, contable_style, date_style):
    """Crea la hoja de ventas con datos y formateo"""
    # Definir encabezados
    headers = [
        "ID", "Fecha Salida", "Agente", "Fecha Depósito", "Carga", "PO",
        "Producto", "Cantidad", "Monto", "Cliente", "Sucursal", "Cuenta"
    ]
    ws.append(headers)
    
    # Aplicar estilo a los encabezados
    for col_num, _ in enumerate(headers, 1):
        ws.cell(row=1, column=col_num).style = header_style
    
    # Obtener datos de ventas
    ventas = Ventas.objects.all().order_by('-fecha_salida_manifiesto')
    
    # Agregar datos a la hoja
    for venta in ventas:
        row = [
            venta.id,
            remove_timezone(venta.fecha_salida_manifiesto),
            venta.agente_id.nombre if hasattr(venta.agente_id, 'nombre') else str(venta.agente_id),
            remove_timezone(venta.fecha_deposito),
            venta.carga,
            venta.PO,
            venta.producto.nombre,
            venta.cantidad,
            venta.monto,
            venta.cliente.nombre if hasattr(venta.cliente, 'nombre') else str(venta.cliente),
            venta.sucursal_id.nombre,
            venta.cuenta.numero_cuenta if hasattr(venta, 'cuenta') and venta.cuenta else ""
        ]
        ws.append(row)
    
    # Aplicar estilos a las celdas
    for row_idx, _ in enumerate(ventas, 2):
        # Estilo para fechas
        ws.cell(row=row_idx, column=2).style = date_style
        ws.cell(row=row_idx, column=4).style = date_style
        
        # Estilo para montos
        ws.cell(row=row_idx, column=9).style = contable_style
    
    # Agregar totales
    total_row = len(ventas) + 2
    ws.cell(row=total_row, column=8, value="Total:")
    ws.cell(row=total_row, column=8).font = Font(bold=True)
    
    total_formula = f"=SUM(I2:I{total_row-1})"
    ws.cell(row=total_row, column=9, value=total_formula)
    ws.cell(row=total_row, column=9).style = contable_style
    ws.cell(row=total_row, column=9).font = Font(bold=True)
    
    # Ajustar ancho de columnas
    for col_idx, column in enumerate(ws.columns, 1):
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

def _create_anticipos_sheet(ws, header_style, contable_style, date_style):
    """Crea la hoja de anticipos con datos y formateo"""
    # Definir encabezados
    headers = [
        "ID", "Fecha", "Cliente", "Sucursal", "Cuenta", "Monto", 
        "Descripción", "Estado del Anticipo"
    ]
    ws.append(headers)
    
    # Aplicar estilo a los encabezados
    for col_num, _ in enumerate(headers, 1):
        ws.cell(row=1, column=col_num).style = header_style
    
    # Obtener datos de anticipos
    anticipos = Anticipo.objects.all().order_by('-fecha')
    
    # Agregar datos a la hoja
    for anticipo in anticipos:
        row = [
            anticipo.id,
            remove_timezone(anticipo.fecha),
            anticipo.cliente.nombre if hasattr(anticipo.cliente, 'nombre') else str(anticipo.cliente),
            anticipo.sucursal.nombre if hasattr(anticipo.sucursal, 'nombre') else str(anticipo.sucursal),
            anticipo.cuenta.numero_cuenta if hasattr(anticipo.cuenta, 'numero_cuenta') else str(anticipo.cuenta),
            anticipo.monto,
            anticipo.descripcion,
            anticipo.estado_anticipo
        ]
        ws.append(row)
    
    # Aplicar estilos a las celdas
    for row_idx, _ in enumerate(anticipos, 2):
        # Estilo para fechas
        ws.cell(row=row_idx, column=2).style = date_style
        
        # Estilo para montos
        ws.cell(row=row_idx, column=6).style = contable_style
    
    # Agregar totales
    total_row = len(anticipos) + 2
    ws.cell(row=total_row, column=5, value="Total:")
    ws.cell(row=total_row, column=5).font = Font(bold=True)
    
    total_formula = f"=SUM(F2:F{total_row-1})"
    ws.cell(row=total_row, column=6, value=total_formula)
    ws.cell(row=total_row, column=6).style = contable_style
    ws.cell(row=total_row, column=6).font = Font(bold=True)
    
    # Ajustar ancho de columnas
    for col_idx, column in enumerate(ws.columns, 1):
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

def _create_balance_sheet(ws, header_style, contable_style, wb):
    """Crea la hoja de balance mensual con datos y formateo"""
    # Definir encabezados
    headers = [
        "Año", "Mes", "Cuenta", "Banco", "Saldo Inicial", "Gastos", 
        "Compras", "Ventas", "Anticipos", "Saldo Final"
    ]
    ws.append(headers)
    
    # Aplicar estilo a los encabezados
    for col_num, _ in enumerate(headers, 1):
        ws.cell(row=1, column=col_num).style = header_style
    
    # Obtener todas las cuentas
    cuentas = Cuenta.objects.all()
    
    # Obtener rango de fechas para el análisis (usamos el último año por defecto)
    current_year = datetime.now().year
    
    # Diccionario para almacenar los saldos de cada cuenta
    saldos_acumulados = {}
    
    # Para cada cuenta, calcular el balance mensual
    row_idx = 2
    for cuenta in cuentas:
        # Inicializamos el saldo acumulado para esta cuenta
        saldos_acumulados[cuenta.id] = 0
        
        # Verificamos si hay un saldo inicial definido para enero
        try:
            saldo_enero = SaldoMensual.objects.filter(
                cuenta=cuenta,
                mes=1
            )
            # Filtramos por año en Python
            for saldo in saldo_enero:
                year_value = None
                # Intentamos diferentes nombres de campo
                for field_name in ['año', 'anio', 'year']:
                    try:
                        year_value = getattr(saldo, field_name, None)
                        if year_value is not None and int(year_value) == current_year:
                            saldos_acumulados[cuenta.id] = float(saldo.saldo_inicial)
                            break
                    except:
                        continue
                if saldos_acumulados[cuenta.id] != 0:
                    break
        except:
            pass
        
        for month in range(1, 13):
            # Obtener el nombre del mes
            month_name = calendar.month_name[month]
            
            # El saldo inicial del mes es el saldo acumulado hasta el momento
            saldo_inicial = saldos_acumulados[cuenta.id]
            
            # Calcular movimientos del mes
            gastos_mes = Gastos.objects.filter(
                id_cuenta_banco=cuenta,
                fecha__year=current_year,
                fecha__month=month
            ).aggregate(total=Sum('monto'))['total'] or 0
            
            compras_mes = Compra.objects.filter(
                cuenta=cuenta,
                fecha_compra__year=current_year,
                fecha_compra__month=month
            ).aggregate(total=Sum('monto_total'))['total'] or 0
            
            ventas_mes = Ventas.objects.filter(
                cuenta=cuenta,
                fecha_deposito__year=current_year,
                fecha_deposito__month=month
            ).aggregate(total=Sum('monto'))['total'] or 0
            
            anticipos_mes = Anticipo.objects.filter(
                cuenta=cuenta,
                fecha__year=current_year,
                fecha__month=month
            ).aggregate(total=Sum('monto'))['total'] or 0
            
            # Convertimos a float para evitar problemas con tipos de datos
            gastos_mes = float(gastos_mes) if gastos_mes is not None else 0
            compras_mes = float(compras_mes) if compras_mes is not None else 0
            ventas_mes = float(ventas_mes) if ventas_mes is not None else 0
            anticipos_mes = float(anticipos_mes) if anticipos_mes is not None else 0
            
            # Calcular saldo final del mes
            # Los gastos son salidas, compras/ventas/anticipos son entradas
            saldo_final = saldo_inicial - gastos_mes + compras_mes + ventas_mes + anticipos_mes
            
            # Actualizar el saldo acumulado para el próximo mes
            saldos_acumulados[cuenta.id] = saldo_final
            
            # Crear fila con los datos
            ws.append([
                current_year,
                month_name,
                cuenta.numero_cuenta,
                cuenta.id_banco.nombre,
                saldo_inicial,
                gastos_mes,
                compras_mes,
                ventas_mes,
                anticipos_mes,
                saldo_final
            ])
            
            # Aplicar estilos a las celdas de montos
            for col in range(5, 11):
                ws.cell(row=row_idx, column=col).style = contable_style
            
            row_idx += 1
    
    # Agregar totales generales
    total_row = row_idx
    ws.cell(row=total_row, column=4, value="TOTALES:")
    ws.cell(row=total_row, column=4).font = Font(bold=True)
    
    # Añadir fórmulas para los totales por columna
    for col in range(5, 11):
        col_letter = get_column_letter(col)
        formula = f"=SUM({col_letter}2:{col_letter}{total_row-1})"
        ws.cell(row=total_row, column=col, value=formula)
        ws.cell(row=total_row, column=col).style = contable_style
        ws.cell(row=total_row, column=col).font = Font(bold=True)
    
    # Ajustar ancho de columnas
    for col_idx, column in enumerate(ws.columns, 1):
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
    
    # Crear un gráfico de barras para el saldo final mensual
    chart = BarChart()
    chart.title = "Saldo Final Mensual por Cuenta"
    chart.style = 10
    chart.x_axis.title = "Cuentas y Meses"
    chart.y_axis.title = "Monto ($)"
    
    # Definir rangos de datos para el gráfico (saldos finales)
    data = Reference(ws, min_col=10, min_row=1, max_row=total_row-1, max_col=10)
    cats = Reference(ws, min_col=2, min_row=2, max_row=total_row-1)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    # Añadir el gráfico a la hoja
    ws.add_chart(chart, "M2")
    
    # Agregar una hoja adicional con gráficos de evolución de saldos
    ws_charts = wb.create_sheet("Evolución de Saldos")
    _create_saldo_evolution_charts(wb, ws_charts, cuentas, current_year)

def _create_saldo_evolution_charts(wb, ws, cuentas, year):
    """Crea gráficos de línea que muestran la evolución del saldo a lo largo del año para cada cuenta"""
    # Títulos en la hoja
    ws.cell(row=1, column=1, value="Evolución de Saldos por Cuenta")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    
    # Obtener datos de balance mensual desde la hoja de balance
    balance_sheet = wb["Balance Mensual"]
    
    # Posición inicial para los gráficos
    current_row = 3
    
    for cuenta in cuentas:
        # Crear un nuevo gráfico de línea para cada cuenta
        chart = LineChart()
        chart.title = f"Evolución de Saldo - {cuenta.numero_cuenta} ({cuenta.id_banco.nombre})"
        chart.style = 13  # Estilo del gráfico
        chart.x_axis.title = "Mes"
        chart.y_axis.title = "Saldo ($)"
        chart.height = 10  # Alto del gráfico
        chart.width = 20   # Ancho del gráfico
        
        # Recopilar datos para esta cuenta
        meses = []
        saldos_iniciales = []
        saldos_finales = []
        
        # Filtrar filas que corresponden a esta cuenta
        for row_idx in range(2, balance_sheet.max_row):
            cuenta_num = balance_sheet.cell(row=row_idx, column=3).value
            if cuenta_num == cuenta.numero_cuenta:
                mes = balance_sheet.cell(row=row_idx, column=2).value
                saldo_inicial = balance_sheet.cell(row=row_idx, column=5).value
                saldo_final = balance_sheet.cell(row=row_idx, column=10).value
                
                meses.append(mes)
                saldos_iniciales.append(saldo_inicial)
                saldos_finales.append(saldo_final)
        
        # Solo crear el gráfico si hay datos
        if meses:
            # Agregar datos a la hoja
            ws.cell(row=current_row, column=1, value=f"Cuenta: {cuenta.numero_cuenta} - {cuenta.id_banco.nombre}")
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1
            
            # Encabezados
            ws.cell(row=current_row, column=1, value="Mes")
            ws.cell(row=current_row, column=2, value="Saldo Inicial")
            ws.cell(row=current_row, column=3, value="Saldo Final")
            for col in range(1, 4):
                ws.cell(row=current_row, column=col).font = Font(bold=True)
            current_row += 1
            
            # Datos
            start_row = current_row
            for i, mes in enumerate(meses):
                ws.cell(row=current_row, column=1, value=mes)
                ws.cell(row=current_row, column=2, value=saldos_iniciales[i])
                ws.cell(row=current_row, column=3, value=saldos_finales[i])
                ws.cell(row=current_row, column=2).number_format = "$#,##0.00"
                ws.cell(row=current_row, column=3).number_format = "$#,##0.00"
                current_row += 1
            
            # Configurar el gráfico
            data = Reference(ws, min_col=2, min_row=current_row-len(meses)-1, max_row=current_row-1, max_col=3)
            cats = Reference(ws, min_col=1, min_row=current_row-len(meses), max_row=current_row-1)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            # Establecer colores de línea
            s1 = chart.series[0]
            s1.graphicalProperties.line.solidFill = "4472C4"  # Azul para saldo inicial
            s1.graphicalProperties.line.width = 20000  # ancho de línea
            
            s2 = chart.series[1]
            s2.graphicalProperties.line.solidFill = "ED7D31"  # Naranja para saldo final
            s2.graphicalProperties.line.width = 20000  # ancho de línea
            
            # Añadir leyenda
            chart.legend.position = 'b'
            
            # Añadir el gráfico a la hoja
            ws.add_chart(chart, f"E{start_row}")
            
            # Espacio para el siguiente gráfico
            current_row += 15
        else:
            # No hay datos para esta cuenta, añadir un mensaje
            ws.cell(row=current_row, column=1, value=f"No hay datos para la cuenta {cuenta.numero_cuenta}")
            current_row += 2
    
    # Ajustar ancho de columnas
    for col_idx in range(1, 4):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20
        for row_idx in range(1, current_row):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                max_length = len(str(cell.value)) + 2
                ws.column_dimensions[get_column_letter(col_idx)].width = max(max_length, ws.column_dimensions[get_column_letter(col_idx)].width)
    # Ajustar el alto de las filas para que se vean bien los gráficos
    for row_idx in range(1, current_row):
        ws.row_dimensions[row_idx].height = 20

@login_required
def balances_view(request):
    cuenta_id = request.GET.get('cuenta_id', '')
    year = request.GET.get('year', datetime.now().year)
    month = request.GET.get('month', datetime.now().month)
    periodo = request.GET.get('periodo', 'diario')  # 'diario', 'semanal' o 'mensual'
    dia = request.GET.get('dia', datetime.now().strftime('%Y-%m-%d'))
    fecha_inicio = request.GET.get('fecha_inicio', '')
    fecha_fin = request.GET.get('fecha_fin', '')

    # Obtener los años disponibles
    available_years = Gastos.objects.dates('fecha', 'year')

    # Lista de meses
    months = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

    # Filtrar y agrupar los datos de los gastos según el periodo seleccionado
    filters = {'fecha__year': year}
    if cuenta_id:
        filters['id_cuenta_banco_id'] = cuenta_id
    if month:
        filters['fecha__month'] = month

    if periodo == 'diario':
        if dia:
            filters['fecha'] = dia
        elif fecha_inicio and fecha_fin:
            filters['fecha__range'] = [fecha_inicio, fecha_fin]
        balances = Gastos.objects.filter(**filters).values(
            'id_cuenta_banco__id', 
            'id_cuenta_banco__numero_cuenta',
            'id_cuenta_banco__id_banco__nombre',
            'id_cuenta_banco__id_sucursal__nombre',
            'id_cat_gastos__nombre',
            'fecha'
        ).annotate(
            total_gastos=Sum('monto')
        ).order_by('id_cuenta_banco__id', 'fecha')
    elif periodo == 'semanal':
        balances = Gastos.objects.filter(**filters).annotate(
            semana=TruncWeek('fecha')
        ).values(
            'id_cuenta_banco__id', 
            'id_cuenta_banco__numero_cuenta',
            'id_cuenta_banco__id_banco__nombre',
            'id_cuenta_banco__id_sucursal__nombre',
            'id_cat_gastos__nombre',
            'semana'
        ).annotate(
            total_gastos=Sum('monto')
        ).order_by('id_cuenta_banco__id', 'semana')
    elif periodo == 'mensual':
        balances = Gastos.objects.filter(**filters).annotate(
            mes=TruncMonth('fecha')
        ).values(
            'id_cuenta_banco__id', 
            'id_cuenta_banco__numero_cuenta',
            'id_cuenta_banco__id_banco__nombre',
            'id_cuenta_banco__id_sucursal__nombre',
            'id_cat_gastos__nombre',
            'mes'
        ).annotate(
            total_gastos=Sum('monto')
        ).order_by('id_cuenta_banco__id', 'mes')
    else:
        balances = []

    # Calcular el acumulado de la suma de montos
    acumulado = 0
    for balance in balances:
        acumulado += balance['total_gastos']
        balance['acumulado'] = acumulado

    total_gastos = sum(balance['total_gastos'] for balance in balances)
    promedio_gastos = Gastos.objects.filter(**filters).aggregate(promedio=Avg('monto'))['promedio']
    numero_transacciones = Gastos.objects.filter(**filters).count()
    gasto_maximo = Gastos.objects.filter(**filters).aggregate(maximo=Max('monto'))['maximo']
    gasto_minimo = Gastos.objects.filter(**filters).aggregate(minimo=Min('monto'))['minimo']
    gastos = list(Gastos.objects.filter(**filters).values_list('monto', flat=True))
    gasto_mediano = np.median(gastos) if gastos else 0

    # Obtener las categorías de gasto máximo y mínimo
    categoria_gasto_maximo = Gastos.objects.filter(monto=gasto_maximo).values('id_cat_gastos__nombre').first()
    categoria_gasto_minimo = Gastos.objects.filter(monto=gasto_minimo).values('id_cat_gastos__nombre').first()

    cuentas = Cuenta.objects.all()
    context = {
        'balances': balances,
        'cuentas': cuentas,
        'selected_cuenta_id': cuenta_id,
        'selected_year': year,
        'selected_month': month,
        'selected_periodo': periodo,
        'selected_dia': dia,
        'selected_fecha_inicio': fecha_inicio,
        'selected_fecha_fin': fecha_fin,
        'available_years': available_years,
        'months': months,
        'total_gastos': total_gastos,
        'promedio_gastos': promedio_gastos,
        'numero_transacciones': numero_transacciones,
        'gasto_maximo': gasto_maximo,
        'gasto_minimo': gasto_minimo,
        'gasto_mediano': gasto_mediano,
        'categoria_gasto_maximo': categoria_gasto_maximo['id_cat_gastos__nombre'] if categoria_gasto_maximo else None,
        'categoria_gasto_minimo': categoria_gasto_minimo['id_cat_gastos__nombre'] if categoria_gasto_minimo else None,
        'meses_rango': range(1, 13),  # Agregar el rango de meses al contexto
    }
    return render(request, 'balances.html', context)


