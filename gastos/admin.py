from django.contrib import admin
from django.contrib.admin import ModelAdmin
from django.contrib.admin import SimpleListFilter
from django.template.response import TemplateResponse
from django.urls import path
from django.http import HttpResponse, HttpResponseRedirect
from import_export import resources, fields
from import_export.widgets import ForeignKeyWidget
from import_export.admin import ImportExportModelAdmin
from import_export.forms import ExportForm, ImportForm
from .models import CatGastos, Banco, Cuenta, Gastos, Compra, SaldoMensual, ComprobanteGasto
from django.utils.html import format_html
from django.utils.text import slugify
from django.utils import timezone
from catalogo.models import Sucursal, Productor, Producto
from app.widgets import MoneyWidget


class SucursalGastoFilter(SimpleListFilter):
    title = 'Sucursal'
    parameter_name = 'sucursal'

    def lookups(self, request, model_admin):
        return [
            (str(sucursal.id), sucursal.nombre)
            for sucursal in Sucursal.objects.order_by('nombre')
        ]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(id_sucursal_id=self.value())
        return queryset


class CategoriaGastoFilter(SimpleListFilter):
    title = 'Categoría'
    parameter_name = 'categoria'

    def lookups(self, request, model_admin):
        return [
            (str(categoria.id), categoria.nombre)
            for categoria in CatGastos.objects.order_by('nombre')
        ]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(id_cat_gastos_id=self.value())
        return queryset


class CuentaGastoFilter(SimpleListFilter):
    title = 'Cuenta bancaria'
    parameter_name = 'cuenta'

    def lookups(self, request, model_admin):
        return [
            (str(cuenta.id), f'{cuenta.numero_cuenta} - {cuenta.id_banco.nombre}')
            for cuenta in Cuenta.objects.select_related('id_banco').order_by('numero_cuenta')
        ]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(id_cuenta_banco_id=self.value())
        return queryset


class MontoGastoFilter(SimpleListFilter):
    title = 'Rango de monto'
    parameter_name = 'rango_monto'

    def lookups(self, request, model_admin):
        return [
            ('0-1000', '$0 - $1,000'),
            ('1000-5000', '$1,000 - $5,000'),
            ('5000-10000', '$5,000 - $10,000'),
            ('10000-50000', '$10,000 - $50,000'),
            ('50000+', '$50,000+'),
        ]

    def queryset(self, request, queryset):
        value = self.value()
        if value == '0-1000':
            return queryset.filter(monto__amount__range=[0, 1000])
        if value == '1000-5000':
            return queryset.filter(monto__amount__range=[1000, 5000])
        if value == '5000-10000':
            return queryset.filter(monto__amount__range=[5000, 10000])
        if value == '10000-50000':
            return queryset.filter(monto__amount__range=[10000, 50000])
        if value == '50000+':
            return queryset.filter(monto__amount__gte=50000)
        return queryset


class CatGastoResource(resources.ModelResource):
    fields = ('id', 'nombre', 'fecha_registro')
    class Meta:
        model = CatGastos

@admin.register(CatGastos)
class CatGastosAdmin(ImportExportModelAdmin, ModelAdmin):
    resource_class = CatGastoResource
    import_form_class = ImportForm
    export_form_class = ExportForm
    list_display = ('id', 'nombre')
    search_fields = ('id', 'nombre', 'fecha_registro')
    list_filter = ('nombre', 'fecha_registro')
    list_per_page = 12
    fieldsets = (
        ('Datos del Registro', {
            'fields': ('nombre', 'fecha_registro')
        }),
    )


class BancoResource(resources.ModelResource):
    class Meta:
        model = Banco
        fields = ('id', 'nombre', 'telefono', 'direccion', 'logotipo', 'fecha_registro')

@admin.register(Banco)
class BancoAdmin(ImportExportModelAdmin, ModelAdmin):
    resource_class = BancoResource
    import_form_class = ImportForm
    export_form_class = ExportForm
    list_display = ('id', 'nombre', 'telefono', 'direccion', 'fecha_registro', 'mostrar_logotipo')
    search_fields = ('nombre', 'telefono', 'direccion')
    list_filter = ('nombre', 'telefono', 'direccion', 'fecha_registro')
    list_per_page = 12
    fieldsets = (
        ('Información General', {
            'fields': ('nombre', 'telefono', 'direccion')
        }),
        ('Imagen', {
            'fields': ('logotipo',),
        }),
        ('Metadatos', {
            'fields': ('fecha_registro',),
        })
    )


class CuentaResource(resources.ModelResource):
    class Meta:
        model = Cuenta
        fields = ('id', 'id_banco', 'id_sucursal', 'numero_cuenta', 'numero_cliente', 'rfc', 'clabe', 'fecha_registro')

@admin.register(Cuenta)
class CuentaAdmin(ImportExportModelAdmin, ModelAdmin):
    resource_class = CuentaResource
    import_form_class = ImportForm
    export_form_class = ExportForm
    list_display = ('id', 'mostrar_logotipo_banco', 'id_sucursal', 'numero_cuenta', 'numero_cliente', 'rfc', 'clabe')
    search_fields = ('id_banco__nombre', 'id_sucursal__nombre', 'numero_cuenta', 'numero_cliente')
    list_filter = ('id_banco', 'id_sucursal', 'numero_cuenta', 'numero_cliente', 'rfc')
    list_per_page = 12
    fieldsets = (
        ('Datos de la Cuenta', {
            'fields': ('id_banco', 'id_sucursal', 'numero_cuenta', 'numero_cliente', 'rfc', 'clabe')
        }),
        ('Metadatos', {
            'fields': ('fecha_registro',),
        })
    )
    
    def mostrar_logotipo_banco(self, obj):
        return obj.id_banco.mostrar_logotipo()
    mostrar_logotipo_banco.short_description = 'Banco'
    
class GastosResource(resources.ModelResource):
    sucursal = fields.Field(
        column_name='sucursal',
        attribute='id_sucursal',
        widget=ForeignKeyWidget(Sucursal, field='nombre'))
    
    categoria = fields.Field(
        column_name='categoria',
        attribute='id_cat_gastos',
        widget=ForeignKeyWidget(CatGastos, field='nombre'))
    
    cuenta = fields.Field(
        column_name='cuenta',
        attribute='id_cuenta_banco',
        widget=ForeignKeyWidget(Cuenta, field='numero_cuenta'))
    
    monto = fields.Field(
        column_name='monto',
        attribute='monto',
        widget=MoneyWidget())
    
    class Meta:
        model = Gastos
        fields = ('id', 'sucursal', 'categoria', 'cuenta', 'monto', 'descripcion', 'fecha', 'fecha_registro')
        export_order = ('id', 'sucursal', 'categoria', 'cuenta', 'monto', 'descripcion', 'fecha', 'fecha_registro')
        import_id_fields = ('id',)

    def dehydrate_categoria(self, gasto):
        return gasto.id_cat_gastos.nombre
    
    def dehydrate_sucursal(self, gasto):
        return gasto.id_sucursal.nombre
    
    def dehydrate_cuenta(self, gasto):
        return gasto.id_cuenta_banco.numero_cuenta

@admin.register(Gastos)
class GastosAdmin(ModelAdmin):
    change_list_template = 'admin/gastos/gastos/change_list.html'
    list_display = ('id', 'id_sucursal', 'id_cat_gastos',
                    'id_cuenta_banco', 'monto', 'descripcion', 'fecha', 'fecha_registro')
    search_fields = ('descripcion', 'id_sucursal__nombre', 'id_cat_gastos__nombre', 'id_cuenta_banco__numero_cuenta', 'id_cuenta_banco__id_banco__nombre')
    # La cuenta ya identifica el banco; la fecha del gasto se filtra desde
    # date_hierarchy. Evitamos duplicar controles que confunden al usuario.
    list_filter = (SucursalGastoFilter, CategoriaGastoFilter, CuentaGastoFilter, MontoGastoFilter)
    date_hierarchy = 'fecha'
    ordering = ('fecha', 'fecha_registro', 'id')
    list_select_related = ('id_sucursal', 'id_cat_gastos', 'id_cuenta_banco', 'id_cuenta_banco__id_banco')
    list_per_page = 20
    class Media:
        css = {
            'all': ('css/admin/gastos_changelist.css',)
        }
    fieldsets = (
        ('Datos del Registro', {
            'fields': ('id_sucursal', 'id_cat_gastos', 'id_cuenta_banco', 'monto', 'descripcion', 'fecha')
        }),
    )
    # La exportacion usa el boton superior y respeta los filtros activos.
    # Deshabilitamos acciones masivas para evitar una barra vacia y casillas
    # de seleccion que ya no forman parte de este flujo.
    actions = None

    def get_export_filename(self, request):
        """Construye un nombre legible a partir de los filtros activos."""
        parts = ['gastos']

        year = request.GET.get('fecha__year')
        month = request.GET.get('fecha__month')
        day = request.GET.get('fecha__day')
        if year:
            date_parts = [year]
            if month:
                date_parts.append(month.zfill(2))
            if day:
                date_parts.append(day.zfill(2))
            parts.append(f"fecha-{'-'.join(date_parts)}")

        sucursal_id = request.GET.get('sucursal')
        if sucursal_id:
            nombre = Sucursal.objects.filter(pk=sucursal_id).values_list('nombre', flat=True).first()
            parts.append(f"sucursal-{slugify(nombre or sucursal_id)[:32]}")

        categoria_id = request.GET.get('categoria')
        if categoria_id:
            nombre = CatGastos.objects.filter(pk=categoria_id).values_list('nombre', flat=True).first()
            parts.append(f"categoria-{slugify(nombre or categoria_id)[:32]}")

        cuenta_id = request.GET.get('cuenta')
        if cuenta_id:
            cuenta = Cuenta.objects.select_related('id_banco').filter(pk=cuenta_id).first()
            if cuenta:
                cuenta_label = f'{cuenta.id_banco.nombre}-{cuenta.numero_cuenta}'
            else:
                cuenta_label = cuenta_id
            parts.append(f"cuenta-{slugify(cuenta_label)[:42]}")

        rango_monto = request.GET.get('rango_monto')
        if rango_monto:
            monto_labels = {
                '0-1000': '0-a-1000',
                '1000-5000': '1000-a-5000',
                '5000-10000': '5000-a-10000',
                '10000-50000': '10000-a-50000',
                '50000+': '50000-o-mas',
            }
            parts.append(f"monto-{monto_labels.get(rango_monto, slugify(rango_monto))}")

        search = request.GET.get('q', '').strip()
        if search:
            parts.append(f"busqueda-{slugify(search)[:30]}")

        generated_at = timezone.localtime().strftime('%Y%m%d-%H%M%S')
        return f"{'_'.join(parts)}_{generated_at}.xlsx"

    def export_to_excel(self, request, queryset):
        import openpyxl
        import datetime
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from collections import defaultdict

        NAVY  = "1E3A5F"
        TEAL  = "1AADBC"
        LIGHT = "D6EAF8"
        WHITE = "FFFFFF"
        MONEY = '"$"#,##0.00'

        def _nav_hdr(cell, bg=NAVY):
            cell.font      = Font(bold=True, color=WHITE)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        def _section_title(ws, row, col_start, col_end, text):
            cell = ws.cell(row=row, column=col_start, value=text)
            cell.font = Font(bold=True, color=WHITE, size=11)
            cell.fill = PatternFill("solid", fgColor=NAVY)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if col_end > col_start:
                ws.merge_cells(
                    start_row=row, start_column=col_start,
                    end_row=row,   end_column=col_end,
                )
            ws.row_dimensions[row].height = 20

        def _add_table(ws, name, hdr_row, data_end_row, col_end):
            if data_end_row < hdr_row + 1:
                return
            tab = Table(
                displayName=name,
                ref=f"A{hdr_row}:{get_column_letter(col_end)}{data_end_row}",
            )
            tab.tableStyleInfo = TableStyleInfo(
                name="TableStyleLight2",
                showFirstColumn=False, showLastColumn=False,
                showRowStripes=True,   showColumnStripes=False,
            )
            ws.add_table(tab)

        def _fecha_registro_excel(gasto):
            if not gasto.fecha_registro:
                return None
            if timezone.is_aware(gasto.fecha_registro):
                return timezone.localtime(gasto.fecha_registro).replace(tzinfo=None)
            return gasto.fecha_registro

        queryset = queryset.select_related(
            'id_sucursal', 'id_cat_gastos', 'id_cuenta_banco'
        ).order_by('fecha', 'fecha_registro', 'id')
        data = list(queryset)

        wb = openpyxl.Workbook()

        # ── HOJA 1 — Detalle ─────────────────────────────────────────────────
        ws = wb.active
        ws.title = "Detalle"

        COLUMNS = [
            ("Fecha",        lambda g: g.fecha,                          14,  "DD/MM/YYYY"),
            ("Fecha registro", lambda g: _fecha_registro_excel(g),        20,  "DD/MM/YYYY HH:MM"),
            ("Sucursal",     lambda g: g.id_sucursal.nombre,             22,  "@"),
            ("Categoría",    lambda g: g.id_cat_gastos.nombre,           22,  "@"),
            ("Cuenta",       lambda g: g.id_cuenta_banco.numero_cuenta,  22,  "@"),
            ("Monto",        lambda g: float(g.monto.amount),            16,  MONEY),
            ("Descripción",  lambda g: g.descripcion or "",              40,  "@"),
        ]

        for ci, (hdr, _, width, _) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=ci, value=hdr)
            _nav_hdr(cell)
            ws.column_dimensions[get_column_letter(ci)].width = width

        for ri, gasto in enumerate(data, 2):
            for ci, (_, getter, _, fmt) in enumerate(COLUMNS, 1):
                cell = ws.cell(row=ri, column=ci, value=getter(gasto))
                cell.number_format = fmt
                cell.alignment = Alignment(vertical="center")

        last_data = len(data) + 1  # last row containing data

        # Tabla dinámica de Excel (activa filtros, ordenamiento y bandas)
        if data:
            tab_main = Table(
                displayName="TablaGastos",
                ref=f"A1:{get_column_letter(len(COLUMNS))}{last_data}",
            )
            tab_main.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False, showLastColumn=False,
                showRowStripes=True,   showColumnStripes=False,
            )
            ws.add_table(tab_main)

        # Fila TOTAL fuera de la tabla
        total_r = last_data + 1
        mc = 6
        lbl = ws.cell(row=total_r, column=mc - 1, value="TOTAL")
        lbl.font = Font(bold=True)
        lbl.alignment = Alignment(horizontal="right")
        mc_ltr = get_column_letter(mc)
        tc = ws.cell(row=total_r, column=mc,
                     value=f"=SUM({mc_ltr}2:{mc_ltr}{last_data})")
        tc.number_format = MONEY
        tc.font = Font(bold=True)
        tc.fill = PatternFill("solid", fgColor=LIGHT)

        ws.freeze_panes = "A2"

        # ── HOJA 2 — Resumen ejecutivo ───────────────────────────────────────
        ws2 = wb.create_sheet("Resumen")

        if not data:
            ws2.cell(row=1, column=1, value="Sin datos seleccionados.")
        else:
            grand_total = sum(float(g.monto.amount) for g in data)
            dates = [g.fecha for g in data]
            fecha_min, fecha_max = min(dates), max(dates)
            avg = grand_total / len(data)

            by_cat   = defaultdict(lambda: {'count': 0, 'total': 0.0})
            by_suc   = defaultdict(lambda: {'count': 0, 'total': 0.0})
            by_month = defaultdict(lambda: {'count': 0, 'total': 0.0})

            for g in data:
                amt  = float(g.monto.amount)
                cat  = g.id_cat_gastos.nombre
                suc  = g.id_sucursal.nombre
                mkey = f"{g.fecha.year}-{g.fecha.month:02d}"
                by_cat[cat]['count']    += 1
                by_cat[cat]['total']    += amt
                by_suc[suc]['count']    += 1
                by_suc[suc]['total']    += amt
                by_month[mkey]['count'] += 1
                by_month[mkey]['total'] += amt

            # Anchos de columnas
            for col, w in zip('ABCD', [28, 14, 18, 14]):
                ws2.column_dimensions[col].width = w

            # ── Encabezado del reporte ────────────────────────────────────────
            r = 1
            hdr_cell = ws2.cell(row=r, column=1,
                                value="  REPORTE EJECUTIVO DE GASTOS")
            hdr_cell.font      = Font(bold=True, color=WHITE, size=14)
            hdr_cell.fill      = PatternFill("solid", fgColor=NAVY)
            hdr_cell.alignment = Alignment(horizontal="left", vertical="center")
            ws2.merge_cells(f"A{r}:D{r}")
            ws2.row_dimensions[r].height = 30
            r += 1

            sub = ws2.cell(
                row=r, column=1,
                value=(f"  Generado: {datetime.date.today().strftime('%d/%m/%Y')}"
                       f"   |   Período: {fecha_min.strftime('%d/%m/%Y')}"
                       f" – {fecha_max.strftime('%d/%m/%Y')}"),
            )
            sub.font = Font(italic=True, size=9, color="555555")
            sub.fill = PatternFill("solid", fgColor="EBF5FB")
            ws2.merge_cells(f"A{r}:D{r}")
            r += 2  # blank row separator

            # ── KPIs — (value, number_format) tuples so cells stay numeric ──
            kpis = [
                ("Total Gastos",       grand_total,     MONEY),
                ("N° de Registros",    len(data),       "0"),
                ("Promedio por Gasto", avg,             MONEY),
                ("Categorías",         len(by_cat),     "0"),
                ("Sucursales",         len(by_suc),     "0"),
                ("Meses con gastos",   len(by_month),   "0"),
            ]
            # 2 KPIs per row, 3 rows
            for i, (label, value, fmt) in enumerate(kpis):
                kpi_row = r + (i // 2)
                kpi_col = 1 + (i % 2) * 2
                lc = ws2.cell(row=kpi_row, column=kpi_col, value=label)
                lc.font      = Font(bold=True, size=9, color="666666")
                lc.alignment = Alignment(horizontal="left")
                lc.fill      = PatternFill("solid", fgColor="F8FBFD")
                vc = ws2.cell(row=kpi_row, column=kpi_col + 1, value=value)
                vc.number_format = fmt
                vc.font      = Font(bold=True, size=11, color=NAVY)
                vc.alignment = Alignment(horizontal="right")
                vc.fill      = PatternFill("solid", fgColor="F8FBFD")
            r += 3 + 1  # 3 KPI rows + blank separator

            # ── Sección: Por Categoría ────────────────────────────────────────
            _section_title(ws2, r, 1, 4, "  RESUMEN POR CATEGORÍA")
            r += 1
            for ci, h in enumerate(["Categoría", "N° Gastos", "Total", "% del Total"], 1):
                _nav_hdr(ws2.cell(row=r, column=ci, value=h), TEAL)
            cat_hdr = r
            r += 1
            cat_start = r
            for cat, v in sorted(by_cat.items(), key=lambda x: -x[1]['total']):
                pct = v['total'] / grand_total if grand_total else 0
                ws2.cell(row=r, column=1, value=cat)
                ws2.cell(row=r, column=2, value=v['count']).alignment = Alignment(horizontal="center")
                c3 = ws2.cell(row=r, column=3, value=v['total'])
                c3.number_format = MONEY
                c4 = ws2.cell(row=r, column=4, value=pct)
                c4.number_format = '0.00%'
                r += 1
            cat_end = r - 1
            # total row
            ws2.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
            ws2.cell(row=r, column=2, value=len(data)).font = Font(bold=True)
            ct = ws2.cell(row=r, column=3, value=grand_total)
            ct.number_format = MONEY
            ct.font = Font(bold=True)
            ct.fill = PatternFill("solid", fgColor=LIGHT)
            cp = ws2.cell(row=r, column=4, value=1.0)
            cp.number_format = '0.00%'
            cp.font = Font(bold=True)
            r += 2

            _add_table(ws2, "TablaCategorias", cat_hdr, cat_end, 4)

            # ── Sección: Por Sucursal ─────────────────────────────────────────
            _section_title(ws2, r, 1, 4, "  RESUMEN POR SUCURSAL")
            r += 1
            for ci, h in enumerate(["Sucursal", "N° Gastos", "Total", "% del Total"], 1):
                _nav_hdr(ws2.cell(row=r, column=ci, value=h), TEAL)
            suc_hdr = r
            r += 1
            suc_start = r
            for suc, v in sorted(by_suc.items(), key=lambda x: -x[1]['total']):
                pct = v['total'] / grand_total if grand_total else 0
                ws2.cell(row=r, column=1, value=suc)
                ws2.cell(row=r, column=2, value=v['count']).alignment = Alignment(horizontal="center")
                s3 = ws2.cell(row=r, column=3, value=v['total'])
                s3.number_format = MONEY
                s4 = ws2.cell(row=r, column=4, value=pct)
                s4.number_format = '0.00%'
                r += 1
            suc_end = r - 1
            ws2.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
            ws2.cell(row=r, column=2, value=len(data)).font = Font(bold=True)
            st = ws2.cell(row=r, column=3, value=grand_total)
            st.number_format = MONEY
            st.font = Font(bold=True)
            st.fill = PatternFill("solid", fgColor=LIGHT)
            sp = ws2.cell(row=r, column=4, value=1.0)
            sp.number_format = '0.00%'
            sp.font = Font(bold=True)
            r += 2

            _add_table(ws2, "TablaSucursales", suc_hdr, suc_end, 4)

            # ── Sección: Evolución Mensual ────────────────────────────────────
            _section_title(ws2, r, 1, 3, "  EVOLUCIÓN MENSUAL")
            r += 1
            for ci, h in enumerate(["Mes", "N° Gastos", "Total"], 1):
                _nav_hdr(ws2.cell(row=r, column=ci, value=h), TEAL)
            mon_hdr = r
            r += 1
            mon_start = r
            for mkey in sorted(by_month):
                v = by_month[mkey]
                ws2.cell(row=r, column=1, value=mkey)
                ws2.cell(row=r, column=2, value=v['count']).alignment = Alignment(horizontal="center")
                m3 = ws2.cell(row=r, column=3, value=v['total'])
                m3.number_format = MONEY
                r += 1
            mon_end = r - 1
            ws2.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
            ws2.cell(row=r, column=2, value=len(data)).font = Font(bold=True)
            mt = ws2.cell(row=r, column=3, value=grand_total)
            mt.number_format = MONEY
            mt.font = Font(bold=True)
            mt.fill = PatternFill("solid", fgColor=LIGHT)

            _add_table(ws2, "TablaMensual", mon_hdr, mon_end, 3)

            ws2.freeze_panes = "A4"

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = self.get_export_filename(request)
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    export_to_excel.short_description = "Exportar a Excel (.xlsx)"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                'exportar-excel/',
                self.admin_site.admin_view(self.exportar_excel_admin_view),
                name='gastos_gastos_export_excel',
            ),
            path(
                'balances/',
                self.admin_site.admin_view(self.balances_admin_view),
                name='gastos_gastos_balances',
            ),
            path(
                'balances/resumen-excel/',
                self.admin_site.admin_view(self.balances_resumen_excel_view),
                name='gastos_gastos_balances_resumen_excel',
            ),
        ]
        return custom_urls + urls

    def exportar_excel_admin_view(self, request):
        if not self.has_view_permission(request):
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied
        # La exportacion refleja exactamente el listado: filtros, busqueda y
        # jerarquia de fecha activos, sin depender de la pagina actual.
        filter_params = {
            key for key in request.GET
            if key not in {'all', 'p', 'o', 'ot', '_popup', '_to_field'}
        }
        if not filter_params:
            self.message_user(
                request,
                'Aplica al menos un filtro, fecha o b\u00fasqueda antes de exportar.',
                level='warning',
            )
            return HttpResponseRedirect('../')

        changelist = self.get_changelist_instance(request)
        return self.export_to_excel(request, changelist.queryset)

    def balances_admin_view(self, request):
        from app.services.balance_service import BalanceAnalysisService
        from django.http import JsonResponse
        from django.template.loader import render_to_string
        import json
        balance_service = BalanceAnalysisService()
        context = balance_service.get_full_context(request)
        context.update(self.admin_site.each_context(request))
        context.update({
            'title': 'Acumulado de Gastos',
            'subtitle': None,
            'opts': self.model._meta,
            'has_view_permission': self.has_view_permission(request),
        })

        is_htmx = request.headers.get('HX-Request') == 'true'

        if is_htmx:
            results_html = render_to_string(
                'admin/gastos/partials/balances_results.html',
                context,
                request=request,
            )
            response = HttpResponse(results_html)
            response['HX-Trigger'] = json.dumps({
                'showToast': {
                    'type': 'info',
                    'title': 'Filtros aplicados',
                    'message': 'La informaci&oacute;n ha sido filtrada seg&uacute;n los criterios seleccionados.'
                }
            })
            return response

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            table_html = render_to_string(
                'admin/gastos/partials/balances_table.html',
                context,
                request=request,
            )
            return JsonResponse({
                'kpis': {
                    'total_gastos': float(context.get('total_gastos') or 0),
                    'promedio_gastos': float(context.get('promedio_gastos') or 0),
                    'numero_transacciones': context.get('numero_transacciones', 0),
                    'gasto_maximo': float(context.get('gasto_maximo') or 0),
                    'gasto_minimo': float(context.get('gasto_minimo') or 0),
                    'gasto_mediano': float(context.get('gasto_mediano') or 0),
                    'categoria_gasto_maximo': context.get('categoria_gasto_maximo') or '',
                    'categoria_gasto_minimo': context.get('categoria_gasto_minimo') or '',
                },
                'table_html': table_html,
            })

        return TemplateResponse(request, 'admin/gastos/balances.html', context)

    def balances_resumen_excel_view(self, request):
        if not self.has_view_permission(request):
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied

        from app.services.balance_service import BalanceAnalysisService
        from collections import defaultdict
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        balance_service = BalanceAnalysisService()
        context = balance_service.get_full_context(request)
        balances = context.get('balances', [])
        periodo = context.get('periodo', 'diario')
        sucursal_id = context.get('sucursal_id', '')

        por_categoria = balance_service.get_accumulated_by_category(balances)
        por_sucursal = None
        if not sucursal_id:
            por_sucursal = balance_service.get_accumulated_by_category_per_sucursal(balances)

        NAVY = "2F4550"
        LIGHT = "E6F3FF"
        WHITE = "FFFFFF"
        MONEY = '"$"#,##0.00'

        def nav_hdr(cell):
            cell.font = Font(bold=True, color=WHITE)
            cell.fill = PatternFill("solid", fgColor=NAVY)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        def money_cell(cell, bold=False):
            cell.number_format = MONEY
            cell.alignment = Alignment(horizontal="right")
            if bold:
                cell.font = Font(bold=True)

        def balance_fecha(balance):
            if periodo == 'diario':
                return balance.get('fecha')
            if periodo == 'semanal':
                return balance.get('semana')
            return balance.get('mes')

        def fecha_text(value):
            if not value:
                return ""
            if hasattr(value, 'strftime'):
                if periodo == 'mensual':
                    return value.strftime('%Y-%m')
                return value.strftime('%Y-%m-%d')
            return str(value)

        def fill_row(ws, row, col_end, bg, font=None):
            for cidx in range(1, col_end + 1):
                cell = ws.cell(row=row, column=cidx)
                cell.fill = PatternFill("solid", fgColor=bg)
                if font is not None:
                    cell.font = font

        wb = openpyxl.Workbook()

        # ── HOJA 1 — Resumen ──────────────────────────────────────────────────
        ws = wb.active
        ws.title = "Resumen"

        title_cell = ws.cell(row=1, column=1, value="Resumen de Gastos")
        title_cell.font = Font(bold=True, size=14, color=NAVY)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

        total = float(context.get('total_gastos') or 0)

        lbl = ws.cell(row=3, column=4, value="Sumatoria total de gastos filtrados:")
        lbl.font = Font(bold=True)
        lbl.alignment = Alignment(horizontal="right")
        tc = ws.cell(row=3, column=5, value=total)
        tc.number_format = MONEY
        tc.font = Font(bold=True, color=WHITE)
        tc.alignment = Alignment(horizontal="right")
        tc.fill = PatternFill("solid", fgColor=NAVY)

        headers = ["Fecha", "Sucursal", "Cuenta", "Categoría", "Total"]
        hdr_row = 5
        for ci, h in enumerate(headers, 1):
            nav_hdr(ws.cell(row=hdr_row, column=ci, value=h))

        ordered = sorted(
            balances,
            key=lambda b: (
                str(b.get('id_sucursal__nombre') or ''),
                str(b.get('id_cuenta_banco__numero_cuenta') or ''),
            ),
        )

        r = hdr_row + 1
        last_sucursal = None
        last_cuenta = None
        suc_total = 0.0
        cta_total = 0.0
        grand_total = 0.0

        def write_detail_row(r, fecha, sucursal, cuenta, categoria, total):
            ws.cell(row=r, column=1, value=fecha)
            ws.cell(row=r, column=2, value=sucursal)
            ws.cell(row=r, column=3, value=cuenta)
            ws.cell(row=r, column=4, value=categoria)
            money_cell(ws.cell(row=r, column=5, value=total))
            return r + 1

        def write_subtotal_row(r, label_col, label, total):
            cell = ws.cell(row=r, column=label_col, value=label)
            cell.font = Font(bold=True)
            money_cell(ws.cell(row=r, column=5, value=total), bold=True)
            fill_row(ws, r, 5, LIGHT)
            return r + 1

        for balance in ordered:
            sucursal = balance.get('id_sucursal__nombre') or ''
            cuenta = balance.get('id_cuenta_banco__numero_cuenta') or ''
            total_b = float(balance.get('total_gastos') or 0)

            if last_sucursal is not None and sucursal != last_sucursal:
                if last_cuenta:
                    r = write_subtotal_row(r, 3, f"{last_cuenta} - SUBTOTAL", cta_total)
                r = write_subtotal_row(r, 2, f"{last_sucursal} - SUBTOTAL", suc_total)
                grand_total += suc_total
                suc_total = 0.0
                cta_total = 0.0
                last_cuenta = None

            if last_cuenta is not None and cuenta != last_cuenta and sucursal == last_sucursal:
                r = write_subtotal_row(r, 3, f"{last_cuenta} - SUBTOTAL", cta_total)
                cta_total = 0.0

            r = write_detail_row(
                r,
                fecha_text(balance_fecha(balance)),
                sucursal,
                cuenta,
                balance.get('id_cat_gastos__nombre') or '',
                total_b,
            )
            cta_total += total_b
            suc_total += total_b
            last_sucursal = sucursal
            last_cuenta = cuenta

        if last_cuenta:
            r = write_subtotal_row(r, 3, f"{last_cuenta} - SUBTOTAL", cta_total)
        if last_sucursal is not None:
            r = write_subtotal_row(r, 2, f"{last_sucursal} - SUBTOTAL", suc_total)
            grand_total += suc_total

        # TOTAL GENERAL
        ws.cell(row=r, column=2, value="TOTAL GENERAL")
        money_cell(ws.cell(row=r, column=5, value=grand_total), bold=True)
        fill_row(ws, r, 5, NAVY, font=Font(bold=True, color=WHITE))
        r += 2

        # ── Gastos por Categoría ──
        sec = ws.cell(row=r, column=1, value="Gastos por Categoría")
        sec.font = Font(bold=True, color=WHITE, size=11)
        sec.fill = PatternFill("solid", fgColor=NAVY)
        sec.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        r += 1

        for ci, h in enumerate(["#", "Categoría", "Total", "%"], 1):
            nav_hdr(ws.cell(row=r, column=ci, value=h))
        r += 1

        cat_total = sum(item['total'] for item in por_categoria) or 0
        for idx, item in enumerate(por_categoria, 1):
            ws.cell(row=r, column=1, value=idx)
            ws.cell(row=r, column=2, value=item['categoria'])
            money_cell(ws.cell(row=r, column=3, value=item['total']))
            pct = (item['total'] / cat_total * 100) if cat_total else 0.0
            pc = ws.cell(row=r, column=4, value=round(pct, 1) / 100)
            pc.number_format = '0.0%'
            pc.alignment = Alignment(horizontal="right")
            r += 1

        ws.cell(row=r, column=2, value="TOTAL").font = Font(bold=True)
        money_cell(ws.cell(row=r, column=3, value=cat_total), bold=True)
        pc = ws.cell(row=r, column=4, value=1.0)
        pc.number_format = '0.0%'
        pc.alignment = Alignment(horizontal="right")
        pc.font = Font(bold=True)
        fill_row(ws, r, 4, LIGHT)

        for ci, w in enumerate([12, 22, 22, 24, 16], 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = "A6"

        # ── HOJA 2 — Por Sucursal (solo cuando el filtro es todas las sucursales) ──
        if por_sucursal is not None:
            ws2 = wb.create_sheet("Por Sucursal")
            sucursales = por_sucursal['sucursales']
            categorias = por_sucursal['categorias']
            matrix = por_sucursal['matrix']

            total_col = 2 + len(sucursales)

            title2 = ws2.cell(row=1, column=1, value="Gastos por Categoría por Sucursal")
            title2.font = Font(bold=True, size=14, color=NAVY)
            ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_col)

            nav_hdr(ws2.cell(row=2, column=1, value="Categoría"))
            for ci, suc in enumerate(sucursales, 2):
                nav_hdr(ws2.cell(row=2, column=ci, value=suc))
            nav_hdr(ws2.cell(row=2, column=total_col, value="Total"))

            col_totals = defaultdict(float)
            rr = 3
            for categoria in categorias:
                ws2.cell(row=rr, column=1, value=categoria).font = Font(bold=True)
                row_total = 0.0
                for ci, suc in enumerate(sucursales, 2):
                    val = matrix.get((suc, categoria), 0.0)
                    money_cell(ws2.cell(row=rr, column=ci, value=val))
                    row_total += val
                    col_totals[suc] += val
                money_cell(ws2.cell(row=rr, column=total_col, value=row_total), bold=True)
                fill_row(ws2, rr, total_col, LIGHT)
                rr += 1

            ws2.cell(row=rr, column=1, value="Total")
            grand = 0.0
            for ci, suc in enumerate(sucursales, 2):
                val = col_totals.get(suc, 0.0)
                money_cell(ws2.cell(row=rr, column=ci, value=val), bold=True)
                grand += val
            money_cell(ws2.cell(row=rr, column=total_col, value=grand), bold=True)
            fill_row(ws2, rr, total_col, NAVY, font=Font(bold=True, color=WHITE))

            ws2.column_dimensions['A'].width = 26
            for ci in range(2, total_col + 1):
                ws2.column_dimensions[get_column_letter(ci)].width = 16
            ws2.freeze_panes = "B3"

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"gastos-resumen-{timezone.now().strftime('%Y-%m-%d')}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

class ComprasResource(resources.ModelResource):
    
    productor = fields.Field(
        column_name='productor',
        attribute='productor',
        widget=ForeignKeyWidget(Productor, field='id'))
    
    producto = fields.Field(
        column_name='producto',
        attribute='producto',
        widget=ForeignKeyWidget(Producto, field='nombre'))
    
    cuenta = fields.Field(
        column_name='cuenta',
        attribute='cuenta',
        widget=ForeignKeyWidget(Cuenta, field='numero_cuenta'))
    
    precio_unitario = fields.Field(
        column_name='precio_unitario',
        attribute='precio_unitario',
        widget=MoneyWidget())
    
    monto_total = fields.Field(
        column_name='monto_total',
        attribute='monto_total',
        widget=MoneyWidget())
    
    class Meta:
        model = Compra
        fields = ('id', 'fecha_compra', 'productor', 'producto', 'cantidad', 'precio_unitario', 'monto_total', 'fecha_registro', 'cuenta', 'tipo_pago')
        import_id_fields = ('id',)
        
    def dehydrate_productor(self, compra):
        return compra.productor.nombre_completo
    
    def dehydrate_producto(self, compra):
        return compra.producto.nombre
    
    def dehydrate_cuenta(self, compra):
        return compra.cuenta.numero_cuenta if compra.cuenta else ""
    
@admin.register(Compra)
class ComprasAdmin(ImportExportModelAdmin, ModelAdmin):
        resource_class = ComprasResource
        import_form_class = ImportForm
        export_form_class = ExportForm
        list_display = ('id', 'fecha_compra','fecha_registro', 'productor', 'producto', 'cantidad', 'precio_unitario', 'monto_total', 'cuenta', 'tipo_pago')
        search_fields = ('tipo_pago', 'productor__nombre', 'producto__nombre', 'cuenta__numero_cuenta')
        list_filter = ('fecha_compra', 'productor', 'producto', 'monto_total')
        list_per_page = 20
        fieldsets = (
            ('Datos del Registro', {
                'fields': ('fecha_compra', 'productor', 'producto', 'cantidad', 'precio_unitario', 'monto_total', 'cuenta', 'tipo_pago')
            }),
        )
        
        class Media:
            js = (
                'js/compra_calculator.js',
            )
 
class SaldoMensualResource(resources.ModelResource):
    cuenta = fields.Field(
            column_name='cuenta',
            attribute='cuenta',
            widget=ForeignKeyWidget(Cuenta, field='numero_cuenta'))
    
    saldo_inicial = fields.Field(
        column_name='saldo_inicial',
        attribute='saldo_inicial',
        widget=MoneyWidget())
    
    saldo_final = fields.Field(
        column_name='saldo_final',
        attribute='saldo_final',
        widget=MoneyWidget())
    
    class Meta:
        model = SaldoMensual
        fields = ('id', 'cuenta', 'año', 'mes', 'saldo_inicial', 'saldo_final', 'fecha_registro', 'ultima_modificacion')
        import_id_fields = ('id',)
        
    def dehydrate_cuenta(self, saldo):
        return saldo.cuenta.numero_cuenta
          
@admin.register(SaldoMensual)
class SaldoMensualAdmin(ImportExportModelAdmin, ModelAdmin):
    resource_class = SaldoMensualResource
    import_form_class = ImportForm
    export_form_class = ExportForm
    list_display = ('cuenta', 'año', 'mes', 'saldo_inicial', 'saldo_final', 'fecha_registro', 'ultima_modificacion')
    search_fields = ('cuenta__numero_cuenta', 'año', 'mes')
    list_filter = ('cuenta', 'año', 'mes')
    list_per_page = 12
    fieldsets = (
        ('Datos del Registro', {
            'fields': ('cuenta', 'año', 'mes', 'saldo_inicial', 'saldo_final')
        }),
    )
@admin.register(ComprobanteGasto)
class ComprobanteGastoAdmin(ModelAdmin):
    """Consulta y auditoria de comprobantes cargados por el flujo OCR."""

    list_display = (
        'id', 'vista_previa', 'nombre_original', 'estado_visual', 'confianza',
        'gasto', 'creado_por', 'creado_en',
    )
    list_filter = ('estado', 'content_type', 'creado_en')
    search_fields = ('id', 'nombre_original', 'sha256', 'texto_ocr', 'creado_por__username')
    list_select_related = ('gasto', 'creado_por')
    readonly_fields = (
        'storage_key', 'archivo_enlace', 'vista_previa_grande', 'nombre_original',
        'content_type', 'tamano_bytes', 'sha256', 'estado', 'datos_extraidos',
        'texto_ocr', 'confianza', 'error_procesamiento', 'creado_por', 'gasto',
        'creado_en', 'procesado_en',
    )
    ordering = ('-creado_en',)
    list_per_page = 25
    date_hierarchy = 'creado_en'
    fieldsets = (
        ('Archivo', {'fields': ('vista_previa_grande', 'archivo_enlace', 'nombre_original', 'content_type', 'tamano_bytes', 'sha256')}),
        ('Procesamiento OCR', {'fields': ('estado', 'confianza', 'datos_extraidos', 'texto_ocr', 'error_procesamiento')}),
        ('Relacion', {'fields': ('gasto', 'creado_por')}),
        ('Metadatos', {'fields': ('storage_key', 'creado_en', 'procesado_en')}),
    )

    @admin.display(description='Archivo')
    def vista_previa(self, obj):
        if not obj.archivo:
            return '-'
        try:
            url = obj.archivo.url
        except (ValueError, OSError):
            return '-'
        if obj.content_type.startswith('image/'):
            return format_html(
                '<a href="{}" target="_blank" title="Abrir imagen"><img src="{}" alt="" style="width:64px;height:48px;object-fit:cover;border-radius:6px;border:1px solid #cbd5d1" /></a>',
                url, url,
            )
        return format_html('<a href="{}" target="_blank">Abrir archivo</a>', url)

    @admin.display(description='Estado', ordering='estado')
    def estado_visual(self, obj):
        colors = {
            ComprobanteGasto.Estado.PENDIENTE: '#8a6d1d',
            ComprobanteGasto.Estado.PROCESANDO: '#1d6fa5',
            ComprobanteGasto.Estado.REVISION: '#16834f',
            ComprobanteGasto.Estado.ERROR: '#b83232',
            ComprobanteGasto.Estado.REGISTRADO: '#246b48',
        }
        return format_html('<strong style="color:{}">{}</strong>', colors.get(obj.estado, '#374151'), obj.get_estado_display())

    @admin.display(description='Archivo vinculado')
    def archivo_enlace(self, obj):
        if not obj.archivo:
            return '-'
        try:
            return format_html('<a href="{}" target="_blank">{} <span aria-hidden="true">↗</span></a>', obj.archivo.url, obj.nombre_original)
        except (ValueError, OSError):
            return 'Archivo no disponible'

    @admin.display(description='Vista previa')
    def vista_previa_grande(self, obj):
        if not obj.archivo or not obj.content_type.startswith('image/'):
            return 'La vista previa esta disponible para imagenes.'
        try:
            return format_html('<a href="{}" target="_blank"><img src="{}" alt="{}" style="max-width:560px;max-height:420px;border-radius:8px;border:1px solid #d1d5db" /></a>', obj.archivo.url, obj.archivo.url, obj.nombre_original)
        except (ValueError, OSError):
            return 'Imagen no disponible'

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return request.user.has_perm('gastos.view_comprobantegasto')
